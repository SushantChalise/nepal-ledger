"""Department of Customs Foreign Trade Statistics (FTS) parser — deterministic.

Source id: ``customs-monthly-trade``. The Department of Customs
(customs.gov.np) publishes, per fiscal year, a series of machine-readable XLSX
workbooks of foreign-trade statistics compiled from the ASYCUDA World system —
a cumulative-to-date snapshot for each month plus a full-year annual file
(Foreign Trade Statistics portal → ``/category/fts-<bs-fy>/``). Each workbook
shares one fixed 10-sheet layout (see ``README.md`` for the acquisition record).

Output (ADR-0015 dimensional facts → ``dne_facts``):
    One base measure sliced by exactly one dimension. We emit seven fact streams
    from the four single-dimension sheets plus the two commodity×partner
    cross-tabs (the latter via a COMPOSITE dimension — see below):

      sheet 5 (Imports by Commodity)  → ``customs-merchandise-imports`` × commodity (HS)
      sheet 7 (Exports by Commodity)  → ``customs-merchandise-exports`` × commodity (HS)
      sheet 3 (Trade Balance Country) → imports & exports × country
      sheet 9 (Customs-wise Trade)    → imports & exports × customs_office
      sheet 4 (Imports by Commodity & Partner) → imports × customs-import-source
      sheet 6 (Exports by Commodity & Partner) → exports × customs-export-destination

    ``dimension_value`` for a commodity is the **HS code** (faithfully preserved,
    6- or 8-digit); ``dimension_label`` is the commodity description. For a
    country/customs office the value is the kebab slug of the name and the label
    is the raw name.

Composite dimension (ADR-0018 — commodity×partner cross-tab):
    Sheets 4 & 6 are long-form cross-tabs: each row is one (HS code, description,
    partner country, …, value) tuple — a fact sliced by TWO dimensions. Rather
    than a new table/migration we encode the pair into the existing
    one-dimension ``dne_facts`` contract via a composite ``dimension_value``:

      dimension_kind  = ``customs-import-source``  (sheet 4, imports by source)
                        ``customs-export-destination`` (sheet 6, exports by dest.)
      dimension_value = ``<hs-code>__<country-slug>`` — the 8-digit HS code and
                        the kebab country slug joined by a ``__`` separator. Both
                        parts are separator-stable: HS codes are pure digits and
                        country slugs contain only ``[a-z0-9-]`` (verified: no
                        slug collision, none contains ``__``), so the split is
                        unambiguous.
      dimension_label = ``<commodity description> → <country>`` (human readable).
      base_indicator_slug = the SAME single-dimension measure slug
                        (``customs-merchandise-imports`` / ``-exports``), so the
                        cross-tab is a strict disaggregation of the commodity
                        totals — it reconciles exactly (ADR-0011, verified below).

    The trailing grand-total row (blank HS + "Total" description) is the only
    aggregate and is excluded — it is already the single-dimension headline total.

Deliberately DEFERRED (kept honest; not fabricated):
    - The ``Imports_Revenue`` column (customs duty collected) and the derived
      ``Trade_Balance`` column — separate measures, not part of the trade-volume
      brief; can be promoted later with their own base slugs.
    - sheet 2 (Trade Balance by HS Chapter — redundant with the commodity sheets),
      sheet 8 (ID value comparison), sheet 1 (headline totals — single series, not
      dimensional).

Unit (ADR-0011 magnitude verification — read the header, don't fuzzy-match):
    Every value sheet states its unit literally as "(figures are in Rs.
    Thousands)" / "(... are in Rs. Thousand)" and the headline table writes
    "Imports (Rs.in `000)". So values are NPR **THOUSAND** — ``npr_thousand`` —
    NOT million and NOT lakh. Sanity check (FY 2081/82 annual): total imports =
    1,804,122,731 thousand = NPR 1.804 TRILLION (~NPR 150bn/month); total exports
    = 277,030,201 thousand = NPR 277 billion (~NPR 23bn/month) — exactly the
    Nepal-scale magnitude expected (ADR-0011 band).

    Cross-tab reconciliation (ADR-0011): summing a commodity's cross-tab partner
    cells reproduces its single-dimension commodity total exactly. Verified FY
    2081/82 annual across ALL 5,264 import + 1,236 export commodities — worst
    relative difference 0.0% (e.g. Diesel HS 27101930 = 128,761,649.231 thousand
    in both). The two grand totals also match the headline imports/exports.

Period dating (ADR-0013):
    The index sheet's row-0 descriptor is self-describing and carries the EXACT
    AD span, e.g.::

        Based on Annual data (Shrawan-Asar) of FY 2081/82 (Mid July 2024 to Mid July 2025)
        Based on First Month (Shrawan) of FY 2081/82 (Mid July 2024 to Mid August 2024)
        Based on First Eleven Months (Shrawan-Jestha) of FY 2081/82 (Mid July 2024 to Mid June 2025)

    We parse the BS fiscal year, the scope, and the literal "Mid <Month> <Year>"
    AD span directly (no mid-month approximation needed — the source states it):
      - "Annual data (...)"        → ``reporting_period_type='annual'``; period
        label = the BS FY (e.g. "2081/82").
      - "First Month (Shrawan)"    → exactly month 1 → ``'monthly'``; period label
        = "Shrawan <bs-year>".
      - "First N Months (...-End)" → cumulative year-to-date → ``'year_to_date'``;
        period label = the END month ("Jestha <bs-year>"), flagged in
        ``base_indicator_name``-adjacent note as cumulative-from-Shrawan.
    The end BS year is derived from the AD span end year via the +57 fiscal
    offset and the BS-month→AD-month break (Shrawan..Poush in AD lead year,
    Magh..Ashadh in the next). Unparseable descriptor → ``RegexMismatch`` and the
    file is rejected (never a silently mis-dated fact).

Confidence: ``A`` — the source registry sets ``confidenceDefault: 'A'`` for
``customs-monthly-trade`` (transaction-level ASYCUDA customs declarations, the
authoritative trade record). Higher than NRB's compiled ``B``.

ADR: ADR-0003 — no LLM / AI calls; pure file-in → dataclass-out. ADR-0015 —
dimensional fact contract. ADR-0011 — unit/magnitude verified above.

Versioning: bump ``PARSER_VERSION`` on any behaviour change.
    0.1.0 — initial: four single-dimension sheets (5/7/3/9) → 5 fact streams.
    0.2.0 — add the commodity×partner cross-tabs (sheets 4 & 6) via a composite
            ``<hs>__<country-slug>`` dimension (kinds ``customs-import-source`` /
            ``customs-export-destination``); reconciles to the commodity totals.
"""

from __future__ import annotations

import json
import re
import sys
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Final

import openpyxl

from _common.periods import BS_MONTHS, BsMonth, fiscal_year_ad_label, fiscal_year_label
from _common.types import ParserError, ParserStatus, ReportingPeriodType

PARSER_VERSION: Final[str] = "0.3.0"
SOURCE_ID: Final[str] = "customs-monthly-trade"

# Confidence default — see module docstring (source registry = 'A').
_CONFIDENCE: Final[str] = "A"

# Unit verbatim from every value sheet's header "(... in Rs. Thousand[s])" and
# the headline "Rs.in `000" — NPR thousand (ADR-0011). NOT million, NOT lakh.
_UNIT_NPR_THOUSAND: Final[str] = "npr_thousand"

# Base measures (dimension-agnostic). The dimension distinguishes commodity /
# country / customs office; the measure distinguishes the trade direction.
_IMPORTS_SLUG: Final[str] = "customs-merchandise-imports"
_IMPORTS_NAME: Final[str] = "Merchandise imports (customs)"
_EXPORTS_SLUG: Final[str] = "customs-merchandise-exports"
_EXPORTS_NAME: Final[str] = "Merchandise exports (customs)"

# Composite-dimension kinds for the commodity×partner cross-tabs (ADR-0018). The
# dimension encodes BOTH the commodity (HS) and the partner country; the measure
# (imports/exports) is the SAME base slug as the single-dimension facts, so the
# cross-tab is a strict disaggregation that reconciles to the commodity totals.
_DIM_IMPORT_SOURCE: Final[str] = "customs-import-source"  # sheet 4: imports by source
_DIM_EXPORT_DESTINATION: Final[str] = "customs-export-destination"  # sheet 6: by dest.

# Separator joining the two parts of a composite dimension_value
# (``<hs-code>__<country-slug>``). Unambiguous: HS codes are pure digits and
# country slugs are ``[a-z0-9-]`` only (no slug contains ``__``; no collisions).
_COMPOSITE_SEP: Final[str] = "__"

# Sheet names (stable across the monthly + annual editions; verified on FY2081/82).
_SHEET_INDEX: Final[str] = "0_Index and Key notes"
_SHEET_IMPORTS_COMMODITY: Final[str] = "5_Imports_By_Commodity"
_SHEET_EXPORTS_COMMODITY: Final[str] = "7_Exports_By_Commodity"
_SHEET_COUNTRY: Final[str] = "3_Trade_Balance_Country"
_SHEET_CUSTOMS: Final[str] = "9_Customswise_Trade"
_SHEET_IMPORTS_COMMODITY_PARTNER: Final[str] = "4_Imports_By_Commodity_Partner"
_SHEET_EXPORTS_COMMODITY_PARTNER: Final[str] = "6_Exports_By_Commodity_Partner"

# Header is row index 2 (0-based) on every value sheet; data starts at row 3.
_DATA_START_ROW_IDX: Final[int] = 3

# Column index of the "Partner Countries" column in the cross-tab sheets (4 & 6):
# 0=HSCode, 1=Description, 2=Partner, 3=Unit, 4=Quantity, 5=value.
_PARTNER_COL_IDX: Final[int] = 2

# Aggregate/total label tokens that are NOT a dimension member (ADR-0015: totals
# are excluded). Lowercased before comparison. The commodity/country/customs
# sheets end with a single "Total" row whose code/SN cell is blank.
_TOTAL_LABELS: Final[frozenset[str]] = frozenset({"total", "grand total", "sum"})

# A valid HS code is 6 or 8 ASCII digits (Nepal publishes 8-digit national HS
# subheadings; a few legacy rows are 6-digit). Anything else in the code column
# (e.g. the trailing blank "Total" row) is not a commodity.
_HS_CODE_RE: Final = re.compile(r"^\d{6,8}$")

# Index-descriptor parser. Captures: (1) scope phrase, (2) BS FY lead, (3) BS FY
# tail, (4) AD start month, (5) AD start year, (6) AD end month, (7) AD end year.
#   "Based on Annual data (Shrawan-Asar) of FY 2081/82 (Mid July 2024 to Mid July 2025)"
_DESCRIPTOR_RE: Final = re.compile(
    r"Based on\s+(?P<scope>.+?)\s+of\s+FY\s+(?P<fy_lead>\d{4})\s*/\s*(?P<fy_tail>\d{2,4})\s*"
    r"\(\s*Mid\s+(?P<ad_start_mon>[A-Za-z]+)\s+(?P<ad_start_yr>\d{4})\s+to\s+"
    r"Mid\s+(?P<ad_end_mon>[A-Za-z]+)\s+(?P<ad_end_yr>\d{4})\s*\)",
    re.IGNORECASE,
)

# AD Gregorian month name → number, for the "Mid <Month> <Year>" span edges.
_AD_MONTH_NAME_TO_NUM: Final[dict[str, int]] = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

# Customs spells the END BS month inside the scope parenthetical, e.g.
# "(Shrawan-Jestha)" or "(Shrawan-Asar)". We read THAT (the source's own BS
# label) rather than inferring from the AD edge. Map Customs' spellings to the
# canonical ``BsMonth`` vocab (``_common.periods.BsMonth``): note "Asar"/"Asadh"
# = "Ashadh", "Ashoj"/"Asoj" = "Ashwin", "Marg" = "Mangsir", "Baisakh"/"Baishakh".
_BS_MONTH_ALIASES: Final[dict[str, BsMonth]] = {
    "shrawan": "Shrawan",
    "sawan": "Shrawan",
    "saun": "Shrawan",
    "bhadra": "Bhadra",
    "bhadau": "Bhadra",
    "ashwin": "Ashwin",
    "ashoj": "Ashwin",
    "asoj": "Ashwin",
    "aswin": "Ashwin",
    "kartik": "Kartik",
    "kattik": "Kartik",
    "mangsir": "Mangsir",
    "marg": "Mangsir",
    "marga": "Mangsir",
    "poush": "Poush",
    "push": "Poush",
    "pous": "Poush",
    "magh": "Magh",
    "falgun": "Falgun",
    "fagun": "Falgun",
    "phalgun": "Falgun",
    "chait": "Chait",
    "chaitra": "Chait",
    "baisakh": "Baisakh",
    "baishakh": "Baisakh",
    "baisakh.": "Baisakh",
    "jestha": "Jestha",
    "jeth": "Jestha",
    "jyestha": "Jestha",
    "ashadh": "Ashadh",
    "asar": "Ashadh",
    "asadh": "Ashadh",
    "ashar": "Ashadh",
}

# Fiscal-year position of each BS month (Shrawan = month 1 of the fiscal year …
# Ashadh = month 12). Used to derive the cumulative month-count AND the end BS
# year (positions 1–6 fall in the FY lead BS year; 7–12 in lead+1).
_BS_MONTH_FY_POSITION: Final[dict[BsMonth, int]] = {m: i + 1 for i, m in enumerate(BS_MONTHS)}

# Positions 1–6 (Shrawan..Poush) are in the FY's lead BS year; 7–12 (Magh..Ashadh)
# roll into lead+1. (e.g. FY 2081/82: Shrawan 2081 … Poush 2081, Magh 2082 …
# Ashadh 2082.) This boundary is the BS-calendar equivalent of the AD mid-July break.
_BS_FY_YEAR_ROLLOVER_POSITION: Final[int] = 6


# ---------------------------------------------------------------------------
# Dimensional fact contract (ADR-0015) — mirrors the DNE / Yellow Book parsers
# field-for-field so the cloned ``ingest-customs-trade.ts`` CLI reads the same
# ``dimensional_rows`` JSON. Intentionally NOT in _common/types.py (DNE-local).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class DimensionalRowDraft:
    """One dimensional fact: a base measure sliced by exactly one dimension."""

    base_indicator_slug: str
    base_indicator_name: str
    dimension_kind: str
    dimension_value: str
    dimension_label: str
    value: float
    unit: str
    reporting_period_type: ReportingPeriodType
    reporting_period_bs: str
    reporting_period_ad_start: datetime
    reporting_period_ad_end: datetime
    fiscal_year_bs: str
    fiscal_year_ad_label: str
    confidence_grade: str

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "base_indicator_slug": self.base_indicator_slug,
            "base_indicator_name": self.base_indicator_name,
            "dimension_kind": self.dimension_kind,
            "dimension_value": self.dimension_value,
            "dimension_label": self.dimension_label,
            "value": self.value,
            "unit": self.unit,
            "reporting_period_type": self.reporting_period_type,
            "reporting_period_bs": self.reporting_period_bs,
            "reporting_period_ad_start": self.reporting_period_ad_start.isoformat(),
            "reporting_period_ad_end": self.reporting_period_ad_end.isoformat(),
            "fiscal_year_bs": self.fiscal_year_bs,
            "fiscal_year_ad_label": self.fiscal_year_ad_label,
            "confidence_grade": self.confidence_grade,
        }


@dataclass(frozen=True)
class CustomsResult:
    """Customs FTS parser result carrying dimensional output only.

    Mirrors the DNE CLI's expected JSON shape: a ``dimensional_rows`` array plus
    ``status`` / ``errors``. Customs FTS emits no single-series ``staging_rows``.
    """

    status: ParserStatus
    parser_version: str
    dimensional_rows: list[DimensionalRowDraft]
    errors: list[ParserError]

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "parser_version": self.parser_version,
            "dimensional_rows": [r.to_json_dict() for r in self.dimensional_rows],
            "errors": [e.to_json_dict() for e in self.errors],
        }


# ---------------------------------------------------------------------------
# Period resolution — parsed from the index descriptor (self-describing).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _Period:
    """Resolved reporting period for a whole FTS workbook (all sheets share it)."""

    reporting_period_type: ReportingPeriodType
    reporting_period_bs: str
    reporting_period_ad_start: datetime
    reporting_period_ad_end: datetime
    fiscal_year_bs: str
    fiscal_year_ad_label: str
    # Appended to base_indicator_name for cumulative files so the YTD semantics
    # travel with the fact (e.g. "(cumulative Shrawan–Jestha)").
    cumulative_note: str | None


def _norm(cell: object) -> str:
    """Stringify a cell and collapse internal whitespace/newlines to a space."""
    if cell is None:
        return ""
    return " ".join(str(cell).split())


def _ad_month_num(name: str) -> int | None:
    return _AD_MONTH_NAME_TO_NUM.get(name.strip().lower())


def _bs_month_from_alias(token: str) -> BsMonth | None:
    """Normalise a Customs-spelled BS month token → canonical ``BsMonth``."""
    return _BS_MONTH_ALIASES.get(token.strip().lower().rstrip("."))


def _scope_end_bs_month(scope: str) -> BsMonth | None:
    """Read the END BS month from the scope parenthetical (the source's own label).

    "First Month (Shrawan)"               → Shrawan (single month → start=end)
    "First Eleven Months (Shrawan-Jestha)"→ Jestha
    "Annual data (Shrawan-Asar)"          → Ashadh ("Asar" alias)
    The parenthetical is "(Start)" or "(Start-End)"; we take the LAST token. The
    AD span is only month edges, so we never infer the BS month from it. Returns
    None if no recognised BS month is present.
    """
    m = re.search(r"\(([^)]*)\)", scope)
    if not m:
        return None
    inner = m.group(1)
    # Split on hyphen/en-dash/"to"; the end month is the last recognised token.
    tokens = re.split(r"\s*(?:-|–|to)\s*", inner)
    for tok in reversed(tokens):
        bs = _bs_month_from_alias(tok)
        if bs is not None:
            return bs
    return None


def _end_bs_year(fy_lead: int, end_bs_month: BsMonth) -> int:
    """BS year of ``end_bs_month`` within FY whose lead BS year is ``fy_lead``.

    Months at fiscal positions 1–6 (Shrawan..Poush) stay in ``fy_lead``; positions
    7–12 (Magh..Ashadh) roll into ``fy_lead + 1`` (e.g. Jestha of FY 2081/82 is
    2082).
    """
    pos = _BS_MONTH_FY_POSITION[end_bs_month]
    return fy_lead if pos <= _BS_FY_YEAR_ROLLOVER_POSITION else fy_lead + 1


@dataclass(frozen=True)
class _DescriptorHeader:
    """The fiscal-year + AD-span + scope primitives parsed from a descriptor."""

    fy_lead: int
    ad_start: datetime
    ad_end: datetime
    scope: str


def _parse_descriptor_header(
    descriptor: str,
) -> tuple[_DescriptorHeader | None, ParserError | None]:
    """Match + validate the descriptor's FY, AD span, and scope (or typed error)."""
    m = _DESCRIPTOR_RE.search(descriptor)
    if not m:
        return None, ParserError(
            error_class="RegexMismatch",
            error_detail=(
                "index descriptor did not match the expected "
                "'Based on <scope> of FY <YYYY/YY> (Mid <Mon> <Yr> to Mid <Mon> <Yr>)' form"
            ),
            source_excerpt=descriptor[:200],
        )
    fy_lead = int(m.group("fy_lead"))
    fy_tail = int(m.group("fy_tail")) % 100
    if fy_tail != (fy_lead + 1) % 100:
        return None, ParserError(
            error_class="RegexMismatch",
            error_detail=f"FY label tail {fy_tail:02d} != (lead+1) for lead {fy_lead}",
            source_excerpt=descriptor[:200],
        )
    start_mon = _ad_month_num(m.group("ad_start_mon"))
    end_mon = _ad_month_num(m.group("ad_end_mon"))
    if start_mon is None or end_mon is None:
        return None, ParserError(
            error_class="RegexMismatch",
            error_detail="AD span month name not recognised",
            source_excerpt=descriptor[:200],
        )
    return (
        _DescriptorHeader(
            fy_lead=fy_lead,
            ad_start=datetime(int(m.group("ad_start_yr")), start_mon, 15, tzinfo=UTC),
            ad_end=datetime(int(m.group("ad_end_yr")), end_mon, 15, tzinfo=UTC),
            scope=m.group("scope").strip(),
        ),
        None,
    )


def parse_period_descriptor(descriptor: str) -> tuple[_Period | None, ParserError | None]:
    """Resolve the index-sheet row-0 descriptor → a ``_Period`` (or typed error).

    The AD span is taken verbatim from the descriptor's "Mid <Month> <Year>" month
    edges. The BS period label is read from the scope parenthetical (the source's
    own BS month names), NOT inferred from the AD edge — the AD bounds are
    mid-month boundaries, so "to Mid June" is the CLOSE of Jestha, and inferring a
    month from it would mislabel the period. Mapping:

      - "Annual data (...)"     → ``annual``; label = the BS FY ("2081/82").
      - "First Month (Shrawan)" → exactly month 1 → ``monthly``; label = "Shrawan <yr>".
      - "First N Months (..-End)"→ cumulative → ``year_to_date``; label = the END
        month ("Jestha <yr>"), month-count = the end month's fiscal position.

    Returns (None, error) when the descriptor or its month names do not match.
    """
    header, header_err = _parse_descriptor_header(descriptor)
    if header is None:
        return None, header_err
    fy_lead, ad_start, ad_end, scope = (
        header.fy_lead,
        header.ad_start,
        header.ad_end,
        header.scope,
    )
    fy_bs = fiscal_year_label(fy_lead)
    fy_ad = fiscal_year_ad_label(fy_lead)

    if "annual" in scope.lower():
        return (
            _Period(
                reporting_period_type="annual",
                reporting_period_bs=fy_bs,
                reporting_period_ad_start=ad_start,
                reporting_period_ad_end=ad_end,
                fiscal_year_bs=fy_bs,
                fiscal_year_ad_label=fy_ad,
                cumulative_note=None,
            ),
            None,
        )

    end_bs_month = _scope_end_bs_month(scope)
    if end_bs_month is None:
        return None, ParserError(
            error_class="RegexMismatch",
            error_detail="could not read the end BS month from the scope parenthetical",
            source_excerpt=scope[:120],
        )
    end_bs_year = _end_bs_year(fy_lead, end_bs_month)
    months = _BS_MONTH_FY_POSITION[end_bs_month]  # fiscal position == # months elapsed
    period_label = f"{end_bs_month} {end_bs_year}"

    if months == 1:
        # Exactly month 1 (Shrawan): a genuine single month, not cumulative.
        return (
            _Period(
                reporting_period_type="monthly",
                reporting_period_bs=period_label,
                reporting_period_ad_start=ad_start,
                reporting_period_ad_end=ad_end,
                fiscal_year_bs=fy_bs,
                fiscal_year_ad_label=fy_ad,
                cumulative_note=None,
            ),
            None,
        )

    return (
        _Period(
            reporting_period_type="year_to_date",
            reporting_period_bs=period_label,
            reporting_period_ad_start=ad_start,
            reporting_period_ad_end=ad_end,
            fiscal_year_bs=fy_bs,
            fiscal_year_ad_label=fy_ad,
            cumulative_note=f"cumulative Shrawan–{end_bs_month} ({months} months) of FY {fy_bs}",
        ),
        None,
    )


# ---------------------------------------------------------------------------
# Value parsing helpers.
# ---------------------------------------------------------------------------


def _parse_value(cell: object) -> float | None:
    """Parse a numeric trade-value cell → float; None for blank/dash/non-numeric.

    A genuine source ``0`` is preserved as ``0.0`` (never fabricated, never
    dropped). Commas are tolerated though the source ships bare floats.
    """
    if isinstance(cell, bool):  # bool is an int subclass — never a value here
        return None
    if isinstance(cell, int | float):
        v = float(cell)
        return None if v != v else v  # drop NaN  # noqa: PLR0124
    s = str(cell).strip()
    if s in ("", "-", "--", "–", "—", "N/A", "n/a", "NA", "...", "."):
        return None
    try:
        v = float(s.replace(",", ""))
    except (TypeError, ValueError):
        return None
    return None if v != v else v  # noqa: PLR0124


def _slugify_name(label: str) -> str:
    """Kebab slug of a country / customs-office name for ``dimension_value``.

    Lowercases ASCII, folds punctuation/whitespace runs to single hyphens. Two
    distinct names never collapse to one slug (the raw label is kept separately as
    ``dimension_label``). Returns the trimmed lowercase fallback if stripping would
    empty the slug, so it stays non-empty and traceable.
    """
    s = re.sub(r"[^a-z0-9]+", " ", label.lower())
    slug = re.sub(r"\s+", "-", s.strip()).strip("-")
    return slug or label.strip().lower()


def _base_name(base_name: str, period: _Period) -> str:
    """Append the cumulative note (if any) to the base measure name."""
    if period.cumulative_note is None:
        return base_name
    return f"{base_name} [{period.cumulative_note}]"


def _is_total_row(label: str, code_or_sn: str) -> bool:
    """True for the trailing aggregate row (blank code/SN + a total-ish label)."""
    return code_or_sn.strip() == "" and label.strip().lower() in _TOTAL_LABELS


def _make_row(
    base_slug: str,
    base_name: str,
    dimension_kind: str,
    dimension_value: str,
    dimension_label: str,
    value: float,
    period: _Period,
) -> DimensionalRowDraft:
    """Build one DimensionalRowDraft for a base measure + dimension + value."""
    return DimensionalRowDraft(
        base_indicator_slug=base_slug,
        base_indicator_name=_base_name(base_name, period),
        dimension_kind=dimension_kind,
        dimension_value=dimension_value,
        dimension_label=dimension_label,
        value=value,
        unit=_UNIT_NPR_THOUSAND,
        reporting_period_type=period.reporting_period_type,
        reporting_period_bs=period.reporting_period_bs,
        reporting_period_ad_start=period.reporting_period_ad_start,
        reporting_period_ad_end=period.reporting_period_ad_end,
        fiscal_year_bs=period.fiscal_year_bs,
        fiscal_year_ad_label=period.fiscal_year_ad_label,
        confidence_grade=_CONFIDENCE,
    )


# ---------------------------------------------------------------------------
# Deterministic core — one function per sheet kind, exercised by tests against
# synthesized rows. Each takes the raw rows-as-tuples for the sheet.
# ---------------------------------------------------------------------------


def extract_commodity_rows(
    sheet_rows: Sequence[tuple[object, ...]],
    base_slug: str,
    base_name: str,
    period: _Period,
    value_col: int,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Commodity sheet → one fact per HS-code row.

    Columns: 0=HSCode, 1=Description, 2=Unit, 3=Quantity, ``value_col``=trade
    value. The trailing "Total" row (blank HS code) is excluded as an aggregate.
    A row whose code is non-blank but not a valid HS code, or whose value does not
    parse, surfaces a typed error (visible, never silent) and is skipped.
    """
    rows: list[DimensionalRowDraft] = []
    errors: list[ParserError] = []
    seen: set[str] = set()
    for r_idx in range(_DATA_START_ROW_IDX, len(sheet_rows)):
        row = sheet_rows[r_idx]
        if value_col >= len(row):
            continue
        code = _norm(row[0]) if row else ""
        label = _norm(row[1]) if len(row) > 1 else ""
        if _is_total_row(label, code):
            continue
        if not code and not label:
            continue
        if not _HS_CODE_RE.fullmatch(code):
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"row {r_idx}: code {code!r} is not a 6/8-digit HS code",
                    source_excerpt=f"{code} | {label}",
                )
            )
            continue
        if not label:
            # Older FTS editions (e.g. FY2076/77–2079/80) occasionally carry a
            # valid HS-coded row with a BLANK description cell. The value is real;
            # fall back to the HS code as the label so we never emit an empty
            # dimension_label (the HS code is itself a stable identifier). Never
            # drop the row — that would silently lose a real trade fact.
            label = code
        value = _parse_value(row[value_col])
        if value is None:
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"HS {code}: trade-value cell did not parse ({row[value_col]!r})",
                    source_excerpt=f"{code} | {label}",
                )
            )
            continue
        if code in seen:
            continue  # duplicate HS code in a single sheet — keep the first only
        seen.add(code)
        rows.append(_make_row(base_slug, base_name, "commodity", code, label, value, period))
    return rows, errors


def extract_commodity_partner_rows(
    sheet_rows: Sequence[tuple[object, ...]],
    base_slug: str,
    base_name: str,
    dimension_kind: str,
    period: _Period,
    value_col: int,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Commodity×partner cross-tab (long form) → one fact per (HS code, partner).

    Sheets 4 (imports) & 6 (exports) repeat the commodity sheet's geometry with a
    partner column inserted: 0=HSCode, 1=Description, 2=Partner Countries, 3=Unit,
    4=Quantity, ``value_col``=trade value. Each row is ONE (commodity, partner)
    cell, so it is encoded as a COMPOSITE dimension (ADR-0018):
    ``dimension_value`` = ``<hs-code>__<country-slug>``,
    ``dimension_label`` = ``<description> → <country>``. The base measure slug is
    the SAME as the single-dimension commodity facts, so summing a commodity's
    partner cells reconciles to its commodity total (ADR-0011).

    Excludes only the trailing grand-total row (blank HS + "Total" description) —
    that is already the single-dimension headline total. A row whose code is
    non-blank but not a valid HS code, whose partner is blank, or whose value does
    not parse surfaces a typed error (visible, never silent — Rule 6) and is
    skipped; a genuine source ``0`` is preserved.
    """
    rows: list[DimensionalRowDraft] = []
    errors: list[ParserError] = []
    seen: set[str] = set()
    for r_idx in range(_DATA_START_ROW_IDX, len(sheet_rows)):
        row = sheet_rows[r_idx]
        if value_col >= len(row):
            continue
        code = _norm(row[0]) if row else ""
        label = _norm(row[1]) if len(row) > 1 else ""
        partner = _norm(row[_PARTNER_COL_IDX]) if len(row) > _PARTNER_COL_IDX else ""
        # Grand-total row: blank HS + a total-ish description (the only aggregate).
        if _is_total_row(label, code):
            continue
        if not code and not label and not partner:
            continue
        if not _HS_CODE_RE.fullmatch(code):
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"row {r_idx}: code {code!r} is not a 6/8-digit HS code",
                    source_excerpt=f"{code} | {label} | {partner}",
                )
            )
            continue
        if not partner:
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"HS {code}: blank partner country in cross-tab row {r_idx}",
                    source_excerpt=f"{code} | {label}",
                )
            )
            continue
        value = _parse_value(row[value_col])
        if value is None:
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"HS {code} × {partner!r}: trade-value cell did not parse "
                    f"({row[value_col]!r})",
                    source_excerpt=f"{code} | {label} | {partner}",
                )
            )
            continue
        composite_value = f"{code}{_COMPOSITE_SEP}{_slugify_name(partner)}"
        if composite_value in seen:
            continue  # duplicate (HS, partner) within a sheet — keep the first only
        seen.add(composite_value)
        composite_label = f"{label} → {partner}"  # "<description> → <country>"
        rows.append(
            _make_row(
                base_slug,
                base_name,
                dimension_kind,
                composite_value,
                composite_label,
                value,
                period,
            )
        )
    return rows, errors


def extract_country_rows(
    sheet_rows: Sequence[tuple[object, ...]],
    period: _Period,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Trade-Balance-by-Country sheet → imports & exports facts per country.

    Columns: 0=SN, 1=Partner Countries, 2=Imports_Value, 3=Exports_Value,
    4=Trade_Balance (Trade_Balance is a derived measure — DEFERRED, not emitted).
    The trailing "Total" row (blank SN) is excluded. Emits a separate fact for
    each direction whose value parses (a genuine 0 is kept).
    """
    return _extract_two_direction_rows(
        sheet_rows, period, dimension_kind="country", imports_col=2, exports_col=3
    )


def extract_customs_rows(
    sheet_rows: Sequence[tuple[object, ...]],
    period: _Period,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Customs-office-wise sheet → imports & exports facts per customs office.

    Columns: 0=SN, 1=Customs, 2=Imports_Value, 3=Import_Share, 4=Exports_Value,
    5=Export_Share (the share columns are derived — DEFERRED). The trailing
    "Total" row (blank SN) is excluded.
    """
    return _extract_two_direction_rows(
        sheet_rows, period, dimension_kind="customs_office", imports_col=2, exports_col=4
    )


def _extract_two_direction_rows(
    sheet_rows: Sequence[tuple[object, ...]],
    period: _Period,
    dimension_kind: str,
    imports_col: int,
    exports_col: int,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Shared core for the country + customs sheets (SN, name, imports, exports).

    For each named row (non-blank SN) that is not the trailing Total, emit an
    imports fact when its imports value parses and an exports fact when its
    exports value parses. A row whose BOTH values fail to parse surfaces one
    ``ValueUnparseable`` (data loss is visible, never silent — Rule 6).
    """
    rows: list[DimensionalRowDraft] = []
    errors: list[ParserError] = []
    seen: set[tuple[str, str]] = set()
    max_col = max(imports_col, exports_col)
    for r_idx in range(_DATA_START_ROW_IDX, len(sheet_rows)):
        row = sheet_rows[r_idx]
        if max_col >= len(row):
            continue
        sn = _norm(row[0]) if row else ""
        name = _norm(row[1]) if len(row) > 1 else ""
        if _is_total_row(name, sn):
            continue
        if not name:
            continue
        dim_value = _slugify_name(name)
        imp = _parse_value(row[imports_col])
        exp = _parse_value(row[exports_col])
        if imp is None and exp is None:
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=f"{dimension_kind} {name!r}: neither imports nor exports parsed",
                    source_excerpt=f"{sn} | {name}",
                )
            )
            continue
        if imp is not None and (_IMPORTS_SLUG, dim_value) not in seen:
            seen.add((_IMPORTS_SLUG, dim_value))
            rows.append(
                _make_row(
                    _IMPORTS_SLUG, _IMPORTS_NAME, dimension_kind, dim_value, name, imp, period
                )
            )
        if exp is not None and (_EXPORTS_SLUG, dim_value) not in seen:
            seen.add((_EXPORTS_SLUG, dim_value))
            rows.append(
                _make_row(
                    _EXPORTS_SLUG, _EXPORTS_NAME, dimension_kind, dim_value, name, exp, period
                )
            )
    return rows, errors


# ---------------------------------------------------------------------------
# XLSX reading — locate the sheets and feed their rows to the deterministic core.
# ---------------------------------------------------------------------------


def _sheet_rows(wb: object, name: str) -> list[tuple[object, ...]] | None:
    """Return a worksheet's rows as value tuples, or None if the sheet is absent."""
    if name not in wb.sheetnames:  # type: ignore[attr-defined]
        return None
    ws = wb[name]  # type: ignore[index]
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _read_descriptor(wb: object) -> str:
    """Return the index sheet's row-0 period descriptor text (col C), or ''.

    The descriptor lives in the third cell of the first row of the index sheet
    ("Based on Annual data (Shrawan-Asar) of FY 2081/82 (...)"). We scan the first
    few cells defensively in case of a column shift.
    """
    rows = _sheet_rows(wb, _SHEET_INDEX)
    if not rows:
        return ""
    for row in rows[:3]:
        for cell in row:
            text = _norm(cell)
            if _DESCRIPTOR_RE.search(text):
                return text
    return ""


# Uniform per-sheet extractor signature for the dispatch table in
# ``parse_customs_fts``: ``(sheet_rows, period) -> (rows, errors)``. The four
# adapters below bind each sheet's measure / dimension / value column to this
# shape so the dispatch loop is a flat table iteration (no if/elif chain).
_SheetHandler = Callable[
    [Sequence[tuple[object, ...]], "_Period"],
    tuple[list["DimensionalRowDraft"], list[ParserError]],
]


def _handle_imports_commodity(
    rows: Sequence[tuple[object, ...]], period: _Period
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    # Imports-by-commodity value column is index 4 (Imports_Value).
    return extract_commodity_rows(rows, _IMPORTS_SLUG, _IMPORTS_NAME, period, 4)


def _handle_exports_commodity(
    rows: Sequence[tuple[object, ...]], period: _Period
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    # Exports-by-commodity value column is index 4 (Exports_Value).
    return extract_commodity_rows(rows, _EXPORTS_SLUG, _EXPORTS_NAME, period, 4)


def _handle_imports_commodity_partner(
    rows: Sequence[tuple[object, ...]], period: _Period
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    # Cross-tab imports: 0=HS,1=Desc,2=Partner,3=Unit,4=Qty,5=Imports_Value.
    return extract_commodity_partner_rows(
        rows, _IMPORTS_SLUG, _IMPORTS_NAME, _DIM_IMPORT_SOURCE, period, 5
    )


def _handle_exports_commodity_partner(
    rows: Sequence[tuple[object, ...]], period: _Period
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    # Cross-tab exports: 0=HS,1=Desc,2=Partner,3=Unit,4=Qty,5=Exports_Value.
    return extract_commodity_partner_rows(
        rows, _EXPORTS_SLUG, _EXPORTS_NAME, _DIM_EXPORT_DESTINATION, period, 5
    )


def parse_customs_fts(source_document_path: str, source_document_id: str) -> CustomsResult:
    """Parse a Department-of-Customs FTS workbook → dimensional facts (ADR-0015).

    Reads the period from the index descriptor (self-describing), then walks the
    four single-dimension value sheets (imports-by-commodity, exports-by-commodity,
    trade-balance-by-country, customs-wise) plus the two commodity×partner
    cross-tabs (sheets 4 & 6 → composite ``<hs>__<country>`` dimension, ADR-0018),
    emitting one fact per dimension member.
    Never raises on bad data: a missing file / unreadable workbook / unparseable
    descriptor yields a typed error and ``status='failure'``; missing individual
    sheets degrade to ``'partial'``.
    """
    _ = source_document_id  # threaded for orchestrator-contract symmetry

    path = Path(source_document_path)
    if not path.exists():
        return CustomsResult(
            status="failure",
            parser_version=PARSER_VERSION,
            dimensional_rows=[],
            errors=[
                ParserError(error_class="Other", error_detail=f"source file not found: {path}")
            ],
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises a variety of types
        return CustomsResult(
            status="failure",
            parser_version=PARSER_VERSION,
            dimensional_rows=[],
            errors=[
                ParserError(
                    error_class="EncodingError",
                    error_detail=f"openpyxl could not read {path.name}: {exc}",
                )
            ],
        )

    try:
        descriptor = _read_descriptor(wb)
        period, period_err = parse_period_descriptor(descriptor)
        if period is None:
            return CustomsResult(
                status="failure",
                parser_version=PARSER_VERSION,
                dimensional_rows=[],
                errors=[
                    period_err
                    or ParserError(
                        error_class="RegexMismatch",
                        error_detail="period descriptor not found on the index sheet",
                    )
                ],
            )

        all_rows: list[DimensionalRowDraft] = []
        all_errors: list[ParserError] = []

        # Per-sheet handler table: each adapter has the uniform signature
        # (sheet_rows, period) -> (rows, errors). Iterating a table (rather than an
        # if/elif chain) keeps this dispatch flat; a missing sheet is a typed
        # PageLayoutChanged that degrades status to partial (never a crash).
        plan: list[tuple[str, _SheetHandler]] = [
            (_SHEET_IMPORTS_COMMODITY, _handle_imports_commodity),
            (_SHEET_EXPORTS_COMMODITY, _handle_exports_commodity),
            (_SHEET_COUNTRY, extract_country_rows),
            (_SHEET_CUSTOMS, extract_customs_rows),
            (_SHEET_IMPORTS_COMMODITY_PARTNER, _handle_imports_commodity_partner),
            (_SHEET_EXPORTS_COMMODITY_PARTNER, _handle_exports_commodity_partner),
        ]
        for sheet_name, handler in plan:
            rows = _sheet_rows(wb, sheet_name)
            if rows is None:
                all_errors.append(
                    ParserError(
                        error_class="PageLayoutChanged",
                        error_detail=f"expected sheet {sheet_name!r} not present — layout changed",
                    )
                )
                continue
            r, e = handler(rows, period)
            all_rows.extend(r)
            all_errors.extend(e)
    finally:
        wb.close()

    if not all_rows:
        return CustomsResult(
            status="failure",
            parser_version=PARSER_VERSION,
            dimensional_rows=[],
            errors=all_errors
            or [ParserError(error_class="Other", error_detail="NoDataExtracted: no facts parsed")],
        )

    status: ParserStatus = "partial" if all_errors else "success"
    return CustomsResult(
        status=status,
        parser_version=PARSER_VERSION,
        dimensional_rows=all_rows,
        errors=all_errors,
    )


def _main() -> None:
    """CLI entrypoint (orchestrator contract — mirror of nrb_dne / mof_yellowbook).

    Argv: ``parser.py <source_document_path> <source_document_id>``. Writes the
    result JSON (including the ``dimensional_rows`` key the ingest CLI reads) to
    stdout. Datetimes are ISO-8601 strings. Exit codes: 0 = ran (status may be
    failure), 2 = usage error.
    """
    expected_argv = 3  # progname + path + doc id
    if len(sys.argv) != expected_argv:
        sys.stderr.write("usage: parser.py <source_document_path> <source_document_id>\n")
        sys.exit(2)
    result = parse_customs_fts(sys.argv[1], sys.argv[2])
    json.dump(result.to_json_dict(), sys.stdout)


if __name__ == "__main__":
    _main()

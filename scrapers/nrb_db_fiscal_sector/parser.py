"""NRB Database on Nepalese Economy — Fiscal Sector XLSX parser.

Source: Nepal Rastra Bank, Database on Nepalese Economy, Fiscal Sector
URL:    https://www.nrb.org.np/database-on-nepalese-economy/fiscal-sector/
Format: XLSX (openpyxl)

Strategy:
    Four XLSX files, each covering a different fiscal indicator. The parser
    opens each file with openpyxl (data_only=True), navigates to the most
    recent sheet (the active sheet), skips header rows, then locates target
    rows by scanning column A or B for known label strings. Month columns are
    resolved from the header row (Aug–Jul for monthly files; fiscal-year labels
    for the annual file).

    The parser is intentionally called once per file (source_document_path,
    source_document_id) with a ``file_key`` argument that identifies which of
    the four datasets is being parsed. Alternatively the caller may pass the
    filepath and the parser auto-detects the dataset from the filename.

    For monthly files (revenue, expenditure, domestic debt):
        - One sheet per fiscal year (e.g. "Rev Col 2024-25").
        - Month header row maps AD month names (Aug–Jul) to BS month names.
        - Values are cumulative YTD (not incremental).
        - Emits one StagingRowDraft per mid-month column.

    For the annual file (foreign debt):
        - One sheet covering FY 2010/11–2022/23.
        - Column headers are AD fiscal-year labels (e.g. "2010/11").
        - Row 25 "Net Outstanding Foreign Debt" — unit Rs. in 10 million;
          multiply by 10 to convert to Rs. in million before emitting.
        - Emits one StagingRowDraft per fiscal year column.

Month → BS month mapping (AD month label in the sheet → BS month):
    Aug  → Bhadra   (mid-Bhadra ≈ mid-August in the BS calendar)
    Sep  → Ashwin
    Oct  → Kartik
    Nov  → Mangsir
    Dec  → Poush
    Jan  → Magh
    Feb  → Falgun
    Mar  → Chait
    Apr  → Baisakh
    May  → Jestha
    Jun  → Ashadh
    Jul  → Shrawan (next BS year)

Known breakage modes (source profile §"Known breakage modes"):
    - Upload URL embeds the upload date; the caller resolves the URL.
    - Revenue/expenditure figures are preliminary and revised monthly.
    - Row labels shift slightly across years; lookup is fuzzy (startswith +
      strip). If a target label is not found, ColumnMissing is emitted.
    - Domestic debt has 6 header rows (rows 0–5); revenue/expenditure have 5.
    - External debt file uses Rs. in 10 million; multiplied by 10 on emit.

Versioning:
    Bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import (
    BsMonth,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ParserError,
    ParserResult,
    ParserStatus,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-db-fiscal-sector"

# ── Month label → BS month (for monthly files, Aug–Jul columns) ───────────────

_AD_MONTH_TO_BS: Final[dict[str, BsMonth]] = {
    "aug": "Bhadra",
    "sep": "Ashwin",
    "oct": "Kartik",
    "nov": "Mangsir",
    "dec": "Poush",
    "jan": "Magh",
    "feb": "Falgun",
    "mar": "Chait",
    "apr": "Baisakh",
    "may": "Jestha",
    "jun": "Ashadh",
    "jul": "Shrawan",
}

# BS months that belong to the *next* BS year relative to the FY start year.
# FY starts Shrawan (BS year N). Aug–Jun map to BS year N; Jul = Shrawan maps
# to BS year N+1 (start of the following FY).
_NEXT_BS_YEAR_MONTHS: Final[frozenset[BsMonth]] = frozenset({"Shrawan"})

# ── Fiscal year label parsing ─────────────────────────────────────────────────

_FY_LABEL_RE: Final[re.Pattern[str]] = re.compile(
    r"(\d{4})/(\d{2,4})",
)


def _parse_fy_label(label: str) -> int | None:
    """Parse an AD fiscal-year label (e.g. '2024/25') → BS start year."""
    m = _FY_LABEL_RE.search(str(label).strip())
    if not m:
        return None
    ad_start = int(m.group(1))
    return ad_start + 57  # approximate BS start year


def _parse_fy_label_ad(label: str) -> int | None:
    """Parse an AD fiscal-year label → AD start year."""
    m = _FY_LABEL_RE.search(str(label).strip())
    if not m:
        return None
    return int(m.group(1))


# ── Sheet → fiscal year detection ────────────────────────────────────────────


def _sheet_bs_fy_start(sheet_title: str) -> int | None:
    """Extract BS FY start year from a sheet name like 'Rev Col 2024-25'."""
    # Normalise separators: "2024-25" or "2024/25"
    normalised = re.sub(r"[-/]", "/", sheet_title)
    m = re.search(r"(\d{4})/\d{2,4}", normalised)
    if m:
        ad_start = int(m.group(1))
        return ad_start + 57
    return None


# ── Row label matchers ────────────────────────────────────────────────────────

def _label_matches(cell_value: object, target: str) -> bool:
    """True if the stripped string form of *cell_value* equals *target*
    (case-insensitive).  Handles leading/trailing whitespace and non-breaking
    spaces."""
    if cell_value is None:
        return False
    cleaned = str(cell_value).replace(" ", " ").strip()
    return cleaned.lower() == target.lower()


def _label_startswith(cell_value: object, prefix: str) -> bool:
    """True if stripped cell text starts with *prefix* (case-insensitive)."""
    if cell_value is None:
        return False
    cleaned = str(cell_value).replace(" ", " ").strip()
    return cleaned.lower().startswith(prefix.lower())


# ── Column header parsing (monthly files) ────────────────────────────────────


def _parse_month_headers(
    header_row: tuple[object, ...],
    first_data_col: int,
) -> list[tuple[int, str, BsMonth]]:
    """Return list of (col_index, ad_label, bs_month) for month columns.

    *header_row* is the 0-indexed row tuple (already read by openpyxl).
    *first_data_col* is the 0-based column index where data begins (after
    label columns).
    """
    result: list[tuple[int, str, BsMonth]] = []
    for col_idx in range(first_data_col, len(header_row)):
        raw = header_row[col_idx]
        if raw is None:
            continue
        label = str(raw).strip()
        bs_month = _AD_MONTH_TO_BS.get(label.lower())
        if bs_month is not None:
            result.append((col_idx, label, bs_month))
    return result


# ── File-type detection ───────────────────────────────────────────────────────

class _FileType:
    REVENUE = "revenue"
    EXPENDITURE = "expenditure"
    DOMESTIC_DEBT = "domestic_debt"
    FOREIGN_DEBT = "foreign_debt"


_FILENAME_TO_FILE_TYPE: Final[dict[str, str]] = {
    "government-revenue": _FileType.REVENUE,
    "government-budgetary": _FileType.EXPENDITURE,
    "outstanding-government-debt": _FileType.DOMESTIC_DEBT,
    "loan-and-debt-servicing": _FileType.FOREIGN_DEBT,
}


def _detect_file_type(path: Path) -> str | None:
    stem = path.stem.lower()
    for key, ft in _FILENAME_TO_FILE_TYPE.items():
        if key in stem:
            return ft
    return None


# ── Monthly file parsers ──────────────────────────────────────────────────────

_REVENUE_SLUG = "nrb-fiscal-revenue-cumulative-ytd"
_EXPENDITURE_SLUG = "nrb-fiscal-expenditure-cumulative-ytd"
_DOMESTIC_DEBT_SLUG = "nrb-fiscal-debt-domestic-outstanding"
_FOREIGN_DEBT_SLUG = "nrb-fiscal-debt-external-outstanding"


def _parse_monthly_file(
    ws: object,  # openpyxl Worksheet
    target_slug: str,
    target_label: str,
    label_col: int,
    value_col_start: int,
    month_header_row_idx: int,
    bs_fy_start: int,
    unit: str,
    pub_ad: datetime,
    pub_bs: str,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse a monthly YTD worksheet and emit one row per month column.

    Args:
        ws: openpyxl worksheet object (iterable rows).
        target_slug: indicator slug to assign.
        target_label: string to match in the label column.
        label_col: 0-based column index holding row labels.
        value_col_start: 0-based first value column index.
        month_header_row_idx: 0-based row index of the month-label header.
        bs_fy_start: BS year the fiscal year starts (e.g. 2082 for FY 2082/83).
        unit: unit string to emit.
        pub_ad, pub_bs: publication date fields.
        errors: mutable list; ColumnMissing appended when target row not found.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # Resolve month headers from the header row.
    header_row = all_rows[month_header_row_idx]
    month_cols = _parse_month_headers(header_row, value_col_start)
    if not month_cols:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"slug {target_slug!r}: no month columns found in header row "
                f"{month_header_row_idx + 1}"
            ),
        ))
        return []

    # Find the target data row by scanning label_col and label_col±1.
    target_row: tuple[object, ...] | None = None
    for row in all_rows:
        if (
            _label_matches(row[label_col] if len(row) > label_col else None, target_label)
            or (
                label_col + 1 < len(row)
                and _label_matches(row[label_col + 1], target_label)
            )
        ):
            target_row = row
            break

    if target_row is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"slug {target_slug!r}: label {target_label!r} not found in "
                f"column {label_col} or {label_col + 1}"
            ),
        ))
        return []

    fy_bs = fiscal_year_label(bs_fy_start)
    fy_ad = fiscal_year_ad_label(bs_fy_start)
    fy_start_shrawan = mid_month_ad("Shrawan", bs_fy_start)

    staging: list[StagingRowDraft] = []
    for col_idx, ad_label, bs_month in month_cols:
        raw = target_row[col_idx] if col_idx < len(target_row) else None
        if raw is None:
            continue
        try:
            value = float(raw)
        except (TypeError, ValueError):
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"slug {target_slug!r}, month {ad_label}: "
                    f"cannot parse {raw!r} as float"
                ),
            ))
            continue

        # Determine BS year for this mid-month
        # Jul (Shrawan) maps to bs_fy_start+1; all others to bs_fy_start
        if bs_month in _NEXT_BS_YEAR_MONTHS:
            bs_year_for_month = bs_fy_start + 1
        else:
            bs_year_for_month = bs_fy_start

        period_end = mid_month_ad(bs_month, bs_year_for_month)
        period_bs = f"{fy_bs} {bs_month}"

        staging.append(StagingRowDraft(
            indicator_slug_raw=target_slug,
            value=value,
            unit=unit,
            reporting_period_type="year_to_date",
            reporting_period_bs=period_bs,
            reporting_period_ad_start=fy_start_shrawan,
            reporting_period_ad_end=period_end,
            publication_date_ad=pub_ad,
            publication_date_bs=pub_bs,
            fiscal_year_bs=fy_bs,
            fiscal_year_ad_label=fy_ad,
            confidence_grade_proposed="B",
            parser_notes="Cumulative YTD; NRB compiles from MoF; preliminary figures",
        ))

    return staging


# ── Revenue parser ────────────────────────────────────────────────────────────

def _parse_revenue(path: Path, pub_ad: datetime, errors: list[ParserError]) -> list[StagingRowDraft]:
    """Government-revenue-1.xlsx → nrb-fiscal-revenue-cumulative-ytd."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    # Detect BS FY start from sheet title.
    bs_fy_start = _sheet_bs_fy_start(ws.title)
    if bs_fy_start is None:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"revenue: cannot parse fiscal year from sheet title {ws.title!r}"
            ),
        ))
        return []

    fy_bs = fiscal_year_label(bs_fy_start)
    pub_bs = f"~{fy_bs} (heuristic)"

    # Layout: row 1=title, row 2=blank, row 3=unit, row 4=HEADS/Mid-Months, row 5=Aug..Jul
    # Header row is index 4 (0-based).
    # Label cols: col A (index 0) = major head label, col B (index 1) = sub-label.
    # Data starts col C (index 2).
    return _parse_monthly_file(
        ws=ws,
        target_slug=_REVENUE_SLUG,
        target_label="Total Revenue",
        label_col=1,
        value_col_start=2,
        month_header_row_idx=4,
        bs_fy_start=bs_fy_start,
        unit="npr_million",
        pub_ad=pub_ad,
        pub_bs=pub_bs,
        errors=errors,
    )


# ── Expenditure parser ────────────────────────────────────────────────────────

def _parse_expenditure(path: Path, pub_ad: datetime, errors: list[ParserError]) -> list[StagingRowDraft]:
    """Government-budgetary-operation.xlsx → nrb-fiscal-expenditure-cumulative-ytd."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    bs_fy_start = _sheet_bs_fy_start(ws.title)
    if bs_fy_start is None:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"expenditure: cannot parse fiscal year from sheet title {ws.title!r}"
            ),
        ))
        return []

    fy_bs = fiscal_year_label(bs_fy_start)
    pub_bs = f"~{fy_bs} (heuristic)"

    # Layout: row 1=title, row 2=blank, row 3=unit, row 4=Heading/FY label, row 5=Aug..Jul
    # Header row is index 4 (0-based).
    # Label col A (index 0) — no col B sub-label for "Total Expenditure".
    # Data starts col B (index 1).
    return _parse_monthly_file(
        ws=ws,
        target_slug=_EXPENDITURE_SLUG,
        target_label="Total Expenditure",
        label_col=0,
        value_col_start=1,
        month_header_row_idx=4,
        bs_fy_start=bs_fy_start,
        unit="npr_million",
        pub_ad=pub_ad,
        pub_bs=pub_bs,
        errors=errors,
    )


# ── Domestic debt parser ──────────────────────────────────────────────────────

def _parse_domestic_debt(path: Path, pub_ad: datetime, errors: list[ParserError]) -> list[StagingRowDraft]:
    """Outstanding-government-debt-1.xlsx → nrb-fiscal-debt-domestic-outstanding."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    bs_fy_start = _sheet_bs_fy_start(ws.title)
    if bs_fy_start is None:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"domestic_debt: cannot parse fiscal year from sheet title {ws.title!r}"
            ),
        ))
        return []

    fy_bs = fiscal_year_label(bs_fy_start)
    pub_bs = f"~{fy_bs} (heuristic)"

    # Layout: row 1=Table ref, row 2=title, row 3=blank, row 4=unit, row 5=Name/Mid-Months, row 6=Aug..Jul
    # 6 header rows → month headers at row index 5 (0-based).
    # Label col A (index 0) = S.N., col B (index 1) = name label.
    # Data starts col C (index 2).
    return _parse_monthly_file(
        ws=ws,
        target_slug=_DOMESTIC_DEBT_SLUG,
        target_label="Total Domestic Debt",
        label_col=1,
        value_col_start=2,
        month_header_row_idx=5,
        bs_fy_start=bs_fy_start,
        unit="npr_million",
        pub_ad=pub_ad,
        pub_bs=pub_bs,
        errors=errors,
    )


# ── Foreign debt parser (annual) ──────────────────────────────────────────────

_FOREIGN_DEBT_LABEL = "Net Outstanding Foreign Debt"
_FOREIGN_DEBT_UNIT_FACTOR: Final[float] = 10.0  # Rs. in 10 million → Rs. in million


def _parse_foreign_debt(path: Path, pub_ad: datetime, errors: list[ParserError]) -> list[StagingRowDraft]:
    """Loan-and-debt-servicing-1.xlsx → nrb-fiscal-debt-external-outstanding.

    The file has annual data only (FY 2010/11–2022/23 on the active sheet).
    Values are in Rs. in 10 million; multiply by 10 to get NPR million.
    Row 25 (1-indexed) = "Net Outstanding Foreign Debt" (Total section).
    Row 5 (1-indexed) = fiscal-year column headers (AD labels).
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))

    # Row 5 (index 4) holds AD fiscal-year headers starting from col 2 (index 2).
    fy_header_row = all_rows[4]

    # Build (col_idx, ad_fy_start_year, bs_fy_start) list.
    fy_cols: list[tuple[int, int, int]] = []
    for col_idx in range(2, len(fy_header_row)):
        raw = fy_header_row[col_idx]
        if raw is None:
            continue
        ad_start = _parse_fy_label_ad(str(raw))
        if ad_start is not None:
            bs_start = ad_start + 57
            fy_cols.append((col_idx, ad_start, bs_start))

    if not fy_cols:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                "foreign_debt: no fiscal-year columns found in row 5 of "
                f"{path.name}"
            ),
        ))
        return []

    # Find the "Net Outstanding Foreign Debt" row in the Total section (row 25,
    # 1-indexed = index 24). Scan all rows for robustness.
    target_row: tuple[object, ...] | None = None
    target_row_num = -1
    # Prefer the row in the Total section (after row 19 which has 'Total' label).
    # Strategy: find last row matching the label.
    for row_idx, row in enumerate(all_rows):
        for ci in range(min(2, len(row))):
            if _label_matches(row[ci], _FOREIGN_DEBT_LABEL):
                target_row = row
                target_row_num = row_idx + 1  # 1-based for logging

    if target_row is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"foreign_debt: label {_FOREIGN_DEBT_LABEL!r} not found in "
                f"{path.name}"
            ),
        ))
        return []

    staging: list[StagingRowDraft] = []
    for col_idx, ad_start, bs_start in fy_cols:
        raw = target_row[col_idx] if col_idx < len(target_row) else None
        if raw is None:
            continue
        try:
            value_10m = float(raw)
        except (TypeError, ValueError):
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"foreign_debt FY{ad_start}/{(ad_start+1)%100:02d}: "
                    f"cannot parse {raw!r} as float"
                ),
            ))
            continue

        value_million = value_10m * _FOREIGN_DEBT_UNIT_FACTOR

        fy_bs = fiscal_year_label(bs_start)
        fy_ad = fiscal_year_ad_label(bs_start)
        pub_bs = f"~{fy_bs} (heuristic)"

        # Annual period: Shrawan of bs_start → Ashadh of bs_start+1.
        period_start = mid_month_ad("Shrawan", bs_start)
        period_end = mid_month_ad("Ashadh", bs_start + 1)

        staging.append(StagingRowDraft(
            indicator_slug_raw=_FOREIGN_DEBT_SLUG,
            value=value_million,
            unit="npr_million",
            reporting_period_type="annual",
            reporting_period_bs=f"FY {fy_bs}",
            reporting_period_ad_start=period_start,
            reporting_period_ad_end=period_end,
            publication_date_ad=pub_ad,
            publication_date_bs=pub_bs,
            fiscal_year_bs=fy_bs,
            fiscal_year_ad_label=fy_ad,
            confidence_grade_proposed="B",
            parser_notes=(
                f"Source row {target_row_num}: Rs. in 10 million × 10 → "
                f"NPR million; NRB compiled from FCGO/PDMO"
            ),
        ))

    return staging


# ── Top-level parse() ─────────────────────────────────────────────────────────

_ALL_SLUGS: Final[frozenset[str]] = frozenset({
    _REVENUE_SLUG,
    _EXPENDITURE_SLUG,
    _DOMESTIC_DEBT_SLUG,
    _FOREIGN_DEBT_SLUG,
})


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse one NRB Fiscal Sector XLSX file; emit staging rows.

    The file type (revenue / expenditure / domestic debt / foreign debt) is
    auto-detected from the filename.  The caller must pass the correct file;
    this function does not download anything.

    Arguments:
        source_document_path: filesystem path to the downloaded XLSX.
        source_document_id:   opaque FK from ``source_documents``; threaded
                              through for orchestrator symmetry.

    Returns:
        ``ParserResult`` with ``status``, ``staging_rows``, ``errors``.
    """
    _ = source_document_id  # threaded through; not used by parser

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source file not found: {path}",
            )],
        )

    file_type = _detect_file_type(path)
    if file_type is None:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=(
                    f"cannot detect file type from filename {path.name!r}. "
                    "Expected one of: government-revenue, "
                    "government-budgetary, outstanding-government-debt, "
                    "loan-and-debt-servicing"
                ),
            )],
        )

    # Approximate publication date: use file mtime if available, else now.
    try:
        mtime = path.stat().st_mtime
        pub_ad = datetime.fromtimestamp(mtime, tz=UTC).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
    except OSError:
        pub_ad = datetime.now(UTC).replace(hour=0, minute=0, second=0, microsecond=0)

    errors: list[ParserError] = []
    staging: list[StagingRowDraft] = []

    try:
        if file_type == _FileType.REVENUE:
            staging = _parse_revenue(path, pub_ad, errors)
        elif file_type == _FileType.EXPENDITURE:
            staging = _parse_expenditure(path, pub_ad, errors)
        elif file_type == _FileType.DOMESTIC_DEBT:
            staging = _parse_domestic_debt(path, pub_ad, errors)
        elif file_type == _FileType.FOREIGN_DEBT:
            staging = _parse_foreign_debt(path, pub_ad, errors)
    except (OSError, KeyError, IndexError, TypeError, ValueError) as exc:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"unexpected error parsing {path.name}: {exc}",
            )],
        )

    if not staging and not errors:
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail=f"no staging rows produced from {path.name}",
        ))

    if not staging:
        status: ParserStatus = "failure"
    elif errors:
        status = "partial"
    else:
        status = "success"

    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging,
        errors=errors,
    )


# ── CLI entrypoint ────────────────────────────────────────────────────────────

def _main() -> None:
    """CLI entrypoint used by the Node ingestion orchestrator.

    Argv: ``parser.py <source_document_path> <source_document_id>``.
    Writes JSON to stdout. Exit codes: 0 = ran (check status); 2 = usage.
    """
    import json
    import sys
    from dataclasses import asdict

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    payload = asdict(result)
    for row in payload.get("staging_rows", []):
        for key in (
            "reporting_period_ad_start",
            "reporting_period_ad_end",
            "publication_date_ad",
        ):
            val = row.get(key)
            if isinstance(val, datetime):
                row[key] = val.isoformat()
    json.dump(payload, sys.stdout)


if __name__ == "__main__":
    _main()

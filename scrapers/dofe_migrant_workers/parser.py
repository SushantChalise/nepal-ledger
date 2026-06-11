"""DoFE migrant-worker labour-permit COUNTS — deterministic parser (ADR-0026).

Source: the NRB "Migrant-Workers-Remittance.xlsx" workbook (filename says
"Remittance" but every sheet holds migrant-worker permit COUNTS, not rupees).
The data is the Department of Foreign Employment (DoFE) labour-permit corpus
re-published by NRB. Three sheets, three cuts of the same monthly outflow:

  1. ``district``        — permits by ORIGIN district  (col0 = district name)
  2. ``Country``         — permits by DESTINATION country (col0 = country name)
  3. ``Migrant Worker``  — monthly New / Renew / Total Worker's Outflow series

Output rows satisfy ``MigrationPermitFactInputSchema``
(``src/lib/ingestion/migration-permit-types.ts``, ADR-0026): one record per
non-empty cell, with NULL in every dimension this cut does not specify
(marginal/aggregate). ``sex`` is always present; ``total`` is the explicit
both-sexes column we READ — never a sum we compute.

Determinism (ADR-0003): pure openpyxl table reading. No LLM, no network. Same
input ⇒ byte-identical JSON output. Numeric coercion only; non-numeric / blank
cells are SKIPPED — counts are never fabricated.

Calendar (ADR-0013):
  - The wide sheets carry AD fiscal-year banners ("2021/22"). Each is converted
    to the BS fiscal year by +57 on the start year (2021/22 AD ⇒ 2078/79 BS),
    matching ``scrapers/nrb_remittance_history/parser.py``.
  - Nepali fiscal month 1–12 from the "Mid-<AD-month>" header at each month
    group's anchor column. Per the source's own note, "August corresponds to
    Shrawan" — so Mid-Aug = month 1 (Shrawan) … Mid-Jul = month 12 (Ashar).
  - The ``Migrant Worker`` sheet keys on an AD calendar date per row; its AD
    (year, month) maps to the same (fiscal_year_bs, month_num) via the inverse
    of the same Aug-anchored cycle.

Wide-sheet layout (verified on the real file, 0-indexed rows):
    r0  : title
    r2  : SPARSE AD fiscal-year banner — one label at each 36-col block's head
    r3  : month labels — "Mid-<AD-month>" at each 3-col group's ANCHOR column
          (a stray BS month name sometimes sits in the +1 / Female slot — ignored)
    r4  : "Male" / "Female" / "Total" repeating per 3-col group
    r5+ : one row per district / country (col0 = name); a trailing "Total"
          aggregate row + footnote rows are excluded as dimensions.
Within a group: anchor = Male, anchor+1 = Female, anchor+2 = Total.

Reconciliation gate (the correctness proof): for a given (fiscal_year_bs,
month_num) the sum of district 'total' permits == the sum of country 'total'
permits == the ``Migrant Worker`` sheet's "Total Worker's Outflow" for that
month (all three are the national monthly outflow). Run with ``--reconcile``
(or ``--verify``) to print it; the test asserts it.

Usage (from scrapers/ on PYTHONPATH):
    python scrapers/dofe_migrant_workers/parser.py <xlsx>             # JSON to stdout
    python scrapers/dofe_migrant_workers/parser.py <xlsx> --verify    # diagnostics → stderr
    python scrapers/dofe_migrant_workers/parser.py <xlsx> --reconcile # reconcile table → stderr
"""

from __future__ import annotations

import datetime as _dt
import json
import re
import sys
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

PARSER_VERSION: Final[str] = "0.1.0"

SHEET_DISTRICT: Final[str] = "district"
SHEET_COUNTRY: Final[str] = "Country"
SHEET_MIGRANT: Final[str] = "Migrant Worker"

AD_TO_BS_OFFSET: Final[int] = 57

UNIT: Final[str] = "permits"

# Row labels in the wide sheets that are NOT a district/country dimension.
# (Aggregate "Total" row + the all-zero placeholder + footnotes.) Matched on a
# normalized col0 label.
_NON_DIMENSION_LABELS: Final[frozenset[str]] = frozenset({"total", "nepal"})

# A fiscal-year banner token: "2021/22" (AD calendar fiscal year).
_FY_RE: Final[re.Pattern[str]] = re.compile(r"^(19|20)(\d{2})\s*/\s*(\d{2})$")

# Nepali fiscal month number (1–12) keyed by the AD month that the "Mid-<month>"
# header names. Source note: "August corresponds to Shrawan" ⇒ Aug = 1.
_AD_MONTH_TO_FISCAL: Final[dict[int, int]] = {
    8: 1,   # Mid-Aug  → Shrawan
    9: 2,   # Mid-Sep  → Bhadra
    10: 3,  # Mid-Oct  → Ashoj
    11: 4,  # Mid-Nov  → Kartik
    12: 5,  # Mid-Dec  → Mangsir
    1: 6,   # Mid-Jan  → Poush
    2: 7,   # Mid-Feb  → Magh
    3: 8,   # Mid-Mar  → Falgun
    4: 9,   # Mid-Apr  → Chaitra
    5: 10,  # Mid-May  → Baisakh
    6: 11,  # Mid-Jun  → Jestha
    7: 12,  # Mid-Jul  → Ashar
}

# Map the leading word of a "Mid-<AD-month>" header (abbreviated OR full) to the
# AD month number. e.g. "midaug"/"midaugust" → 8, "midmar"/"midmarch" → 3.
_MID_PREFIX_TO_AD_MONTH: Final[tuple[tuple[str, int], ...]] = (
    ("jan", 1), ("feb", 2), ("mar", 3), ("apr", 4), ("may", 5), ("jun", 6),
    ("jul", 7), ("aug", 8), ("sep", 9), ("oct", 10), ("nov", 11), ("dec", 12),
)

_SEX_BY_OFFSET: Final[tuple[str, str, str]] = ("male", "female", "total")


def _norm(s: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower()) if s is not None else ""


def _ad_fy_to_bs(token: str) -> str | None:
    """'2021/22' (AD) → '2078/79' (BS), else None."""
    m = _FY_RE.match(token.strip())
    if not m:
        return None
    ad_start = int(m.group(1) + m.group(2))
    bs_start = ad_start + AD_TO_BS_OFFSET
    return f"{bs_start}/{(bs_start + 1) % 100:02d}"


def _mid_label_to_month(label: object) -> int | None:
    """'Mid-Aug' / 'Mid-August' → 1, … 'Mid-Jul' → 12. Non-month → None."""
    if not isinstance(label, str):
        return None
    norm = _norm(label)
    if not norm.startswith("mid"):
        return None
    rest = norm[3:]
    for prefix, ad_month in _MID_PREFIX_TO_AD_MONTH:
        if rest.startswith(prefix):
            return _AD_MONTH_TO_FISCAL[ad_month]
    return None


def _ad_date_to_period(d: _dt.date) -> tuple[str, int]:
    """An AD calendar (year, month) → (fiscal_year_bs, month_num).

    The Nepali fiscal year is Aug-anchored here (Aug = month 1). So AD months
    Aug–Dec belong to the FY whose AD start year is ``d.year``; AD months Jan–Jul
    belong to the FY whose AD start year is ``d.year - 1``.
    """
    month_num = _AD_MONTH_TO_FISCAL[d.month]
    ad_start = d.year if d.month >= 8 else d.year - 1
    bs_start = ad_start + AD_TO_BS_OFFSET
    fy_bs = f"{bs_start}/{(bs_start + 1) % 100:02d}"
    return fy_bs, month_num


def _coerce_count(cell: object) -> int | None:
    """Numeric cell → non-negative int, else None (skip — never fabricate)."""
    if isinstance(cell, bool):
        return None
    if isinstance(cell, int):
        return cell if cell >= 0 else None
    if isinstance(cell, float):
        if cell != cell or cell < 0:  # NaN or negative
            return None
        if float(cell).is_integer():
            return int(cell)
        return None  # fractional permit count is not valid
    return None


def _column_periods(rows: list[tuple], ncol: int) -> dict[int, tuple[str, int]]:
    """Map each month-group ANCHOR column → (fiscal_year_bs, month_num).

    FY banner (row idx2) is forward-filled across columns; month groups are
    detected from "Mid-<month>" labels in row idx3 (each such cell is a group
    anchor / Male column).
    """
    fy_row = rows[2]
    month_row = rows[3]

    # Forward-fill the AD→BS fiscal year across all columns.
    fy_at: dict[int, str | None] = {}
    current: str | None = None
    for ci in range(ncol):
        token = fy_row[ci] if ci < len(fy_row) else None
        if isinstance(token, str):
            bs = _ad_fy_to_bs(token)
            if bs is not None:
                current = bs
        fy_at[ci] = current

    anchors: dict[int, tuple[str, int]] = {}
    for ci in range(ncol):
        label = month_row[ci] if ci < len(month_row) else None
        month_num = _mid_label_to_month(label)
        if month_num is None:
            continue
        fy_bs = fy_at.get(ci)
        if fy_bs is None:
            continue  # a month group with no resolvable FY — skip, do not guess
        anchors[ci] = (fy_bs, month_num)
    return anchors


def _parse_wide_sheet(
    rows: list[tuple], name_field: str
) -> list[dict[str, object]]:
    """Parse a district/Country wide sheet into permit-fact rows.

    ``name_field`` is the dimension key the row's col0 name is emitted under:
    ``originDistrict`` (district sheet) or ``destinationCountry`` (Country sheet).
    """
    if len(rows) < 5:
        return []
    ncol = max((len(r) for r in rows), default=0)
    anchors = _column_periods(rows, ncol)

    out: list[dict[str, object]] = []
    for ri in range(5, len(rows)):
        row = rows[ri]
        raw_name = row[0] if len(row) > 0 else None
        if not isinstance(raw_name, str) or not raw_name.strip():
            continue
        name = raw_name.strip()
        if _norm(name) in _NON_DIMENSION_LABELS:
            continue  # aggregate / placeholder row — not a dimension

        for anchor_col, (fy_bs, month_num) in anchors.items():
            for offset, sex in enumerate(_SEX_BY_OFFSET):
                col = anchor_col + offset
                cell = row[col] if col < len(row) else None
                count = _coerce_count(cell)
                if count is None:
                    continue
                rec: dict[str, object] = {
                    "fiscalYearBs": fy_bs,
                    "monthNum": month_num,
                    "destinationCountry": None,
                    "destinationRegion": None,
                    "originDistrict": None,
                    "skillClass": None,
                    "permitCategory": None,
                    "sex": sex,
                    "permits": str(count),
                    "unit": UNIT,
                    "sourceSheet": _name_field_to_sheet(name_field),
                }
                rec[name_field] = name
                out.append(rec)
    return out


def _name_field_to_sheet(name_field: str) -> str:
    return SHEET_DISTRICT if name_field == "originDistrict" else SHEET_COUNTRY


def _parse_migrant_sheet(rows: list[tuple]) -> list[dict[str, object]]:
    """Parse the ``Migrant Worker`` monthly series.

    One row per AD-dated month → two permit-fact rows: a ``new_individual`` row
    (New Entry column) and a ``reentry`` row (Renew Entry column), both
    ``sex='total'``. Cumulative columns are ignored.
    """
    out: list[dict[str, object]] = []
    for ri in range(2, len(rows)):
        row = rows[ri]
        d = row[0] if len(row) > 0 else None
        if isinstance(d, _dt.datetime):
            d = d.date()
        if not isinstance(d, _dt.date):
            continue
        fy_bs, month_num = _ad_date_to_period(d)
        new_entry = _coerce_count(row[1] if len(row) > 1 else None)
        renew_entry = _coerce_count(row[2] if len(row) > 2 else None)
        for category, count in (("new_individual", new_entry), ("reentry", renew_entry)):
            if count is None:
                continue
            out.append(
                {
                    "fiscalYearBs": fy_bs,
                    "monthNum": month_num,
                    "destinationCountry": None,
                    "destinationRegion": None,
                    "originDistrict": None,
                    "skillClass": None,
                    "permitCategory": category,
                    "sex": "total",
                    "permits": str(count),
                    "unit": UNIT,
                    "sourceSheet": SHEET_MIGRANT,
                }
            )
    return out


def parse(xlsx_path: str | Path) -> list[dict[str, object]]:
    """Extract every permit-fact row from the three sheets (JSON-ready dicts)."""
    path = Path(xlsx_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        records: list[dict[str, object]] = []
        if SHEET_DISTRICT in wb.sheetnames:
            rows = list(wb[SHEET_DISTRICT].iter_rows(values_only=True))
            records.extend(_parse_wide_sheet(rows, "originDistrict"))
        if SHEET_COUNTRY in wb.sheetnames:
            rows = list(wb[SHEET_COUNTRY].iter_rows(values_only=True))
            records.extend(_parse_wide_sheet(rows, "destinationCountry"))
        if SHEET_MIGRANT in wb.sheetnames:
            rows = list(wb[SHEET_MIGRANT].iter_rows(values_only=True))
            records.extend(_parse_migrant_sheet(rows))
        return records
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------


def reconcile(records: list[dict[str, object]]) -> dict[tuple[str, int], dict[str, int]]:
    """Per (fiscal_year_bs, month_num): district-total vs country-total vs
    migrant-worker-total (New + Renew). The three should agree.

    Returns ``{(fy, month): {"district": …, "country": …, "migrant": …}}``.
    """
    agg: dict[tuple[str, int], dict[str, int]] = {}

    def _bucket(rec: dict[str, object]) -> dict[str, int]:
        key = (str(rec["fiscalYearBs"]), int(rec["monthNum"]))  # type: ignore[arg-type]
        return agg.setdefault(key, {"district": 0, "country": 0, "migrant": 0})

    for rec in records:
        if rec.get("monthNum") is None:
            continue
        permits = int(str(rec["permits"]))
        if rec.get("originDistrict") is not None and rec["sex"] == "total":
            _bucket(rec)["district"] += permits
        elif rec.get("destinationCountry") is not None and rec["sex"] == "total":
            _bucket(rec)["country"] += permits
        elif rec.get("sourceSheet") == SHEET_MIGRANT and rec["sex"] == "total":
            # New + Renew = Total Worker's Outflow for the month.
            _bucket(rec)["migrant"] += permits
    return agg


def _verify(records: list[dict[str, object]]) -> None:
    by_sheet: dict[object, int] = {}
    for rec in records:
        by_sheet[rec.get("sourceSheet")] = by_sheet.get(rec.get("sourceSheet"), 0) + 1
    sys.stderr.write(
        f"[dofe-migrant-workers] parser v{PARSER_VERSION} — {len(records)} rows\n"
    )
    for sheet, n in sorted(by_sheet.items(), key=lambda kv: str(kv[0])):
        sys.stderr.write(f"  rows from {sheet!r}: {n}\n")

    agg = reconcile(records)
    # Pick a few well-populated months to display (those with all three present).
    full = sorted(
        (k for k, v in agg.items() if v["district"] and v["country"] and v["migrant"])
    )
    sample = full[:5]
    sys.stderr.write("  reconciliation (district-total == country-total == migrant-total):\n")
    all_ok = True
    for key in sample:
        v = agg[key]
        ok = v["district"] == v["country"] == v["migrant"]
        all_ok = all_ok and ok
        sys.stderr.write(
            f"    FY{key[0]} M{key[1]:>2}: district={v['district']:>8,} "
            f"country={v['country']:>8,} migrant={v['migrant']:>8,}  "
            f"{'OK' if ok else 'MISMATCH'}\n"
        )
    sys.stderr.write(
        f"  reconciliation over sampled months: {'ALL OK' if all_ok else 'FAILURES PRESENT'}\n"
    )


def _main() -> None:
    args = sys.argv[1:]
    if not args:
        sys.stderr.write("usage: parser.py <xlsx_path> [--verify|--reconcile]\n")
        sys.exit(2)
    path = Path(args[0])
    if not path.exists():
        sys.stderr.write(f"file not found: {path}\n")
        sys.exit(2)
    records = parse(path)
    if "--verify" in args or "--reconcile" in args:
        _verify(records)
    else:
        json.dump(records, sys.stdout, ensure_ascii=True)


if __name__ == "__main__":
    _main()

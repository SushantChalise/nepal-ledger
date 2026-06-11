"""NRB Interest-Subsidized (Concessional) Loan XLSX parser — deterministic Python.

Source: Nepal Rastra Bank — monthly concessional/interest-subsidized loan
statistics (XLSX). Published at https://www.nrb.org.np/category/concessional-loan/

Strategy:
    The XLSX Summary sheet contains 10 policy-scheme rows (Preeti-encoded Nepali)
    plus a grand-total row. Columns are fixed: position 17 (0-indexed) is the
    Total Outstanding Loan (Rs. Hajar).

    The file title (rows 6-7) encodes the BS month and year in Preeti:
    row 7 typically contains "(@)*@ r}q d;fGt;Dd_" — we extract the BS year
    and BS month name directly from the filename, which is the most reliable
    signal (NRB uses consistent "Month-YYYY" patterns in filenames).

Target indicators (v0.1.0):
    - nrb-concession-total-outstanding        — grand-total outstanding (col 17, row 23)
    - nrb-concession-agriculture-outstanding  — SN=1 outstanding (col 17, row 11)
    - nrb-concession-sme-outstanding          — NOT FOUND in source; NRB does not
      publish a dedicated SME sub-total; closest proxy is SN=4 Women Entrepreneur
      (dlxnf pBdzLn shf{); emitted with parser_notes documenting the proxy.

Column layout (0-indexed) in Summary sheet:
    Col 1  : S.N. (sequence number)
    Col 2  : Loan type (Preeti-encoded Nepali)
    Col 3  : Commercial Banks — Borrowers (C0f ;++Vof)
    Col 4  : Commercial Banks — Approved Limit (:jLs[t shf{)
    Col 5  : Commercial Banks — Outstanding (AffFsL shf{)
    Col 6  : Development Banks — Borrowers
    Col 7  : Development Banks — Approved Limit
    Col 8  : Development Banks — Outstanding
    Col 9  : Finance Companies — Borrowers
    Col 10 : Finance Companies — Approved Limit
    Col 11 : Finance Companies — Outstanding
    Col 12 : Microfinance — Borrowers
    Col 13 : Microfinance — Approved Limit
    Col 14 : Microfinance — Outstanding
    Col 15 : Total — Borrowers
    Col 16 : Total — Approved Limit
    Col 17 : Total — Outstanding  ← TARGET

Row layout (0-indexed) in Summary sheet:
    Rows 0-8  : Title / header area (merged cells, Preeti text, units label)
    Row 9     : Category headers (ka/kha/ga/gha/total)
    Row 10    : Sub-headers (C0f ;++Vof, :jLs[t shf{, AffFsL shf{)
    Row 11    : SN=1  Agriculture & Livestock (grand-total target: row 23, col 17)
    Row 12    : SN=1.1 (without collateral)
    Row 13    : SN=1.2 (with collateral)
    Row 14    : SN=2  Educated Youth
    Row 15    : SN=3  Return Youth
    Row 16    : SN=4  Women Entrepreneur  ← SME proxy
    Row 17    : SN=5  Dalit Business
    Row 18    : SN=6  Higher Education
    Row 19    : SN=7  Earthquake Housing
    Row 20    : SN=8  Textile Industry
    Row 21    : SN=9  Vocational Training
    Row 22    : SN=10 Youth Self-Employment
    Row 23    : Grand Total (s'n)

Unit: Rs. Hajar (NPR thousands) — stated in row 6, col 17 of the sheet.

Date parsing:
    The NRB filename embeds the BS month name and BS year, e.g.:
    "Interest-subsidized-loan-Chaitra-2082-Publish.xlsx" → Chait BS 2082.
    The parser extracts month + year from the filename using a regex against
    the known BS month name romanisations.
    Fallback: row 7 Preeti text is NOT decoded (Preeti is a proprietary
    font-encoding, not Unicode); filename is the authoritative date source.

Versioning:
    Bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import (
    BsMonth,
    BS_MONTHS,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ParserError,
    ParserResult,
    ParserStatus,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-concessional-loan"

# ── Column / row layout constants ─────────────────────────────────────────────

_SUMMARY_SHEET: Final[str] = "Summary"

# 0-indexed row positions in the Summary sheet.
_ROW_AGRICULTURE: Final[int] = 11   # SN=1  Agriculture & Livestock
_ROW_WOMEN_ENTREPRENEUR: Final[int] = 16  # SN=4  Women Entrepreneur (SME proxy)
_ROW_GRAND_TOTAL: Final[int] = 23   # Grand Total row

# 0-indexed column for Total Outstanding Loan (Rs. Hajar).
_COL_TOTAL_OUTSTANDING: Final[int] = 17

# NRB publication lag heuristic (days from period-end to publication).
_NRB_PUB_LAG_DAYS: Final[int] = 30

# ── BS month name → BsMonth canonical ────────────────────────────────────────
# NRB filenames use mixed romanisations; list all variants encountered.
_FILENAME_MONTH_MAP: Final[dict[str, BsMonth]] = {
    # Shrawan
    "shrawan": "Shrawan", "srawan": "Shrawan",
    # Bhadra
    "bhadra": "Bhadra", "bhadau": "Bhadra", "badau": "Bhadra",
    # Ashwin
    "ashwin": "Ashwin", "aswin": "Ashwin", "ashoj": "Ashwin",
    # Kartik
    "kartik": "Kartik", "kartiki": "Kartik",
    # Mangsir
    "mangsir": "Mangsir", "mansir": "Mangsir",
    # Poush
    "poush": "Poush", "push": "Poush", "paus": "Poush",
    # Magh
    "magh": "Magh", "mag": "Magh",
    # Falgun
    "falgun": "Falgun", "phalgun": "Falgun", "falgan": "Falgun",
    # Chait
    "chaitra": "Chait", "chait": "Chait", "chaita": "Chait",
    # Baisakh
    "baisakh": "Baisakh", "baisak": "Baisakh", "baishakh": "Baisakh",
    # Jestha
    "jestha": "Jestha", "jyestha": "Jestha", "jeth": "Jestha",
    # Ashadh
    "ashadh": "Ashadh", "ashad": "Ashadh", "asar": "Ashadh",
}

# Regex to extract BS month and year from the filename.
# E.g. "Interest-subsidized-loan-Chaitra-2082-Publish.xlsx"
_FILENAME_DATE_RE: Final[re.Pattern[str]] = re.compile(
    r"-(" + "|".join(re.escape(k) for k in _FILENAME_MONTH_MAP) + r")-(\d{4})-",
    re.IGNORECASE,
)

# fiscal-year ordinal (1-indexed, Shrawan=1) for each BS month.
_BS_MONTH_FY_POS: Final[dict[BsMonth, int]] = {
    m: i + 1 for i, m in enumerate(BS_MONTHS)
}


def _parse_date_from_filename(filename: str, errors: list[ParserError]) -> tuple[BsMonth, int] | None:
    """Extract (bs_month, bs_year) from the NRB concessional-loan filename.

    Returns None and appends a PeriodAmbiguous error on failure.
    """
    m = _FILENAME_DATE_RE.search(filename)
    if not m:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"filename {filename!r}: could not extract BS month + year; "
                "expected pattern like 'Chaitra-2082' in the filename"
            ),
        ))
        return None

    raw_month = m.group(1).lower()
    bs_month = _FILENAME_MONTH_MAP.get(raw_month)
    if bs_month is None:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=f"filename: unrecognised BS month token {m.group(1)!r}",
        ))
        return None

    try:
        bs_year = int(m.group(2))
    except ValueError:
        errors.append(ParserError(
            error_class="PeriodAmbiguous",
            error_detail=f"filename: BS year not parseable: {m.group(2)!r}",
        ))
        return None

    return bs_month, bs_year


def _cell_float(rows: list[tuple], row_idx: int, col_idx: int, slug: str, errors: list[ParserError]) -> float | None:
    """Read a cell value as float; append ValueUnparseable on failure."""
    try:
        row = rows[row_idx]
    except IndexError:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=f"{slug}: row {row_idx} does not exist (sheet too short)",
        ))
        return None

    try:
        val = row[col_idx]
    except IndexError:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=f"{slug}: col {col_idx} does not exist in row {row_idx}",
        ))
        return None

    if val is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=f"{slug}: cell [{row_idx},{col_idx}] is empty",
        ))
        return None

    try:
        return float(val)
    except (TypeError, ValueError):
        errors.append(ParserError(
            error_class="ValueUnparseable",
            error_detail=f"{slug}: could not cast {val!r} to float at [{row_idx},{col_idx}]",
            source_excerpt=str(val),
        ))
        return None


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse one NRB concessional-loan XLSX file.

    Arguments:
        source_document_path: filesystem path to the downloaded XLSX.
        source_document_id: opaque FK from ``source_documents``; threaded
            through for orchestrator symmetry.

    Returns:
        ``ParserResult`` with ``status``, ``staging_rows``, ``errors``.
    """
    _ = source_document_id

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source file not found: {path}",
            )],
        )

    errors: list[ParserError] = []

    # ── Date from filename ────────────────────────────────────────────────────
    date_result = _parse_date_from_filename(path.name, errors)
    if date_result is None:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors,
        )
    bs_month, bs_year = date_result

    # Derive fiscal year: Shrawan–Chait belong to FY starting that BS year;
    # Baisakh–Ashadh belong to FY starting the previous BS year.
    fy_pos = _BS_MONTH_FY_POS[bs_month]
    bs_fy_start = bs_year if fy_pos <= 9 else bs_year - 1

    fy_bs = fiscal_year_label(bs_fy_start)
    fy_ad = fiscal_year_ad_label(bs_fy_start)
    period_bs = f"{bs_year}/{(bs_year + 1) % 100:02d} {bs_month}"
    period_end = mid_month_ad(bs_month, bs_year)
    pub_ad = period_end + timedelta(days=_NRB_PUB_LAG_DAYS)
    pub_bs = f"~{fy_bs} (heuristic)"

    # ── Open workbook ─────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl failed to open workbook: {exc}",
            )],
        )

    if _SUMMARY_SHEET not in wb.sheetnames:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="PageLayoutChanged",
                error_detail=(
                    f"sheet {_SUMMARY_SHEET!r} not found; "
                    f"available sheets: {wb.sheetnames}"
                ),
            )],
        )

    ws = wb[_SUMMARY_SHEET]
    rows: list[tuple] = list(ws.iter_rows(values_only=True))

    # ── Base StagingRowDraft template ─────────────────────────────────────────
    base = StagingRowDraft(
        indicator_slug_raw="",
        value=0.0,
        unit="npr_thousand",
        reporting_period_type="monthly",
        reporting_period_bs=period_bs,
        reporting_period_ad_start=period_end,
        reporting_period_ad_end=period_end,
        publication_date_ad=pub_ad,
        publication_date_bs=pub_bs,
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade_proposed="A",
        parser_notes=None,
    )

    staging_rows: list[StagingRowDraft] = []

    # ── 1. Total outstanding (grand total row) ────────────────────────────────
    total_val = _cell_float(rows, _ROW_GRAND_TOTAL, _COL_TOTAL_OUTSTANDING,
                            "nrb-concession-total-outstanding", errors)
    if total_val is not None:
        staging_rows.append(replace(
            base,
            indicator_slug_raw="nrb-concession-total-outstanding",
            value=total_val,
        ))

    # ── 2. Agriculture & Livestock outstanding (SN=1) ────────────────────────
    agri_val = _cell_float(rows, _ROW_AGRICULTURE, _COL_TOTAL_OUTSTANDING,
                           "nrb-concession-agriculture-outstanding", errors)
    if agri_val is not None:
        staging_rows.append(replace(
            base,
            indicator_slug_raw="nrb-concession-agriculture-outstanding",
            value=agri_val,
        ))

    # ── 3. SME proxy — Women Entrepreneur outstanding (SN=4) ─────────────────
    # NRB does NOT publish a dedicated SME sub-total in this series.
    # The 10 schemes are organised by policy program, not economic sector.
    # Closest available proxy: SN=4 Women Entrepreneur Loan (dlxnf pBdzLn shf{).
    # Emitted with confidence grade B and explanatory parser_notes.
    sme_val = _cell_float(rows, _ROW_WOMEN_ENTREPRENEUR, _COL_TOTAL_OUTSTANDING,
                          "nrb-concession-sme-outstanding", errors)
    if sme_val is not None:
        staging_rows.append(replace(
            base,
            indicator_slug_raw="nrb-concession-sme-outstanding",
            value=sme_val,
            confidence_grade_proposed="B",
            parser_notes=(
                "PROXY — NRB does not publish a dedicated SME sub-total in this "
                "series. Mapped to SN=4 Women Entrepreneur Loan (dlxnf pBdzLn shf{), "
                "which is the closest available policy-scheme proxy. "
                "Slug nrb-concession-sme-outstanding cannot be accurately mapped; "
                "confidence downgraded to B."
            ),
        ))

    # ── Status ────────────────────────────────────────────────────────────────
    if not staging_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors or [ParserError(
                error_class="PageLayoutChanged",
                error_detail="no indicators extracted",
            )],
        )

    # All 3 target slugs extracted → success; partial if any errors remain.
    status: ParserStatus = "partial" if errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


def _main() -> None:
    """CLI entrypoint used by the Node ingestion orchestrator.

    Argv: ``parser.py <source_document_path> <source_document_id>``.
    Writes JSON to stdout.
    Exit codes:
      0: parser ran (status may still be 'failure'; consumer reads stdout)
      2: usage error
      1: catastrophic crash (let Python propagate)
    """
    import json
    import sys

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    json.dump(result.to_json_dict(), sys.stdout, default=str)


if __name__ == "__main__":
    _main()

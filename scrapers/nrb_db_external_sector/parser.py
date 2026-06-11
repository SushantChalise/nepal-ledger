"""NRB Database on Nepalese Economy — External Sector parser.

Source: Nepal Rastra Bank "Database on Nepalese Economy" — External Sector
tab. Three XLSX files updated in-place each release cycle.

Files handled:
    MIgrant-Workers_.xlsx — "Migrant Worker" sheet
        Monthly departure counts from the Department of Foreign Employment.
        Date column stores mid-month datetime objects (e.g. 2025-11-25 for
        the Nepali month of Mangsir 2082). The sheet note confirms: "August
        here corresponds to Shrawan month in Nepali Calendar".

    Tourist-arrivals.xlsx — "Tourist Arrival" sheet
        Year-by-month pivot (row = calendar year 1992–present,
        columns Jan–Dec + Total). Emits one StagingRowDraft per
        non-null (year, month) cell.

    Balance-of-Payments-BPM6.xlsx — "BOP BPM6" sheet
        BPM6 classification. Cumulative monthly series within each
        fiscal year (FY starts Shrawan/mid-July). Workers' remittances
        live at row S.N. 1.C.2.1.1 ("O/W Workers' remittances"), Credit
        column only (NPR million).

Target slugs (ADR-0003 / source nrb-db-external-sector):
    nrb-ext-migrant-departures-monthly  — Migrant Worker sheet outflow total
    nrb-ext-tourist-arrivals-monthly    — Tourist Arrival pivot monthly count
    nrb-ext-bop-remittance-workers-monthly — BOP BPM6 workers remittances
                                             cumulative Credit (NPR million)

Unavailable slugs (emitted as ColumnMissing errors):
    nrb-ext-remittance-india-monthly    — no India remittance NPR corridor;
                                          India is excluded from DOFE permit
                                          tracking (zero headcounts)
    nrb-ext-remittance-gulf-monthly     — no Gulf-specific remittance NPR;
                                          BOP BPM6 has no corridor breakdown

Date mapping (AD month → BS month, from periods.py):
    Jul → Ashadh   Aug → Shrawan  Sep → Bhadra  Oct → Ashwin
    Nov → Kartik   Dec → Poush    Jan → Magh     Feb → Falgun
    Mar → Chait    Apr → Baisakh  May → Jestha   Jun → Ashadh

    NB: The migrant-worker datetime object day component (25th, etc.) is
    ignored; only the month (and year) of the datetime object are used to
    identify the BS month.

Confidence grades:
    A — XLSX cells read directly from NRB-published files
    B — cells that carry Excel formula strings (evaluated value unavailable);
        downgraded from A

Versioning: bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Final, Literal

import openpyxl

from _common.periods import (
    BS_MONTHS,
    BsMonth,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ConfidenceGrade,
    ParserError,
    ParserResult,
    ParserStatus,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-db-external-sector"

# ── Slug constants ────────────────────────────────────────────────────────────

SLUG_MIGRANT_DEPARTURES: Final[str] = "nrb-ext-migrant-departures-monthly"
SLUG_TOURIST_ARRIVALS: Final[str] = "nrb-ext-tourist-arrivals-monthly"
SLUG_BOP_REMITTANCE: Final[str] = "nrb-ext-bop-remittance-workers-monthly"

# Slugs that cannot be sourced from these files — emit ColumnMissing.
SLUG_REMITTANCE_INDIA: Final[str] = "nrb-ext-remittance-india-monthly"
SLUG_REMITTANCE_GULF: Final[str] = "nrb-ext-remittance-gulf-monthly"

ALL_TARGET_SLUGS: Final[frozenset[str]] = frozenset({
    SLUG_MIGRANT_DEPARTURES,
    SLUG_TOURIST_ARRIVALS,
    SLUG_BOP_REMITTANCE,
    SLUG_REMITTANCE_INDIA,
    SLUG_REMITTANCE_GULF,
})

# ── AD-month → BS-month mapping ──────────────────────────────────────────────
# Mirrors _BS_MONTH_TO_AD_MONTH from periods.py (inverted).
# When two BS months share an AD calendar month (Shrawan straddles Jul–Aug),
# we use the month whose "mid-month" falls in that AD month.
# The periods.py canonical mapping is: Shrawan→Jul(15), so Aug(datetime month=8)
# is Bhadra. The migrant-worker sheet note confirms Aug datetime = Shrawan
# (which covers mid-Jul to mid-Aug, so NRB assigns the AD-month-8 datetime to
# Shrawan). We honour NRB's own note:
#   Aug→Shrawan, Sep→Bhadra, Oct→Ashwin, Nov→Kartik, Dec→Poush,
#   Jan→Magh, Feb→Falgun, Mar→Chait, Apr→Baisakh, May→Jestha,
#   Jun→Ashadh, Jul→Ashadh   (Jul is end-of-FY, same BS month as Ashadh)
_AD_MONTH_TO_BS: Final[dict[int, BsMonth]] = {
    8: "Shrawan",
    9: "Bhadra",
    10: "Ashwin",
    11: "Kartik",
    12: "Poush",
    1: "Magh",
    2: "Falgun",
    3: "Chait",
    4: "Baisakh",
    5: "Jestha",
    6: "Ashadh",
    7: "Ashadh",
}

# BS month fiscal-year position (1 = Shrawan, 12 = Ashadh).
_BS_MONTH_FY_POS: Final[dict[BsMonth, int]] = {
    m: i + 1 for i, m in enumerate(BS_MONTHS)
}

# Approximate NRB publication lag for heuristic pub-date.
_PUB_LAG_DAYS: Final[int] = 45


def _bs_month_from_ad_month(ad_month: int) -> BsMonth:
    """Map an AD calendar month (1–12) to the corresponding BS month name."""
    return _AD_MONTH_TO_BS[ad_month]


def _bs_year_from_ad(ad_month: int, ad_year: int) -> int:
    """Derive BS year from an AD year + month.

    BS fiscal year N starts mid-July of AD year N-57.
    Months Shrawan(Aug)–Poush(Dec) belong to BS year = AD year + 57.
    Months Magh(Jan)–Ashadh(Jul) belong to BS year = AD year + 57 - 1 = AD year + 56.
    But this is the BS year of the month, not the FY start year.
    """
    if ad_month >= 7:
        # Aug–Dec: BS calendar year matches AD year + 57
        return ad_year + 57
    else:
        # Jan–Jun: still in the same Nepali year that started the previous AD year
        return ad_year + 57 - 1


def _fy_bs_start_from_bs_month_year(bs_month: BsMonth, bs_year: int) -> int:
    """Return the BS fiscal-year start year for a given BS month+year.

    FY start = Shrawan(month-pos 1) of bs_year if month_pos <= 9 (Shrawan–Chait),
               or Shrawan of bs_year-1 if month_pos >= 10 (Baisakh–Ashadh).

    NRB fiscal year runs Shrawan → Ashadh (mid-July → mid-July).
    Months 1–9 (Shrawan–Chait) → FY start year = bs_year.
    Months 10–12 (Baisakh–Ashadh) → FY start year = bs_year - 1.
    """
    pos = _BS_MONTH_FY_POS[bs_month]
    return bs_year if pos <= 9 else bs_year - 1


def _make_base_row(
    bs_month: BsMonth,
    bs_year: int,
    slug: str,
    value: float,
    unit: str,
    confidence: ConfidenceGrade = "A",
    notes: str | None = None,
    period_type: Literal[
        "monthly", "quarterly", "annual", "nine_months_cumulative",
        "year_to_date", "daily", "seasonal"
    ] = "monthly",
) -> StagingRowDraft:
    """Build a StagingRowDraft for a monthly observation."""
    fy_start = _fy_bs_start_from_bs_month_year(bs_month, bs_year)
    fy_bs = fiscal_year_label(fy_start)
    fy_ad = fiscal_year_ad_label(fy_start)
    period_bs = f"{fy_bs} {bs_month}"
    mid = mid_month_ad(bs_month, bs_year)
    pub_ad = mid + timedelta(days=_PUB_LAG_DAYS)
    return StagingRowDraft(
        indicator_slug_raw=slug,
        value=value,
        unit=unit,
        reporting_period_type=period_type,
        reporting_period_bs=period_bs,
        reporting_period_ad_start=mid,
        reporting_period_ad_end=mid,
        publication_date_ad=pub_ad,
        publication_date_bs=f"~{fy_bs} (heuristic)",
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade_proposed=confidence,
        parser_notes=notes,
    )


# ── Migrant Worker parser ─────────────────────────────────────────────────────

def _parse_migrant_workers(
    path: Path,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse MIgrant-Workers_.xlsx → Migrant Worker sheet.

    The "Migrant Worker" sheet has columns:
        A: Month (datetime or None)
        B: New Entry
        C: Renew Entry
        D: Total Worker's Outflow   ← target
        E: Cumulative New Entry
        F: Cumulative Renew Entry
        G: Cumulative Worker's Outflow

    Row 0: title ("Number of Migrant Workers")
    Row 1: column headers
    Rows 2+: data rows until blank/note rows.
    """
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"MIgrant-Workers_.xlsx open failed: {exc}",
        ))
        return []

    if "Migrant Worker" not in wb.sheetnames:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="sheet 'Migrant Worker' not found in MIgrant-Workers_.xlsx",
        ))
        return []

    ws = wb["Migrant Worker"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Verify header row (row index 1).
    if len(rows) < 2:
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail="MIgrant-Workers_ Migrant Worker sheet has fewer than 2 rows",
        ))
        return []

    header = rows[1]
    # "Total Worker's Outflow" is expected at column D (index 3).
    # Tolerate minor variations in apostrophe style.
    total_outflow_col: int | None = None
    for i, h in enumerate(header):
        if h and re.search(r"total\s+worker", str(h), re.IGNORECASE):
            total_outflow_col = i
            break

    if total_outflow_col is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"'Total Worker's Outflow' column not found in row 1 headers: "
                f"{header!r}"
            ),
        ))
        return []

    staging_rows: list[StagingRowDraft] = []

    for row_idx, row in enumerate(rows[2:], start=2):
        date_cell = row[0]
        if date_cell is None:
            continue
        if not isinstance(date_cell, datetime):
            # Skip note/footer rows.
            continue

        value_cell = row[total_outflow_col] if total_outflow_col < len(row) else None
        if value_cell is None:
            continue

        # Formula strings are produced by data_only=False; data_only=True
        # returns the cached evaluated value (int/float) or None.
        if isinstance(value_cell, str):
            # Formula string — cached value unavailable.
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"row {row_idx}: formula string '{value_cell}' for "
                    f"{SLUG_MIGRANT_DEPARTURES} — data_only should have resolved this"
                ),
                source_excerpt=str(value_cell),
            ))
            continue

        try:
            value = float(value_cell)
        except (TypeError, ValueError):
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"row {row_idx}: cannot parse outflow value {value_cell!r}"
                ),
            ))
            continue

        ad_month = date_cell.month
        ad_year = date_cell.year
        if ad_month not in _AD_MONTH_TO_BS:
            errors.append(ParserError(
                error_class="PeriodAmbiguous",
                error_detail=f"row {row_idx}: unexpected AD month {ad_month}",
                source_excerpt=str(date_cell),
            ))
            continue

        bs_month = _bs_month_from_ad_month(ad_month)
        bs_year = _bs_year_from_ad(ad_month, ad_year)

        staging_rows.append(_make_base_row(
            bs_month=bs_month,
            bs_year=bs_year,
            slug=SLUG_MIGRANT_DEPARTURES,
            value=value,
            unit="count",
            confidence="A",
            notes="Source: Department of Foreign Employment (dofe.gov.np); monthly departures",
        ))

    return staging_rows


# ── Tourist Arrivals parser ───────────────────────────────────────────────────

# Mapping AD calendar month name (column header) → AD month integer.
_TOURIST_MONTH_NAMES: Final[dict[str, int]] = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_tourist_arrivals(
    path: Path,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse Tourist-arrivals.xlsx → Tourist Arrival sheet.

    Pivot table: row = calendar year (integer), columns = Jan–Dec + Total.
    Row 0: title
    Row 1: headers ("Year", "Jan", "Feb", ..., "Dec", "Total")
    Rows 2+: data rows.

    Emits one StagingRowDraft per non-null monthly cell (skips Total column).
    """
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"Tourist-arrivals.xlsx open failed: {exc}",
        ))
        return []

    sheet_name = "Tourist Arrival"
    if sheet_name not in wb.sheetnames:
        # Try the first sheet.
        sheet_name = wb.sheetnames[0]
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail=(
                f"sheet 'Tourist Arrival' not found; using first sheet "
                f"'{sheet_name}' instead"
            ),
        ))

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail="Tourist-arrivals.xlsx has fewer than 2 rows",
        ))
        return []

    # Parse column headers from row 1.
    header = rows[1]
    # Map col_index → (ad_month_int | None for Year/Total).
    col_month: dict[int, int] = {}
    year_col: int | None = None
    for i, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key == "year":
            year_col = i
        elif key in _TOURIST_MONTH_NAMES:
            col_month[i] = _TOURIST_MONTH_NAMES[key]
        # "total" column is intentionally skipped.

    if year_col is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=f"'Year' column not found in header: {header!r}",
        ))
        return []

    if not col_month:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="no month columns (Jan–Dec) found in Tourist-arrivals.xlsx",
        ))
        return []

    staging_rows: list[StagingRowDraft] = []

    for row_idx, row in enumerate(rows[2:], start=2):
        year_cell = row[year_col] if year_col < len(row) else None
        if year_cell is None:
            continue
        try:
            ad_year = int(year_cell)
        except (TypeError, ValueError):
            continue
        if ad_year < 1950 or ad_year > 2100:
            continue

        for col_i, ad_month in col_month.items():
            if col_i >= len(row):
                continue
            cell = row[col_i]
            if cell is None:
                continue
            if isinstance(cell, str):
                # Formula string — skip (data_only should resolve these).
                continue
            try:
                value = float(cell)
            except (TypeError, ValueError):
                errors.append(ParserError(
                    error_class="ValueUnparseable",
                    error_detail=(
                        f"row {row_idx}: cannot parse tourist value "
                        f"{cell!r} for year={ad_year} month={ad_month}"
                    ),
                ))
                continue

            if ad_month not in _AD_MONTH_TO_BS:
                errors.append(ParserError(
                    error_class="PeriodAmbiguous",
                    error_detail=(
                        f"row {row_idx}: unexpected AD month {ad_month}"
                    ),
                ))
                continue

            bs_month = _bs_month_from_ad_month(ad_month)
            bs_year = _bs_year_from_ad(ad_month, ad_year)

            staging_rows.append(_make_base_row(
                bs_month=bs_month,
                bs_year=bs_year,
                slug=SLUG_TOURIST_ARRIVALS,
                value=value,
                unit="count",
                confidence="A",
                notes=(
                    f"AD calendar year {ad_year}, month {ad_month}; "
                    "source: NRB Tourist Arrival pivot table"
                ),
            ))

    return staging_rows


# ── BOP BPM6 Workers Remittance parser ───────────────────────────────────────

# FY label suffix patterns (e.g. "2022/23R", "2025/26P").
_FY_LABEL_RE: Final[re.Pattern[str]] = re.compile(
    r"^(\d{4})/(\d{2,4})[A-Z]*$"
)

# Month name → AD month integer for BOP column headers.
_BOP_MONTH_NAMES: Final[dict[str, int]] = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "june": 6, "july": 7,
}


def _parse_bop_remittance(
    path: Path,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse Balance-of-Payments-BPM6.xlsx → BOP BPM6 sheet.

    Column structure (row indices 2–4):
        Row 2: FY labels ("2022/23R", None, ..., "2023/24R", ...) at every 36th col.
        Row 3: Month names ("Aug", None, None, "Sep", ...) at every 3rd col within a FY.
        Row 4: ("Credit ", "Debit", "Net") repeating.

    Data is cumulative within each fiscal year. Each (FY, month) triple gives
    Credit (col 0), Debit (col 1), Net (col 2).

    Workers' remittances are at S.N. 1.C.2.1.1 (row index 55 in the fixture,
    but we search by S.N. value to be layout-stable).

    Emits one row per (FY, month) for the Credit value (NPR million inflow).
    """
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"Balance-of-Payments-BPM6.xlsx open failed: {exc}",
        ))
        return []

    sheet_name = "BOP BPM6"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail=(
                f"sheet 'BOP BPM6' not found; using first sheet "
                f"'{sheet_name}' instead"
            ),
        ))

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail="Balance-of-Payments-BPM6.xlsx has fewer than 5 rows",
        ))
        return []

    fy_row = rows[2]
    month_row = rows[3]

    # Build a map of col_index → (fy_ad_start, ad_month) for Credit columns only.
    # FY label columns mark the start of a new fiscal year block.
    # Each block: FY_col, then repeating triplets (Credit, Debit, Net) per month.
    col_to_period: dict[int, tuple[int, int]] = {}  # col → (fy_ad_start, ad_month)

    cur_fy_ad_start: int | None = None
    for col_i, fy_val in enumerate(fy_row):
        if fy_val is None or col_i < 2:
            continue
        m = _FY_LABEL_RE.match(str(fy_val).strip())
        if m:
            cur_fy_ad_start = int(m.group(1))

    # Re-iterate together with month row to build the per-column map.
    cur_fy_ad_start = None
    for col_i in range(len(fy_row)):
        fy_val = fy_row[col_i] if col_i < len(fy_row) else None
        if fy_val is not None and col_i >= 2:
            m = _FY_LABEL_RE.match(str(fy_val).strip())
            if m:
                cur_fy_ad_start = int(m.group(1))

        month_val = month_row[col_i] if col_i < len(month_row) else None
        if month_val is not None and cur_fy_ad_start is not None:
            month_key = str(month_val).strip().lower()
            if month_key in _BOP_MONTH_NAMES:
                ad_month = _BOP_MONTH_NAMES[month_key]
                # Credit column is at this exact col_i; Debit at col_i+1; Net at col_i+2.
                col_to_period[col_i] = (cur_fy_ad_start, ad_month)

    if not col_to_period:
        errors.append(ParserError(
            error_class="PageLayoutChanged",
            error_detail=(
                "BOP BPM6: could not build column→period map from rows 2–3; "
                "layout may have changed"
            ),
        ))
        return []

    # Find the workers' remittances row by S.N. = "1.C.2.1.1".
    workers_row: tuple | None = None
    workers_sn_target = "1.C.2.1.1"
    for data_row in rows[5:]:
        sn = data_row[0] if data_row else None
        if sn is not None and str(sn).strip() == workers_sn_target:
            workers_row = data_row
            break

    if workers_row is None:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"BOP BPM6: row S.N. '{workers_sn_target}' "
                "(O/W Workers' remittances) not found"
            ),
        ))
        return []

    staging_rows: list[StagingRowDraft] = []

    for credit_col, (fy_ad_start, ad_month) in col_to_period.items():
        if credit_col >= len(workers_row):
            continue
        cell = workers_row[credit_col]
        if cell is None:
            continue
        if isinstance(cell, str):
            # Formula string — should not happen with data_only=True.
            continue
        try:
            value = float(cell)
        except (TypeError, ValueError):
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"BOP BPM6: cannot parse workers remittance Credit "
                    f"{cell!r} at FY_AD_start={fy_ad_start} month={ad_month}"
                ),
            ))
            continue

        if ad_month not in _AD_MONTH_TO_BS:
            continue

        # Derive the actual AD year from the FY label and month.
        # NRB BOP FY "YYYY/YY" starts in Aug of AD year YYYY.
        # Months Aug(8)–Dec(12) are in AD year fy_ad_start.
        # Months Jan(1)–Jul(7) are in AD year fy_ad_start + 1.
        if ad_month >= 8:
            ad_year_actual = fy_ad_start
        else:
            ad_year_actual = fy_ad_start + 1

        bs_month = _bs_month_from_ad_month(ad_month)
        bs_year = _bs_year_from_ad(ad_month, ad_year_actual)
        fy_start = _fy_bs_start_from_bs_month_year(bs_month, bs_year)
        fy_bs = fiscal_year_label(fy_start)
        fy_ad = fiscal_year_ad_label(fy_start)
        period_bs = f"{fy_bs} {bs_month}"
        mid = mid_month_ad(bs_month, bs_year)
        pub_ad = mid + timedelta(days=_PUB_LAG_DAYS)

        staging_rows.append(StagingRowDraft(
            indicator_slug_raw=SLUG_BOP_REMITTANCE,
            value=value,
            unit="npr_million",
            reporting_period_type="year_to_date",
            reporting_period_bs=period_bs,
            reporting_period_ad_start=mid,
            reporting_period_ad_end=mid,
            publication_date_ad=pub_ad,
            publication_date_bs=f"~{fy_bs} (heuristic)",
            fiscal_year_bs=fy_bs,
            fiscal_year_ad_label=fy_ad,
            confidence_grade_proposed="A",
            parser_notes=(
                "BOP BPM6 S.N. 1.C.2.1.1 Workers remittances Credit — "
                "cumulative within FY; NPR million"
            ),
        ))

    return staging_rows


# ── Top-level parse() ─────────────────────────────────────────────────────────


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse the three NRB External Sector XLSX files.

    ``source_document_path`` must be the path to one of:
        - MIgrant-Workers_.xlsx
        - Tourist-arrivals.xlsx
        - Balance-of-Payments-BPM6.xlsx

    The parser auto-detects the file by filename and routes to the
    appropriate sub-parser. For a full ingestion run, call this function
    three times (once per file).

    Returns:
        ParserResult with status "success" / "partial" / "failure",
        staging_rows, and errors.
    """
    _ = source_document_id

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source file not found: {path}",
            )],
        )

    name_lower = path.name.lower()
    errors: list[ParserError] = []
    staging_rows: list[StagingRowDraft] = []

    if "migrant" in name_lower:
        staging_rows = _parse_migrant_workers(path, errors)
        # Add ColumnMissing for unavailable slugs (remittance corridors).
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"{SLUG_REMITTANCE_INDIA}: India migrant-worker departures are "
                "not tracked by DOFE (no labor permit required); remittance NPR "
                "by corridor is not available in this source"
            ),
        ))
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail=(
                f"{SLUG_REMITTANCE_GULF}: Gulf-corridor remittance NPR is not "
                "available in this source; departure headcounts exist per country "
                "but not remittance flows"
            ),
        ))

    elif "tourist" in name_lower:
        staging_rows = _parse_tourist_arrivals(path, errors)

    elif "balance" in name_lower or "bop" in name_lower or "bpm6" in name_lower:
        staging_rows = _parse_bop_remittance(path, errors)

    else:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=(
                    f"unrecognised filename '{path.name}'; expected one of: "
                    "MIgrant-Workers_.xlsx, Tourist-arrivals.xlsx, "
                    "Balance-of-Payments-BPM6.xlsx"
                ),
            )],
        )

    if not staging_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors or [ParserError(
                error_class="PageLayoutChanged",
                error_detail="no staging rows produced",
            )],
        )

    # "partial" if any non-ColumnMissing errors (real parse failures); the
    # ColumnMissing errors for unavailable corridors are expected, not failures.
    hard_errors = [
        e for e in errors
        if e.error_class != "ColumnMissing"
    ]
    status: ParserStatus = "partial" if hard_errors else "success"

    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    print(json.dumps({
        "status": result.status,
        "parser_version": result.parser_version,
        "staging_rows": [r.to_json_dict() for r in result.staging_rows],
        "errors": [
            {
                "error_class": e.error_class,
                "error_detail": e.error_detail,
            }
            for e in result.errors
        ],
    }))

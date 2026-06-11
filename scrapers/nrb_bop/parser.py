"""NRB Balance of Payments — historical BPM5 back-series parser (deterministic).

Source id: ``nrb-bop``.

File:
    ``Financial Data/nrb_dne_historical/Trade-and-Balance-of-Payments.xlsx``
    Sheet: ``BOP 2000-``

Layout:
    The sheet uses a **two-panel** side-by-side format:
      Panel 1 — header row: cols 3–14  (years 2000/01 → 2011/12)
      Panel 2 — header row: cols 18–29 (years 2012/13 → 2023/24P)
    Year labels carry optional revision suffixes: ``R`` (revised), ``P`` (provisional).
    Values are in NPR million ("In Million Rupees").

What is promoted:
    A single series: **``remittance-inflow-bpm5``** (Workers' remittances row).
    No other BoP lines are promoted (ADR-0014: no catalogue pollution).

Methodology note (DATA CONTINUITY PROTOCOL):
    This series uses **BPM5 methodology**.  NRB adopted BPM6 from approximately
    FY2069/70 (AD 2012/13).  The two series are **NOT directly comparable**.
    Every staging row carries an explicit ``parser_notes`` field marking the
    methodology so the UI layer can render the break.  Do NOT splice this series
    onto ``dne-remittance-inflow`` (BPM6, source ``nrb-dne-xlsx``).

ADR: ADR-0003 — no LLM / AI calls.  Pure file-in → dataclass-out.

Version history:
    0.1.0 — initial: Workers' remittances BPM5 annual series FY2000/01→FY2023/24P.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import fiscal_year_ad_label, fiscal_year_label
from _common.types import ParserError, ParserResult, ParserStatus, StagingRowDraft

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-bop"

# The only indicator promoted from this file (ADR-0014: no catalogue pollution).
_INDICATOR_SLUG: Final[str] = "remittance-inflow-bpm5"
_INDICATOR_UNIT: Final[str] = "npr_million"
_CONFIDENCE: Final[str] = "B"

# Target row label (exact match after normalisation).
_TARGET_LABEL: Final[str] = "workers' remittances"

# Sheet name.
_SHEET: Final[str] = "BOP 2000-"

# Regex: AD fiscal-year label like "2000/01", "2021/22R", "2023/24P".
# Group 1 = 4-digit start year, Group 2 = 2-digit tail, Group 3 = suffix (R/P/E/blank).
_FY_RE: Final = re.compile(
    r"^\s*(\d{4})\s*/\s*(\d{2})\s*([RPEQrpeq]?)\s*$"
)

# Approximate Nepali fiscal year boundaries expressed as AD calendar dates.
# Nepal FY runs Shrawan 1 (≈ mid-July) to Ashadh 30 (≈ mid-July next year).
# Mid-month approximation — TS validation layer refines to exact BS-calendar boundaries.
_FY_START_AD_MONTH: Final[int] = 7   # July  (Shrawan)
_FY_START_AD_DAY: Final[int] = 16
_FY_END_AD_MONTH: Final[int] = 7    # July
_FY_END_AD_DAY: Final[int] = 15

# Methodology break annotation carried on every staging row.
_METHODOLOGY_NOTE_TEMPLATE: Final[str] = (
    "BPM5 methodology. "
    "NRB adopted BPM6 from ~FY2069/70 (AD2012/13). "
    "Series not directly comparable to dne-remittance-inflow (BPM6). "
    "Any chart joining both series must show a break at FY2069/70. "
    "Revision suffix: {suffix}. "
    "Source: Trade-and-Balance-of-Payments.xlsx sheet 'BOP 2000-'."
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _norm(val: object) -> str:
    """Normalise a cell value to a lowercase stripped string."""
    return str(val).strip().lower() if val is not None else ""


def _parse_fy_label(raw: object) -> tuple[int, str] | None:
    """Return (ad_start_year, suffix) for a recognised FY label cell, else None.

    ``suffix`` is 'R', 'P', 'E', or '' (empty string = final/no-suffix).
    """
    text = _norm(raw)
    m = _FY_RE.match(text)
    if m is None:
        return None
    return int(m.group(1)), m.group(3).upper()


def _ad_fy_to_bs(ad_start: int) -> tuple[str, str]:
    """Return (bs_fy_label, ad_fy_label) from an AD fiscal-year start year.

    E.g. ad_start=2000 → ('2057/58', '2000/01').
    The BS start year is ad_start + 57 (lightweight mid-year approximation; the
    TS validation layer refines — see docs/CALENDAR_AND_PERIODS.md).
    """
    bs_start = ad_start + 57
    return fiscal_year_label(bs_start), fiscal_year_ad_label(bs_start)


def _fy_period_datetimes(ad_start: int) -> tuple[datetime, datetime]:
    """Approximate AD start/end datetimes for a Nepal fiscal year.

    Start: July 16 of ad_start  (Shrawan ~1)
    End:   July 15 of ad_start+1 (Ashadh ~30)
    """
    start = datetime(ad_start, _FY_START_AD_MONTH, _FY_START_AD_DAY, tzinfo=UTC)
    end = datetime(ad_start + 1, _FY_END_AD_MONTH, _FY_END_AD_DAY, tzinfo=UTC)
    return start, end


# ---------------------------------------------------------------------------
# Core parse logic
# ---------------------------------------------------------------------------


def _build_col_map(
    rows: list[tuple[object, ...]],
) -> dict[int, tuple[int, str]]:
    """Scan rows for FY header labels; return {col_index: (ad_start_year, suffix)}.

    The sheet uses a two-panel layout.  We scan every row for cells that parse
    as FY labels and collect all (col, parsed_fy) mappings.  The first row that
    yields ≥ 2 distinct years is taken as the definitive header row.
    """
    for row in rows:
        mapping: dict[int, tuple[int, str]] = {}
        for col, val in enumerate(row):
            parsed = _parse_fy_label(val)
            if parsed is not None:
                mapping[col] = parsed
        if len(mapping) >= 2:
            return mapping
    return {}


def _find_target_row(
    rows: list[tuple[object, ...]],
    col_map: dict[int, tuple[int, str]],
) -> tuple[object, ...] | None:
    """Return the first row whose label columns contain the Workers' remittances label."""
    # Scan col 0, 1, and 2 (the label cells in both panels) for each data row.
    label_cols = {0, 1, 2, 15, 16, 17}
    for row in rows:
        for c in label_cols:
            if c < len(row) and _norm(row[c]) == _TARGET_LABEL:
                return row
    return None


def _parse_sheet(
    rows: list[tuple[object, ...]],
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Extract Workers' remittances staging rows from the BOP 2000- sheet."""
    col_map = _build_col_map(rows)
    if not col_map:
        return [], [
            ParserError(
                error_class="PageLayoutChanged",
                error_detail="No fiscal-year header row found in sheet 'BOP 2000-'.",
            )
        ]

    target_row = _find_target_row(rows, col_map)
    if target_row is None:
        return [], [
            ParserError(
                error_class="ColumnMissing",
                error_detail=(
                    f"Row labelled '{_TARGET_LABEL}' not found in sheet 'BOP 2000-'."
                ),
            )
        ]

    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []

    pub_date = datetime(1970, 1, 1, tzinfo=UTC)  # sentinel — no embedded pub date

    for col, (ad_start, suffix) in sorted(col_map.items()):
        # Skip if the value column is out of bounds or non-numeric.
        if col >= len(target_row):
            continue
        raw_val = target_row[col]
        if raw_val is None:
            continue
        try:
            value = float(raw_val)
        except (TypeError, ValueError):
            errors.append(
                ParserError(
                    error_class="ValueUnparseable",
                    error_detail=(
                        f"Non-numeric value {raw_val!r} in Workers' remittances "
                        f"col {col} (FY {ad_start}/{(ad_start+1)%100:02d})."
                    ),
                )
            )
            continue

        bs_fy, ad_fy = _ad_fy_to_bs(ad_start)
        period_start, period_end = _fy_period_datetimes(ad_start)

        note = _METHODOLOGY_NOTE_TEMPLATE.format(
            suffix=suffix if suffix else "none"
        )

        staging.append(
            StagingRowDraft(
                indicator_slug_raw=_INDICATOR_SLUG,
                value=value,
                unit=_INDICATOR_UNIT,
                reporting_period_type="annual",
                reporting_period_bs=bs_fy,
                reporting_period_ad_start=period_start,
                reporting_period_ad_end=period_end,
                publication_date_ad=pub_date,
                publication_date_bs="unknown",
                fiscal_year_bs=bs_fy,
                fiscal_year_ad_label=ad_fy,
                confidence_grade_proposed=_CONFIDENCE,
                parser_notes=note,
            )
        )

    return staging, errors


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse(file_path: str, source_document_id: str) -> ParserResult:
    """Parse the NRB historical BPM5 BoP xlsx → Workers' remittances staging rows.

    ``file_path`` must point to ``Trade-and-Balance-of-Payments.xlsx``.
    ``source_document_id`` is the FK into ``source_documents`` supplied by the
    orchestration layer.

    Returns a ``ParserResult``.  Never raises: file errors become typed
    ``ParserError`` entries with ``status='failure'``.
    """
    path = Path(file_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=f"File not found: {file_path}",
                )
            ],
        )

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=[
                ParserError(
                    error_class="EncodingError",
                    error_detail=f"openpyxl could not open {path.name}: {exc}",
                )
            ],
        )

    if _SHEET not in wb.sheetnames:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=(
                        f"Sheet '{_SHEET}' not found in {path.name}. "
                        f"Available: {wb.sheetnames}"
                    ),
                )
            ],
        )

    rows = list(wb[_SHEET].iter_rows(values_only=True))
    staging, errors = _parse_sheet(rows)

    if not staging and not errors:
        status: ParserStatus = "partial"
        errors = [
            ParserError(
                error_class="Other",
                error_detail="No staging rows extracted; sheet may be empty.",
            )
        ]
    elif errors and not staging:
        status = "failure"
    elif errors:
        status = "partial"
    else:
        status = "success"

    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------


def _main() -> None:  # pragma: no cover
    """CLI: ``python -m nrb_bop <file_path> [--source-doc-id <id>]``"""
    import argparse

    ap = argparse.ArgumentParser(
        description="Parse NRB historical BPM5 BoP xlsx → Workers' remittances staging rows."
    )
    ap.add_argument("file_path", help="Path to Trade-and-Balance-of-Payments.xlsx")
    ap.add_argument(
        "--source-doc-id",
        default="nrb-bop-local",
        help="source_documents FK (default: nrb-bop-local)",
    )
    args = ap.parse_args()

    result = parse(args.file_path, source_document_id=args.source_doc_id)
    import dataclasses

    print(json.dumps(dataclasses.asdict(result), indent=2, default=str))
    sys.exit(0 if result.status == "success" else 1)


if __name__ == "__main__":
    _main()

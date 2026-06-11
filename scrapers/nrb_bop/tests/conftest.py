"""Pytest fixtures for the NRB BOP parser.

Builds a synthetic XLSX fixture via openpyxl — regenerated each run, not
committed as a binary.  The fixture mimics the two-panel BOP 2000- layout:
  - Panel 1: cols 3–14, years 2057/58–2068/69 (AD 2000/01–2011/12)
  - Panel 2: cols 18–29, years 2069/70–2078/79R, 2079/80R, 2080/81P
             (AD 2012/13–2021/22, 2022/23R, 2023/24P)

Workers' remittances row uses known test values so tests can assert exact amounts.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"

# Known test values for Workers' remittances (NPR million), one per FY column.
# Panel 1: AD 2000/01 through 2011/12 (12 values, matching row 33 of the real file)
_PANEL1_FY_LABELS = [
    "2000/01", "2001/02", "2002/03", "2003/04", "2004/05", "2005/06",
    "2006/07", "2007/08", "2008/09", "2009/10", "2010/11", "2011/12",
]
_PANEL1_VALUES = [
    47216.1, 47536.3, 54203.3, 58587.6, 65541.2, 97688.5,
    100144.8, 142682.7, 209698.5, 231725.3, 253551.6, 359554.4,
]

# Panel 2: AD 2012/13 through 2023/24P (12 values)
_PANEL2_FY_LABELS = [
    "2012/13", "2013/14", "2014/15", "2015/16", "2016/17", "2017/18",
    "2018/19", "2019/20", "2020/21", "2021/22R", "2022/23R", "2023/24P",
]
_PANEL2_VALUES = [
    434581.7, 543294.1, 617278.8, 665064.3,
    695452.4, 755058.6, 879271.4, 875026.9,
    961054.6, 1007306.9, 1240686.4, 1445315.1,
]


def _build_bop_fixture(path: Path) -> None:
    """Build a synthetic Trade-and-Balance-of-Payments.xlsx with BOP 2000- sheet."""
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BOP 2000-"

    # Row 1: title (row index 0 → openpyxl row 1)
    ws.cell(row=1, column=1, value="Summary of Balance of Payments")

    # Row 4: unit annotation (row index 3 → openpyxl row 4)
    ws.cell(row=4, column=1, value="In Million Rupees")

    # Row 6: header row (row index 5 → openpyxl row 6)
    # Panel 1: cols 1–3 = labels, cols 4–15 = years 2000/01–2011/12
    ws.cell(row=6, column=1, value="Particulars")
    for i, label in enumerate(_PANEL1_FY_LABELS):
        ws.cell(row=6, column=4 + i, value=label)  # cols 4–15 (openpyxl 1-indexed)
    # Panel 2: cols 16–18 = labels, cols 19–30 = years 2012/13–2023/24P
    ws.cell(row=6, column=16, value="Particulars")
    for i, label in enumerate(_PANEL2_FY_LABELS):
        ws.cell(row=6, column=19 + i, value=label)  # cols 19–30

    # ---- Other BoP rows (should NOT be promoted) ----
    # Current Account
    ws.cell(row=7, column=1, value="A. Current Account")
    for i in range(len(_PANEL1_FY_LABELS)):
        ws.cell(row=7, column=4 + i, value=10000.0 + i * 100)
    for i in range(len(_PANEL2_FY_LABELS)):
        ws.cell(row=7, column=19 + i, value=50000.0 + i * 100)

    # Goods exports
    ws.cell(row=8, column=1, value="Goods: exports f.o.b.")
    for i in range(len(_PANEL1_FY_LABELS)):
        ws.cell(row=8, column=4 + i, value=60000.0 + i * 500)
    for i in range(len(_PANEL2_FY_LABELS)):
        ws.cell(row=8, column=19 + i, value=90000.0 + i * 500)

    # ---- Workers' remittances row ----
    # Panel 1: label in col 2 (openpyxl), values in cols 4–15
    ws.cell(row=34, column=2, value="Workers' remittances")
    for i, val in enumerate(_PANEL1_VALUES):
        ws.cell(row=34, column=4 + i, value=val)
    # Panel 2: label in col 17, values in cols 19–30
    ws.cell(row=34, column=17, value="Workers' remittances")
    for i, val in enumerate(_PANEL2_VALUES):
        ws.cell(row=34, column=19 + i, value=val)

    wb.save(str(path))


@pytest.fixture(scope="session")
def bop_xlsx() -> Path:
    """Synthetic Trade-and-Balance-of-Payments.xlsx with BOP 2000- sheet."""
    p = FIXTURE_DIR / "Trade-and-Balance-of-Payments.xlsx"
    if not p.exists():
        _build_bop_fixture(p)
    return p


@pytest.fixture(scope="session")
def missing_sheet_xlsx() -> Path:
    """Workbook that lacks the 'BOP 2000-' sheet."""
    p = FIXTURE_DIR / "missing_sheet.xlsx"
    if not p.exists():
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="no bop sheet here")
        wb.save(str(p))
    return p


@pytest.fixture(scope="session")
def no_remittances_row_xlsx() -> Path:
    """BOP 2000- sheet that has the header row but no Workers' remittances row."""
    p = FIXTURE_DIR / "no_remittances_row.xlsx"
    if not p.exists():
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "BOP 2000-"
        ws.cell(row=6, column=1, value="Particulars")
        for i, label in enumerate(_PANEL1_FY_LABELS):
            ws.cell(row=6, column=4 + i, value=label)
        for i, label in enumerate(_PANEL2_FY_LABELS):
            ws.cell(row=6, column=19 + i, value=label)
        # no Workers' remittances row
        ws.cell(row=7, column=1, value="A. Current Account")
        ws.cell(row=7, column=4, value=12345.0)
        wb.save(str(p))
    return p

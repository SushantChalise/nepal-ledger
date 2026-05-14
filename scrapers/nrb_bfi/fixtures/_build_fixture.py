"""Build a minimal NRB BFI fixture XLSX for unit tests.

Run from the repo root::

    python scrapers/nrb_bfi/fixtures/_build_fixture.py

Produces ``Saun-2082-fixture.xlsx`` next to this script. Tracked in git so
test runs don't depend on the large raw corpus.
"""

from pathlib import Path

from openpyxl import Workbook


def build() -> None:  # noqa: PLR0915 — single-purpose builder, linear layout
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ─── C2 stub (table of contents) ───────────────────────────────────
    c2 = wb.create_sheet("C2")
    c2["A1"] = "Table of Contents"

    # ─── C4 Major Financial Indicators ─────────────────────────────────
    c4 = wb.create_sheet("C4")
    c4["B2"] = "Major Financial Indicators"
    c4["B3"] = "as on Saun End, 2082 (Mid-Aug, 2025)"
    c4["D4"] = 'Class "A"'
    c4["E4"] = 'Class "B"'
    c4["F4"] = 'Class "C"'
    c4["G4"] = "Overall"
    c4["B5"] = "A.  Credit, Deposit Ratios (%)"
    c4["B6"] = 1
    c4["C6"] = "Total Deposit/GDP"
    c4["D6"] = 106.21
    c4["E6"] = 10.20
    c4["F6"] = 2.16
    c4["G6"] = 118.57
    c4["B7"] = 2
    c4["C7"] = "Total Credit/GDP"
    c4["D7"] = 81.10
    c4["E7"] = 8.53
    c4["F7"] = 1.68
    c4["G7"] = 91.31
    c4["B8"] = "B.  Liquidity Ratios (%)"
    c4["B9"] = 1
    c4["C9"] = "Total Liquid Assets/Total Deposit"
    c4["D9"] = 30.5
    c4["E9"] = 25.1
    c4["F9"] = 20.0
    c4["G9"] = 28.9

    # ─── C5 Assets & Liabilities (4-block) ─────────────────────────────
    c5 = wb.create_sheet("C5")
    c5["D2"] = (
        "Statement of  Assets and Liabilities of Banks & Financial "
        "Institutions"
    )
    c5["L2"] = "Statement of  Assets and Liabilities of Commercial Banks"
    c5["T2"] = "Statement of  Assets and Liabilities of Development Banks"
    c5["AB2"] = "Statement of  Assets and Liabilities of Finance Companies"
    c5["J3"] = "Amt in Mn of Rs"
    c5["R3"] = "Amt in Mn of Rs"
    c5["Z3"] = "Amt in Mn of Rs"
    c5["AH3"] = "Amt in Mn of Rs"
    for col in ("D", "L", "T", "AB"):
        c5[f"{col}4"] = "Mid-July "
    c5["G4"] = "Mid-July"
    c5["H4"] = "Mid-Aug"
    c5["I4"] = "% Change"
    c5["O4"] = "Mid-July"
    c5["P4"] = "Mid-Aug"
    c5["Q4"] = "% Change"
    c5["W4"] = "Mid-July"
    c5["X4"] = "Mid-Aug"
    c5["Y4"] = "% Change"
    c5["AE4"] = "Mid-July"
    c5["AF4"] = "Mid-Aug"
    c5["AG4"] = "% Change"
    years = (("D", "E", "F", "G", "H"),
             ("L", "M", "N", "O", "P"),
             ("T", "U", "V", "W", "X"),
             ("AB", "AC", "AD", "AE", "AF"))
    for block in years:
        for col, year in zip(block, (2022, 2023, 2024, 2025, 2025), strict=True):
            c5[f"{col}5"] = year
        for col, n in zip(block, (1, 2, 3, 4, 5), strict=True):
            c5[f"{col}6"] = n
    # Section: CAPITAL FUND
    c5["B7"] = 1
    c5["C7"] = "CAPITAL FUND"
    c5["G7"] = 775394.54
    c5["H7"] = 840143.59
    c5["O7"] = 697777.94
    c5["P7"] = 754999.44
    c5["W7"] = 60146.16
    c5["X7"] = 66650.85
    c5["AE7"] = 17470.44
    c5["AF7"] = 18493.30
    # Sub: a. Paid-up Capital
    c5["B8"] = " "
    c5["C8"] = "a. Paid-up Capital"
    c5["G8"] = 443682.13
    c5["H8"] = 444576.20
    c5["O8"] = 385327.20
    c5["P8"] = 386252.25
    c5["W8"] = 43054.77
    c5["X8"] = 43023.79
    c5["AE8"] = 15300.16
    c5["AF8"] = 15300.16
    # Sub: b. Statutory Reserves
    c5["C9"] = "b. Statutory Reserves"
    c5["G9"] = 100.0
    c5["H9"] = 110.0
    c5["O9"] = 90.0
    c5["P9"] = 95.0
    c5["W9"] = 5.0
    c5["X9"] = 6.0
    c5["AE9"] = 2.0
    c5["AF9"] = 3.0

    # ─── C6 P&L ─────────────────────────────────────────────────────────
    c6 = wb.create_sheet("C6")
    c6["C2"] = (
        "Profit and Loss Statement of Banks & Financial Institutions"
    )
    c6["K2"] = "Profit and Loss Statement of Commercial Banks"
    c6["S2"] = "Profit and Loss Statement of Development Banks"
    c6["AA2"] = "Profit and Loss Statement of Finance Companies"
    c6["I3"] = "Amt in Mn of Rs"
    c6["C4"] = "Mid-July "
    c6["F4"] = "Mid-July"
    c6["G4"] = "Mid-Aug"
    c6["H4"] = "% Change"
    c6["K4"] = "Mid-July "
    c6["N4"] = "Mid-July"
    c6["O4"] = "Mid-Aug"
    c6["P4"] = "% Change"
    c6["S4"] = "Mid-July "
    c6["V4"] = "Mid-July"
    c6["W4"] = "Mid-Aug"
    c6["X4"] = "% Change"
    c6["AA4"] = "Mid-July "
    c6["AD4"] = "Mid-July"
    c6["AE4"] = "Mid-Aug"
    c6["AF4"] = "% Change"
    c6_blocks = (("C", "D", "E", "F", "G"),
                 ("K", "L", "M", "N", "O"),
                 ("S", "T", "U", "V", "W"),
                 ("AA", "AB", "AC", "AD", "AE"))
    for block in c6_blocks:
        for col, year in zip(block, (2022, 2023, 2024, 2025, 2025), strict=True):
            c6[f"{col}5"] = year
        for col, n in zip(block, (1, 2, 3, 4, 5), strict=True):
            c6[f"{col}6"] = n
    c6["B7"] = " 1  Interest Expenses"
    c6["F7"] = 350266.34
    c6["G7"] = 26777.11
    c6["N7"] = 306737.53
    c6["O7"] = 23561.76
    c6["V7"] = 34592.64
    c6["W7"] = 2552.91
    c6["AD7"] = 8936.16
    c6["AE7"] = 662.45
    c6["B8"] = "      1.1  Deposit Liabilities"
    c6["F8"] = 100.0
    c6["G8"] = 110.0
    c6["N8"] = 90.0
    c6["O8"] = 95.0
    c6["V8"] = 8.0
    c6["W8"] = 9.0
    c6["AD8"] = 2.0
    c6["AE8"] = 3.0

    # ─── C7 Sectorwise Loans & Advances ────────────────────────────────
    c7 = wb.create_sheet("C7")
    c7["C2"] = (
        "Statement of Loans and Advances of Banks & Financial Institutions"
    )
    c7["K2"] = "Statement of Loans and Advances of Commercial Banks"
    c7["S2"] = "Statement of Loans and Advances of Development Banks"
    c7["AA2"] = "Statement of Loans and Advances of Finance Companies"
    c7["I3"] = "Amt in Mn of Rs"
    c7["C4"] = "Mid-July "
    c7["F4"] = "Mid-July"
    c7["G4"] = "Mid-Aug"
    c7["H4"] = "% Change"
    c7["K4"] = "Mid-July "
    c7["N4"] = "Mid-July"
    c7["O4"] = "Mid-Aug"
    c7["P4"] = "% Change"
    c7["S4"] = "Mid-July "
    c7["V4"] = "Mid-July"
    c7["W4"] = "Mid-Aug"
    c7["X4"] = "% Change"
    c7["AA4"] = "Mid-July "
    c7["AD4"] = "Mid-July"
    c7["AE4"] = "Mid-Aug"
    c7["AF4"] = "% Change"
    for block in c6_blocks:
        for col, year in zip(block, (2022, 2023, 2024, 2025, 2025), strict=True):
            c7[f"{col}5"] = year
    c7["B6"] = "Sectorwise"
    for block in c6_blocks:
        for col, n in zip(block, (1, 2, 3, 4, 5), strict=True):
            c7[f"{col}6"] = n
    c7["B7"] = "Agricultural and Forest Related"
    c7["F7"] = 356370.47
    c7["G7"] = 352479.81
    c7["N7"] = 318651.22
    c7["O7"] = 313843.87
    c7["V7"] = 30557.99
    c7["W7"] = 31549.56
    c7["AD7"] = 7161.26
    c7["AE7"] = 7086.38
    c7["B8"] = "Construction"
    c7["F8"] = 200000.0
    c7["G8"] = 205000.0
    c7["N8"] = 180000.0
    c7["O8"] = 183000.0
    c7["V8"] = 15000.0
    c7["W8"] = 16000.0
    c7["AD8"] = 5000.0
    c7["AE8"] = 6000.0

    output = Path(__file__).parent / "Saun-2082-fixture.xlsx"
    wb.save(output)
    print(f"wrote {output}")


if __name__ == "__main__":
    build()

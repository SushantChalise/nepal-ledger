"""NRB BFI monthly XLSX parser — period derived from filename.

Source id: ``nrb-bfi-monthly-xlsx``.

Scope (this version):
    Parses sheet **C5** (Statement of Assets & Liabilities) for any BFI
    monthly publication XLSX whose filename encodes the BS month and year
    (e.g. ``Bhadau_2082_Publish.xlsx``, ``Chaitra_2082_Publish.xlsx``).
    Emits one row per (indicator, bank_class) for the latest snapshot column
    (Mid-Sept 2025 for Bhadau 2082; the column layout is constant across
    the corpus — see schema probe docs).

    Schema drift across the 49-month corpus is documented in
    ``docs/research/nrb-bfi-schema-probe.md`` (output of
    ``scrapers.nrb_bfi.schema_probe``). Follow-up parsers for remaining
    months are batched per ``docs/tasks/worker-P2-followup-bfi-batches.md``.

C5 layout (canonical month):
    Sheet has four side-by-side sub-tables, one per bank class. Each sub-table
    shares the descriptive label column (col index 2) and uses a fixed
    column stride of 8 for value columns. Mid-Sept (latest) value column
    indices are: BFI total -> 7, Commercial -> 15, Development -> 23,
    Finance -> 31. Numeric ordinal column (col 1) groups L/A side-of-balance
    rows; the descriptive label (col 2) is the indicator name.

Indicator slug convention:
    ``bfi-c5-<bank-class>-<slugified-label>`` (e.g.
    ``bfi-c5-commercial-deposits``). Bank-class is also emitted as a typed
    dimension via ``BankingSectorFactRow.bank_class``; encoding it in the slug
    keeps the slug unique across bank classes and makes the staging table
    legible.

Output:
    Emits ``BankingSectorFactRow`` dataclasses (NOT ``StagingRowDraft``) —
    the BFI corpus targets ``banking_sector_facts`` directly, not the
    indicator-values staging pipeline. The Node CLI
    ``scripts/ingest-bfi-monthly.ts`` consumes the JSON output via
    ``BankingSectorFactRow.to_json_dict()``.

Versioning:
    Bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Final, Literal

import openpyxl

from _common.periods import BS_MONTHS, BsMonth, fiscal_year_label, mid_month_ad
from _common.types import ParserError, ParserStatus

PARSER_VERSION: Final[str] = "0.2.0"
SOURCE_ID: Final[str] = "nrb-bfi-monthly-xlsx"

BankClass = Literal["commercial", "development", "finance", "system_total"]

# C5 sheet column layout (0-indexed). Mid-Sept (latest snapshot) value column
# per bank-class sub-table.
_LATEST_VALUE_COL_BY_CLASS: Final[dict[BankClass, int]] = {
    "system_total": 7,
    "commercial": 15,
    "development": 23,
    "finance": 31,
}
_LABEL_COL: Final[int] = 2

# C5 indicator labels we lift in v0.2.0. Hand-picked to anchor the canonical
# month: the rows we expect to find on every snapshot. Adding more rows is a
# parser-version bump; do not silently expand.
_C5_INDICATORS: Final[tuple[tuple[str, str], ...]] = (
    # (descriptive label in col 2, slug stem after bank-class)
    ("CAPITAL FUND", "capital-fund"),
    ("a. Paid-up Capital", "paid-up-capital"),
    ("b. Statutory Reserves", "statutory-reserves"),
    ("BORROWINGS", "borrowings-total"),
    ("DEPOSITS", "deposits-total"),
    ("a. Current", "deposits-current"),
    ("b. Savings", "deposits-savings"),
    ("c. Fixed", "deposits-fixed"),
    ("LIQUID FUNDS", "liquid-funds"),
)

_UNIT: Final[str] = "npr_million"

# ---------------------------------------------------------------------------
# BS month-name → canonical BsMonth map.
#
# Canonical spellings are the BsMonth literals from _common/periods.py:
#   Shrawan, Bhadra, Ashwin, Kartik, Mangsir, Poush, Magh, Falgun, Chait,
#   Baisakh, Jestha, Ashadh
#
# Keys are lowercase; all romanisation variants observed in the 59-file corpus
# are mapped here.  Every variant must map to exactly one BsMonth value.
# ---------------------------------------------------------------------------
_MONTH_ALIAS_TO_CANONICAL: Final[dict[str, BsMonth]] = {
    # 1 – Baisakh
    "baisakh": "Baisakh",
    "baishakh": "Baisakh",
    # 2 – Jestha
    "jestha": "Jestha",
    "jeth": "Jestha",
    "jestha2": "Jestha",  # e.g. "Jestha2079" (no separator)
    # 3 – Ashadh
    "asar": "Ashadh",
    "ashadh": "Ashadh",
    "ashar": "Ashadh",
    "ashadh2": "Ashadh",
    # 4 – Shrawan
    "saun": "Shrawan",
    "shrawan": "Shrawan",
    "shrawn": "Shrawan",
    "shrawan2": "Shrawan",
    # 5 – Bhadra
    "bhadau": "Bhadra",
    "bhadra": "Bhadra",
    # 6 – Ashwin
    "asoj": "Ashwin",
    "ashwin": "Ashwin",
    "ashoj": "Ashwin",
    "ashwin2": "Ashwin",
    # 7 – Kartik
    "kartik": "Kartik",
    "kartik2": "Kartik",
    # 8 – Mangsir
    "mangshir": "Mangsir",
    "mangsir": "Mangsir",
    "manghir": "Mangsir",
    # 9 – Poush
    "poush": "Poush",
    "push": "Poush",
    "poush2": "Poush",
    # 10 – Magh
    "magh": "Magh",
    "magh2": "Magh",
    # 11 – Falgun
    "falgun": "Falgun",
    "phalgun": "Falgun",
    "falgun2": "Falgun",
    # 12 – Chait
    "chaitra": "Chait",
    "chaitr": "Chait",
    "chaitra2": "Chait",
}

# Regex to extract the month token and 4-digit BS year from a filename.
# Handles common separators (_/-/none) and optional trailing noise such as
# version suffixes (-1, -2, V1, etc.).
# Examples matched:
#   Bhadau_2082_Publish.xlsx      -> ("Bhadau",  "2082")
#   Chaitra_2082_Publish.xlsx     -> ("Chaitra", "2082")
#   Saun-2082-Publish.xlsx        -> ("Saun",    "2082")
#   Shrawan_2081_Publish-1.xlsx   -> ("Shrawan", "2081")
#   Asoj_2082_Publish.xlsx        -> ("Asoj",    "2082")
#   Baisakh_2080_Publish-1.xlsx   -> ("Baisakh", "2080")
#   Ashwin2078_NFRS_Publish.xlsx  -> ("Ashwin",  "2078")
#   Jestha2079_PublishV1.xlsx     -> ("Jestha",  "2079")
#   Ashadh-2078-2.xlsx            -> ("Ashadh",  "2078")
#   Manghir-2078.xlsx             -> ("Manghir", "2078")
_FILENAME_RE: Final[re.Pattern[str]] = re.compile(
    r"^([A-Za-z]+)[_\-]?(20\d{2})",
    re.IGNORECASE,
)

# BS_MONTHS tuple from _common.periods is ordered Shrawan=1 … Ashadh=12.
# Build a fast canonical → month-number lookup (1-based).
_CANONICAL_TO_MONTH_NUM: Final[dict[BsMonth, int]] = {
    m: i + 1 for i, m in enumerate(BS_MONTHS)
}


def _parse_period_from_filename(filename: str) -> tuple[BsMonth, int] | ParserError:
    """Extract (canonical_bs_month, bs_year) from a BFI XLSX filename.

    Returns a ``ParserError`` (error_class="PeriodAmbiguous") when the
    filename does not match the expected pattern or the month token is
    unrecognised.  Never raises; never falls back silently.
    """
    stem = Path(filename).stem  # strip .xlsx / .xls
    m = _FILENAME_RE.match(stem)
    if not m:
        return ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"filename {filename!r} does not match expected pattern "
                r"<MonthName>[_-]?<YYYY> (e.g. Bhadau_2082_Publish.xlsx)"
            ),
            source_excerpt=filename,
        )

    raw_month = m.group(1).lower()
    bs_year = int(m.group(2))

    canonical = _MONTH_ALIAS_TO_CANONICAL.get(raw_month)
    if canonical is None:
        return ParserError(
            error_class="PeriodAmbiguous",
            error_detail=(
                f"unrecognised BS month token {m.group(1)!r} in filename "
                f"{filename!r}; add it to _MONTH_ALIAS_TO_CANONICAL"
            ),
            source_excerpt=m.group(1),
        )

    return canonical, bs_year


@dataclass(frozen=True)
class BankingSectorFactRow:
    """Python mirror of ``banking_sector_facts.$inferInsert`` minus FKs and
    server-side fields (``id``, ``source_document_id``, ``promoted_*``)
    which the Node ingest layer fills in.
    """

    bank_class: BankClass
    bank_entity_id: str | None  # null for class-aggregate rows
    source_sheet: str
    indicator_slug: str
    value: float
    unit: str
    reporting_period_type: Literal["monthly"]
    reporting_period_bs: str
    reporting_period_ad_start: datetime
    reporting_period_ad_end: datetime
    publication_date_ad: datetime
    publication_date_bs: str
    fiscal_year_bs: str
    confidence_grade: Literal["A", "B", "C"]
    parser_notes: str | None = None

    def to_json_dict(self) -> dict[str, Any]:
        out = asdict(self)
        for k in ("reporting_period_ad_start", "reporting_period_ad_end", "publication_date_ad"):
            out[k] = getattr(self, k).isoformat()
        return out


@dataclass(frozen=True)
class ParserResult:
    """Top-level result. Mirrors ``_common.types.ParserResult`` shape but
    carries ``BankingSectorFactRow`` rather than ``StagingRowDraft``.
    """

    status: ParserStatus
    parser_version: str
    fact_rows: list[BankingSectorFactRow] = field(default_factory=list)
    errors: list[ParserError] = field(default_factory=list)

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "parser_version": self.parser_version,
            "fact_rows": [r.to_json_dict() for r in self.fact_rows],
            "errors": [e.to_json_dict() for e in self.errors],
        }


def _load_c5_sheet(
    path: Path,
) -> tuple[list[tuple[object, ...]], dict[str, int]] | ParserError:
    """Open the XLSX, locate sheet C5, and return (rows, label_to_row_index).

    Returns a ``ParserError`` on any failure so the caller branch count stays
    within ruff's PLR0912 limit.
    """
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except (OSError, KeyError, ValueError) as exc:
        return ParserError(
            error_class="EncodingError",
            error_detail=f"openpyxl could not open {path.name}: {exc}",
        )

    if "C5" not in wb.sheetnames:
        return ParserError(
            error_class="PageLayoutChanged",
            error_detail=f"expected sheet C5 not present in {path.name}",
        )

    ws = wb["C5"]
    rows: list[tuple[object, ...]] = list(ws.iter_rows(values_only=True))

    label_to_row: dict[str, int] = {}
    for r_idx, row in enumerate(rows):
        if len(row) <= _LABEL_COL:
            continue
        lbl = _norm_label(row[_LABEL_COL])
        if lbl and lbl not in label_to_row:
            label_to_row[lbl] = r_idx

    return rows, label_to_row


def _safe_float(raw: object) -> float | None:
    """Coerce a cell value to float; reject NaN, empty, and non-numeric."""
    if raw is None:
        return None
    try:
        v = float(raw)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    if v != v:  # NaN; openpyxl can return float('nan')  # noqa: PLR0124
        return None
    return v


def _norm_label(raw: object) -> str:
    return "" if raw is None else " ".join(str(raw).split())


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse a BFI monthly XLSX, deriving the reporting period from the filename.

    Arguments:
        source_document_path: filesystem path to the XLSX.  The BS month and
            year are extracted from ``Path(source_document_path).name``.
        source_document_id: opaque ID (threaded through; not embedded in rows).

    Returns:
        ``ParserResult`` with ``status``, ``fact_rows``, ``errors``.
        Returns status="failure" with error_class="PeriodAmbiguous" when the
        filename cannot be parsed — never silently falls back to a hardcoded
        period.
    """
    _ = source_document_id

    path = Path(source_document_path)

    # --- Derive period from filename (fail fast if unparseable) ---
    # Done before the existence check so a bad filename is caught even on
    # a missing file (consistent error class for callers).
    period_result = _parse_period_from_filename(path.name)
    if isinstance(period_result, ParserError):
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[period_result],
        )

    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=f"source file not found: {path}",
                )
            ],
        )

    bs_month, bs_year = period_result

    # Canonical reporting_period_bs: "<CanonicalMonthName> <YYYY>".
    reporting_period_bs = f"{bs_month} {bs_year}"

    # AD span: mid-month approximation; refined downstream by validation layer.
    _mid = mid_month_ad(bs_month, bs_year)
    ad_start = datetime(_mid.year, _mid.month, 1, tzinfo=UTC)
    ad_end = datetime(_mid.year, _mid.month, 15, tzinfo=UTC)

    # Fiscal year: all 12 BS months of year Y belong to FY starting Y.
    # (Shrawan=month 1 … Ashadh=month 12, all within the same BS year label.)
    fy_bs = fiscal_year_label(bs_year)

    # Publication date is not derivable from the filename alone; use a
    # placeholder that the ingest layer can override via source-registry
    # metadata.  NRB typically publishes ~6 weeks after period close.
    # We approximate as the 15th of the AD month two months after period close.
    _pub_month_offset = _mid.month + 2
    _pub_year = _mid.year + (_pub_month_offset - 1) // 12
    _pub_month = ((_pub_month_offset - 1) % 12) + 1
    publication_date_ad = datetime(_pub_year, _pub_month, 15, tzinfo=UTC)
    publication_date_bs = f"{bs_year} {bs_month} (approx +6wk)"

    sheet_result = _load_c5_sheet(path)
    if isinstance(sheet_result, ParserError):
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[sheet_result],
        )
    rows, label_to_row = sheet_result

    fact_rows: list[BankingSectorFactRow] = []
    errors: list[ParserError] = []

    for label, slug_stem in _C5_INDICATORS:
        r_idx = label_to_row.get(label)
        if r_idx is None:
            errors.append(
                ParserError(
                    error_class="RegexMismatch",
                    error_detail=f"C5 label not found: {label!r}",
                    source_excerpt=label,
                )
            )
            continue
        row = rows[r_idx]
        for bank_class, col_idx in _LATEST_VALUE_COL_BY_CLASS.items():
            if col_idx >= len(row):
                errors.append(
                    ParserError(
                        error_class="ColumnMissing",
                        error_detail=(
                            f"row {r_idx} ({label!r}/{bank_class}): "
                            f"value column {col_idx} out of range (row len {len(row)})"
                        ),
                        source_excerpt=label,
                    )
                )
                continue
            value = _safe_float(row[col_idx])
            if value is None:
                errors.append(
                    ParserError(
                        error_class="ValueUnparseable",
                        error_detail=(
                            f"row {r_idx} ({label!r}/{bank_class}): "
                            f"could not parse {row[col_idx]!r} as float"
                        ),
                        source_excerpt=label,
                    )
                )
                continue
            fact_rows.append(
                BankingSectorFactRow(
                    bank_class=bank_class,
                    bank_entity_id=None,
                    source_sheet="C5",
                    indicator_slug=f"bfi-c5-{bank_class.replace('_', '-')}-{slug_stem}",
                    value=value,
                    unit=_UNIT,
                    reporting_period_type="monthly",
                    reporting_period_bs=reporting_period_bs,
                    reporting_period_ad_start=ad_start,
                    reporting_period_ad_end=ad_end,
                    publication_date_ad=publication_date_ad,
                    publication_date_bs=publication_date_bs,
                    fiscal_year_bs=fy_bs,
                    confidence_grade="A",
                    parser_notes=None,
                )
            )

    if not fact_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors
            or [
                ParserError(
                    error_class="Other",
                    error_detail="no recognised C5 rows found",
                )
            ],
        )

    status: ParserStatus = "partial" if errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        fact_rows=fact_rows,
        errors=errors,
    )


def _main() -> None:
    """CLI entrypoint used by ``scripts/ingest-bfi-monthly.ts``.

    Argv: ``parser.py <source_document_path> <source_document_id>``.
    Writes ``ParserResult.to_json_dict()`` to stdout.
    """
    import json
    import sys

    if len(sys.argv) != 3:
        sys.stderr.write("usage: parser.py <source_document_path> <source_document_id>\n")
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    json.dump(result.to_json_dict(), sys.stdout)


if __name__ == "__main__":
    _main()

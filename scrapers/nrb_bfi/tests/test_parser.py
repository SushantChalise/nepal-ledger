"""Tests for the NRB BFI monthly parser (v0.2.0).

The canonical-month fixture (bhadau_2082.xlsx) continues to pass; additional
tests verify that different filenames yield different, correct
reporting_period_bs and fiscal_year_bs values, and that an unparseable
filename returns a failure result.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from _common.types import ParserError
from nrb_bfi import PARSER_VERSION, parse
from nrb_bfi.parser import (
    _C5_INDICATORS,
    _C7_INDICATORS,
    _LATEST_VALUE_COL_BY_CLASS,
    BankingSectorFactRow,
    ParserResult,
    _parse_period_from_filename,
)

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
FIXTURE = FIXTURE_DIR / "bhadau_2082.xlsx"


def _build_fixture(path: Path, filename_stem: str | None = None) -> None:
    """Build a trimmed XLSX that matches the real C5 + C7 layout structurally.

    C5 layout (label in col 2, values at cols 7/15/23/31):
      system_total base=1000, commercial=700, development=200, finance=100.

    C7 layout (label in col 1, values at cols 6/14/22/30 — 0-indexed):
      system_total base=5000, commercial=4000, development=700, finance=300.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    default = wb.active
    if default is None:
        raise RuntimeError("openpyxl returned no default sheet")
    default.title = "C5"
    wb.create_sheet("C1")
    wb.create_sheet("C6")
    wb.create_sheet("C7")

    ws = wb["C5"]
    ws.cell(row=2, column=2, value="Liabilities")
    ws.cell(row=4, column=4, value="Mid-July ")
    ws.cell(row=4, column=7, value="Mid-Aug")
    ws.cell(row=4, column=8, value="Mid-Sept")
    ws.cell(row=4, column=16, value="Mid-Sept")
    ws.cell(row=4, column=24, value="Mid-Sept")
    ws.cell(row=4, column=32, value="Mid-Sept")
    for c in (4, 5, 6, 7, 8):
        ws.cell(row=5, column=c, value=2022 + (c - 4))
    body_start = 7
    for offset, (label, _slug) in enumerate(_C5_INDICATORS):
        r = body_start + offset
        ws.cell(row=r, column=3, value=label)   # 0-based col 2
        ws.cell(row=r, column=8, value=1000 + offset + 0.5)   # system_total (col 7)
        ws.cell(row=r, column=16, value=700 + offset + 0.5)   # commercial (col 15)
        ws.cell(row=r, column=24, value=200 + offset + 0.5)   # development (col 23)
        ws.cell(row=r, column=32, value=100 + offset + 0.5)   # finance (col 31)

    # C7: label in col 1 (openpyxl col 2), latest value in cols 6/14/22/30
    # (openpyxl cols 7/15/23/31). Values use distinct bases (5000/4000/700/300)
    # so column-wiring assertions can't accidentally pass on the wrong column.
    ws7 = wb["C7"]
    c7_body_start = 7
    for offset, (label, _slug) in enumerate(_C7_INDICATORS):
        r = c7_body_start + offset
        ws7.cell(row=r, column=2, value=label)              # 0-based col 1
        ws7.cell(row=r, column=7, value=5000 + offset + 0.5)    # system_total (col 6)
        ws7.cell(row=r, column=15, value=4000 + offset + 0.5)   # commercial (col 14)
        ws7.cell(row=r, column=23, value=700 + offset + 0.5)    # development (col 22)
        ws7.cell(row=r, column=31, value=300 + offset + 0.5)    # finance (col 30)

    wb.save(str(path))


def _fixture_for(filename: str) -> Path:
    """Return path to a fixture named ``filename``, building it if absent."""
    dest = FIXTURE_DIR / filename
    if not dest.exists():
        _build_fixture(dest)
    return dest


@pytest.fixture(scope="module", autouse=True)
def ensure_fixture() -> None:
    """Materialise the canonical fixture on first run. Cheap (≤10 rows)."""
    if not FIXTURE.exists():
        _build_fixture(FIXTURE)


@pytest.fixture(scope="module")
def result() -> ParserResult:
    return parse(str(FIXTURE), source_document_id="test-doc-id")


# ---------------------------------------------------------------------------
# Canonical-month tests (Bhadau 2082 → "Bhadra 2082")
# ---------------------------------------------------------------------------

def test_status_success(result: ParserResult) -> None:
    assert result.status == "success", f"errors={result.errors}"


def test_parser_version(result: ParserResult) -> None:
    assert result.parser_version == PARSER_VERSION == "0.3.0"


def test_row_count(result: ParserResult) -> None:
    n_classes = len(_LATEST_VALUE_COL_BY_CLASS)
    expected = (len(_C5_INDICATORS) + len(_C7_INDICATORS)) * n_classes
    assert len(result.fact_rows) == expected


def test_bank_classes_balanced(result: ParserResult) -> None:
    per_class: dict[str, int] = dict.fromkeys(_LATEST_VALUE_COL_BY_CLASS.keys(), 0)
    for row in result.fact_rows:
        per_class[row.bank_class] += 1
    expected_per_class = len(_C5_INDICATORS) + len(_C7_INDICATORS)
    assert per_class == {k: expected_per_class for k in _LATEST_VALUE_COL_BY_CLASS}


def test_required_fields_populated(result: ParserResult) -> None:
    for row in result.fact_rows:
        assert isinstance(row, BankingSectorFactRow)
        assert row.indicator_slug.startswith("bfi-c5-") or row.indicator_slug.startswith("bfi-c7-")
        assert row.unit == "npr_million"
        assert row.source_sheet in ("C5", "C7")
        assert row.reporting_period_type == "monthly"
        assert row.reporting_period_bs == "Bhadra 2082"
        assert row.fiscal_year_bs == "2082/83"
        assert row.confidence_grade == "A"
        assert row.bank_entity_id is None
        assert isinstance(row.value, float)


def test_per_class_value_wiring(result: ParserResult) -> None:
    """The synthetic fixture uses class-specific value bases: system_total
    starts at 1000.5, commercial at 700.5, development at 200.5, finance at
    100.5. Confirms column wiring is correct for all 4 classes.
    """
    by_class: dict[str, list[float]] = {
        "system_total": [],
        "commercial": [],
        "development": [],
        "finance": [],
    }
    for row in result.fact_rows:
        by_class[row.bank_class].append(row.value)

    # Pick first indicator (CAPITAL FUND -> offset 0 -> values 1000.5, 700.5,
    # 200.5, 100.5). The slug suffix is 'capital-fund'.
    cf_rows = [r for r in result.fact_rows if r.indicator_slug.endswith("capital-fund")]
    cf_by_class = {r.bank_class: r.value for r in cf_rows}
    assert cf_by_class["system_total"] == pytest.approx(1000.5)
    assert cf_by_class["commercial"] == pytest.approx(700.5)
    assert cf_by_class["development"] == pytest.approx(200.5)
    assert cf_by_class["finance"] == pytest.approx(100.5)


def test_idempotent() -> None:
    a = parse(str(FIXTURE), source_document_id="x")
    b = parse(str(FIXTURE), source_document_id="x")
    assert a.status == b.status
    assert len(a.fact_rows) == len(b.fact_rows)
    for ra, rb in zip(a.fact_rows, b.fact_rows, strict=True):
        assert ra == rb


def test_missing_file_returns_failure() -> None:
    res = parse("nonexistent-bfi.xlsx", source_document_id="x")
    assert res.status == "failure"
    assert res.errors


def test_no_unexpected_errors(result: ParserResult) -> None:
    assert result.errors == [], f"unexpected errors: {result.errors}"


# ---------------------------------------------------------------------------
# C7 sector-loan tests
# ---------------------------------------------------------------------------

def test_c7_row_count(result: ParserResult) -> None:
    c7_rows = [r for r in result.fact_rows if r.source_sheet == "C7"]
    expected = len(_C7_INDICATORS) * len(_LATEST_VALUE_COL_BY_CLASS)
    assert len(c7_rows) == expected, f"expected {expected} C7 rows, got {len(c7_rows)}"


def test_c7_slugs_correct_prefix(result: ParserResult) -> None:
    for row in result.fact_rows:
        if row.source_sheet == "C7":
            assert row.indicator_slug.startswith("bfi-c7-")


def test_c7_deprived_sector_system_total_value(result: ParserResult) -> None:
    """Deprived Sector Loan system_total should match the synthetic fixture value."""
    offset = next(i for i, (lbl, _) in enumerate(_C7_INDICATORS) if lbl == "Deprived Sector Loan")
    expected = 5000 + offset + 0.5
    row = next(
        (r for r in result.fact_rows if r.source_sheet == "C7"
         and r.indicator_slug == "bfi-c7-system-total-deprived-sector"
         and r.bank_class == "system_total"),
        None,
    )
    # slug is bfi-c7-{class}-{stem}, stem="deprived-sector"
    row = next(
        (r for r in result.fact_rows if r.source_sheet == "C7"
         and r.bank_class == "system_total"
         and r.indicator_slug.endswith("-deprived-sector")),
        None,
    )
    assert row is not None, "deprived-sector system_total row not found"
    assert row.value == pytest.approx(expected)


def test_c7_agriculture_commercial_value(result: ParserResult) -> None:
    """Agriculture-forest commercial bank value wired to col 14 (0-indexed)."""
    offset = next(i for i, (lbl, _) in enumerate(_C7_INDICATORS)
                  if lbl == "Agricultural and Forest Related")
    expected = 4000 + offset + 0.5  # commercial base in fixture
    row = next(
        (r for r in result.fact_rows if r.source_sheet == "C7"
         and r.bank_class == "commercial"
         and r.indicator_slug.endswith("-agriculture-forest")),
        None,
    )
    assert row is not None, "agriculture-forest commercial row not found"
    assert row.value == pytest.approx(expected)


def test_c7_missing_sheet_yields_partial(tmp_path: Path) -> None:
    """A file with no C7 sheet should parse C5 successfully but mark partial."""
    fixture = tmp_path / "Bhadau_2082_Publish.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("no active sheet")
    ws.title = "C5"
    # Minimal C5 data
    body_start = 7
    for offset, (label, _slug) in enumerate(_C5_INDICATORS):
        ws.cell(row=body_start + offset, column=3, value=label)
        ws.cell(row=body_start + offset, column=8, value=1000.0 + offset)
        ws.cell(row=body_start + offset, column=16, value=700.0 + offset)
        ws.cell(row=body_start + offset, column=24, value=200.0 + offset)
        ws.cell(row=body_start + offset, column=32, value=100.0 + offset)
    wb.save(str(fixture))

    res = parse(str(fixture), source_document_id="x")
    assert res.status == "partial"
    c5_rows = [r for r in res.fact_rows if r.source_sheet == "C5"]
    assert len(c5_rows) == len(_C5_INDICATORS) * len(_LATEST_VALUE_COL_BY_CLASS)
    assert any(e.error_class == "PageLayoutChanged" for e in res.errors)


def test_json_serialisable(result: ParserResult) -> None:
    """Ingest CLI consumes the JSON dict; ensure it's JSON-clean."""
    import json

    payload = json.dumps(result.to_json_dict())
    assert "fact_rows" in payload
    assert "parser_version" in payload


# ---------------------------------------------------------------------------
# Filename-derived period tests — different files → different periods
# ---------------------------------------------------------------------------

def test_chaitra_2082_period() -> None:
    """Chaitra_2082_Publish.xlsx → reporting_period_bs='Chait 2082',
    fiscal_year_bs='2082/83'.
    """
    fixture = _fixture_for("Chaitra_2082_Publish.xlsx")
    res = parse(str(fixture), source_document_id="test-chaitra")
    assert res.status == "success", f"errors={res.errors}"
    assert res.fact_rows[0].reporting_period_bs == "Chait 2082"
    assert res.fact_rows[0].fiscal_year_bs == "2082/83"


def test_asoj_2082_period() -> None:
    """Asoj_2082_Publish.xlsx → reporting_period_bs='Ashwin 2082',
    fiscal_year_bs='2082/83'.
    Asoj is a variant spelling for Ashwin (month 6).
    """
    fixture = _fixture_for("Asoj_2082_Publish.xlsx")
    res = parse(str(fixture), source_document_id="test-asoj")
    assert res.status == "success", f"errors={res.errors}"
    assert res.fact_rows[0].reporting_period_bs == "Ashwin 2082"
    assert res.fact_rows[0].fiscal_year_bs == "2082/83"


def test_baisakh_2080_period() -> None:
    """Baisakh_2080_Publish-1.xlsx → reporting_period_bs='Baisakh 2080',
    fiscal_year_bs='2080/81'.
    Baisakh is month 10 (BS FY months run Shrawan=1 … Ashadh=12, but
    Baisakh is month 10 of the FY). The BS *year* label is 2080, so
    fiscal_year_label(2080) → '2080/81'.
    """
    fixture = _fixture_for("Baisakh_2080_Publish-1.xlsx")
    res = parse(str(fixture), source_document_id="test-baisakh")
    assert res.status == "success", f"errors={res.errors}"
    assert res.fact_rows[0].reporting_period_bs == "Baisakh 2080"
    assert res.fact_rows[0].fiscal_year_bs == "2080/81"


def test_three_files_yield_distinct_periods() -> None:
    """Smoke test: parsing three files with distinct month/year combos produces
    three distinct reporting_period_bs values — confirms no global state leak.
    """
    f_chaitra = _fixture_for("Chaitra_2082_Publish.xlsx")
    f_asoj = _fixture_for("Asoj_2082_Publish.xlsx")
    f_baisakh = _fixture_for("Baisakh_2080_Publish-1.xlsx")

    res_c = parse(str(f_chaitra), source_document_id="x")
    res_a = parse(str(f_asoj), source_document_id="x")
    res_b = parse(str(f_baisakh), source_document_id="x")

    periods = {
        res_c.fact_rows[0].reporting_period_bs,
        res_a.fact_rows[0].reporting_period_bs,
        res_b.fact_rows[0].reporting_period_bs,
    }
    assert len(periods) == 3, f"expected 3 distinct periods, got: {periods}"


# ---------------------------------------------------------------------------
# Filename-parsing helper unit tests
# ---------------------------------------------------------------------------

def test_parse_period_known_variants() -> None:
    """Spot-check that romanisation variants all resolve to the right
    canonical month + year without going through the full parse pipeline.
    """
    cases: list[tuple[str, str, int]] = [
        ("Bhadau_2082_Publish.xlsx", "Bhadra", 2082),
        ("Chaitra_2082_Publish.xlsx", "Chait", 2082),
        ("Saun-2082-Publish.xlsx", "Shrawan", 2082),
        ("Asoj_2082_Publish.xlsx", "Ashwin", 2082),
        ("Shrawan_2081_Publish-1.xlsx", "Shrawan", 2081),
        ("Ashwin2078_NFRS_Publish.xlsx", "Ashwin", 2078),
        ("Jestha2079_PublishV1.xlsx", "Jestha", 2079),
        ("Ashadh-2078-2.xlsx", "Ashadh", 2078),
        ("Manghir-2078.xlsx", "Mangsir", 2078),
        ("Mangshir_2082_Publish.xlsx", "Mangsir", 2082),
        ("Falgun_2082_Publish.xlsx", "Falgun", 2082),
        ("Magh_2082_Publish-2.xlsx", "Magh", 2082),
        ("Poush_2082_Publish.xlsx", "Poush", 2082),
        ("Kartik_2082_Publish.xlsx", "Kartik", 2082),
        ("Baisakh_2080_Publish-1.xlsx", "Baisakh", 2080),
        ("Baishakh-2079.xlsx", "Baisakh", 2079),
        ("Asar_2082_Publish.xlsx", "Ashadh", 2082),
        ("Ashar2079_Publish.xlsx", "Ashadh", 2079),
    ]
    for filename, expected_month, expected_year in cases:
        result = _parse_period_from_filename(filename)
        assert not isinstance(result, Exception), f"{filename}: got error {result}"
        assert not isinstance(result, ParserError), (
            f"{filename}: got ParserError {result}"
        )
        month, year = result
        assert month == expected_month, (
            f"{filename}: expected month {expected_month!r}, got {month!r}"
        )
        assert year == expected_year, (
            f"{filename}: expected year {expected_year}, got {year}"
        )


def test_unparseable_filename_returns_failure() -> None:
    """A filename with no recognisable month token must return status='failure'
    with error_class='PeriodAmbiguous', never silently fall back.
    """
    res = parse("unknown_file_2082.xlsx", source_document_id="x")
    assert res.status == "failure"
    assert res.errors
    assert res.errors[0].error_class == "PeriodAmbiguous"


def test_unparseable_filename_no_year_returns_failure() -> None:
    """A filename with no 4-digit year must also fail cleanly."""
    res = parse("Bhadau_Publish.xlsx", source_document_id="x")
    assert res.status == "failure"
    assert res.errors
    assert res.errors[0].error_class == "PeriodAmbiguous"

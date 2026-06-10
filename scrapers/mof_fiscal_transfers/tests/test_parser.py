"""Tests for the MoF Local Fiscal Transfers parser.

Runs against a self-contained fixture XLSX built by conftest.py. The
gitignored real corpus is NEVER touched by these tests — see
``conftest.py::_inject_canonical_table``.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from mof_fiscal_transfers import PARSER_VERSION, FiscalTransferRow, parse
from mof_fiscal_transfers.tests.conftest import SAMPLE_XLSX

EXPECTED_GRANT_TYPES = {
    "equalization_minimum",
    "equalization_formula",
    "equalization_performance",
    "conditional_current",
    "conditional_capital",
    "special_current",
    "special_capital",
    "complementary_capital",
}
EXPECTED_MUNICIPALITIES = 3


@pytest.fixture(scope="module")
def result() -> dict[str, object]:
    assert SAMPLE_XLSX.exists(), f"fixture missing: {SAMPLE_XLSX}"
    return parse(str(SAMPLE_XLSX), source_document_id="test-doc-id")


def test_status_success(result: dict[str, object]) -> None:
    assert result["status"] == "success", (
        f"got status={result['status']!r} errors={result['errors']!r}"
    )


def test_parser_version(result: dict[str, object]) -> None:
    assert result["parser_version"] == PARSER_VERSION == "0.5.0"


def test_reads_data_when_sheet_named_sheet2(tmp_path: Path) -> None:
    """The real Cleaned/ exports ship transfer data on a sheet NOT named
    'Sheet1' (e.g. 'Sheet2'). The parser must fall back to the first sheet
    by position rather than failing with EncodingError. Regression for the
    Fiscal Transfer_2082_82.xlsx ingest blocker.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet2"
    ws.append(["Annex 1: Fiscal Transfer FY 2082/83 (in NPR thousand)"])
    ws.append(
        [
            "S.N.",
            "District",
            "Local Level Name",
            "Equalization Grant (Minimum)",
            "Equalization Grant (Formula-Based)",
            "Equalization Grant (Performance-Based)",
            "Conditional Grant (Current)",
            "Conditional Grant (Capital)",
            "Special Grant (Current)",
            "Special Grant (Capital)",
            "Complementary Grant (Capital)",
        ],
    )
    ws.append([1, "Kathmandu", "Kathmandu", 100000, 250000, 50000, 800000, 600000, 0, 30000, 20000])
    sheet2_xlsx = tmp_path / "sheet2_only.xlsx"
    wb.save(sheet2_xlsx)

    result = parse(str(sheet2_xlsx), source_document_id="test-doc-sheet2")

    assert result["status"] == "success", (
        f"got status={result['status']!r} errors={result['errors']!r}"
    )
    rows = result["rows"]
    assert isinstance(rows, list)
    assert len(rows) > 0


def test_code_column_used_directly_no_fuzzy_collision(tmp_path: Path) -> None:
    """When the workbook has a federal Code column, identity comes from it
    directly — not fuzzy name matching. Two rows with near-identical names but
    distinct codes must map to distinct federal_codes (regression for the
    fuzzy-collision bug that inflated FY 2082/83 totals ~65%).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet2"
    ws.append(["Annex 1: Fiscal Transfer FY 2082/83 (NPR crore)"])
    ws.append(
        [
            "Code",
            "District (English)",
            "Local Level Name (Nepali)",
            "Local Level Name (English)",
            "Local Level Type",
            "Minimum Grant",
            "Conditional Grant Current",
        ],
    )
    # Two near-identically-named rural municipalities, distinct codes.
    rm = "Rural Municipality"
    ws.append([80101101, "Taplejung", "आठराई त्रिवेणी", "Aathrai Tribeni", rm, 5.0, 19.0])
    ws.append([80101102, "Taplejung", "आठराई", "Aathrai", rm, 6.0, 21.0])
    p = tmp_path / "coded.xlsx"
    wb.save(p)

    result = parse(str(p), source_document_id="t")
    rows = result["rows"]
    assert isinstance(rows, list)
    codes = {r["federal_code"] for r in rows}
    assert codes == {"80101101", "80101102"}, codes
    # No federal_code appears more than once per grant_type.
    cc = [r for r in rows if r["grant_type"] == "conditional_current"]
    assert len(cc) == len({r["federal_code"] for r in cc}) == 2
    by_code = {r["federal_code"]: r["amount_npr"] for r in cc}
    assert by_code["80101101"] == pytest.approx(19.0)
    assert by_code["80101102"] == pytest.approx(21.0)


def test_row_count(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    # 3 municipalities × 8 grant types (including zero-valued special_current)
    assert len(rows) == EXPECTED_MUNICIPALITIES * len(EXPECTED_GRANT_TYPES)


def test_grant_types_complete(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    grant_types_emitted = {row["grant_type"] for row in rows}
    assert grant_types_emitted == EXPECTED_GRANT_TYPES


def test_kathmandu_equalization_minimum(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    matches = [
        r
        for r in rows
        if r["federal_code"] == "80101101" and r["grant_type"] == "equalization_minimum"
    ]
    assert len(matches) == 1
    assert matches[0]["amount_npr"] == pytest.approx(100000.0)
    assert matches[0]["unit"] == "npr_crore"
    assert matches[0]["confidence_grade"] == "A"
    assert matches[0]["fiscal_year_bs"] == "2082/83"


def test_pokhara_conditional_current(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    matches = [
        r
        for r in rows
        if r["federal_code"] == "80201101" and r["grant_type"] == "conditional_current"
    ]
    assert len(matches) == 1
    assert matches[0]["amount_npr"] == pytest.approx(600000.0)
    assert matches[0]["district_en"] == "Kaski"


def test_total_row_skipped(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    # No federal_code maps to "Total"; verify no row has impossibly-large amount
    # belonging to the aggregator row.
    for row in rows:
        assert row["municipality_name_en"] != "Total"


def test_all_row_fields_typed(result: dict[str, object]) -> None:
    rows = result["rows"]
    assert isinstance(rows, list)
    for row in rows:
        # Validate by reconstructing the dataclass — strict type check.
        rebuilt = FiscalTransferRow(**row)
        assert rebuilt.federal_code.isdigit()
        assert len(rebuilt.federal_code) == 8
        assert rebuilt.grant_type in EXPECTED_GRANT_TYPES
        assert rebuilt.amount_npr >= 0
        assert rebuilt.unit == "npr_crore"


def test_idempotent() -> None:
    """Per docs/DATA_PIPELINE.md: parsers must be deterministic."""
    first = parse(str(SAMPLE_XLSX), source_document_id="x")
    second = parse(str(SAMPLE_XLSX), source_document_id="x")
    assert first == second


def test_missing_file_returns_failure() -> None:
    res = parse("nonexistent.xlsx", source_document_id="x")
    assert res["status"] == "failure"
    errors = res["errors"]
    assert isinstance(errors, list)
    assert len(errors) >= 1


def test_no_unexpected_errors(result: dict[str, object]) -> None:
    assert result["errors"] == [], f"unexpected parser errors: {result['errors']!r}"


def test_fixture_xlsx_is_real_xlsx() -> None:
    """Sanity: openpyxl must accept the fixture (catches binary corruption)."""
    from openpyxl import load_workbook

    wb = load_workbook(SAMPLE_XLSX, read_only=True)
    assert "Sheet1" in wb.sheetnames
    wb.close()


def test_fixture_path_under_tests_dir() -> None:
    """Fixture lives under the tests/ tree, not in the repo root."""
    assert SAMPLE_XLSX.parent == Path(__file__).resolve().parent / "fixtures"


def test_real_file_short_form_headers(tmp_path: Path) -> None:
    """Parser must extract equalization grants when the real MoF file uses
    short-form column headers: "Minimum Grant", "Formula Based Grant",
    "Performance Based Grant" (no "Equalization" prefix). Regression for
    v0.3.0 equalization extraction from Fiscal Transfer_2082_82.xlsx.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet2"
    ws.append(["Annex 1: Fiscal Transfer FY 2082/83 (in NPR thousand)"])
    ws.append(
        [
            "S.N.",
            "District",
            "Local Level Name",
            "Minimum Grant",
            "Formula Based Grant",
            "Performance Based Grant",
            "Total Equalization",        # aggregator — must be EXCLUDED
            "Conditional Grant (Current)",
            "Conditional Grant (Capital)",
            "Special Grant (Current)",
            "Special Grant (Capital)",
            "Complementary Grant (Capital)",
            "Total",                     # grand-total aggregator — must be EXCLUDED
        ],
    )
    ws.append(
        [
            1, "Kathmandu", "Kathmandu",
            111000, 222000, 333000,   # eq_min, eq_formula, eq_perf
            666000,                   # Total Equalization (excluded)
            800000, 600000, 0, 30000, 20000,  # conditional + special + complementary
            1800000,                  # grand Total (excluded)
        ],
    )
    short_form_xlsx = tmp_path / "short_form_headers.xlsx"
    wb.save(short_form_xlsx)

    result = parse(str(short_form_xlsx), source_document_id="test-short-form")

    assert result["status"] == "success", (
        f"got status={result['status']!r} errors={result['errors']!r}"
    )
    rows = result["rows"]
    assert isinstance(rows, list)

    grant_types_emitted = {r["grant_type"] for r in rows}
    assert "equalization_minimum" in grant_types_emitted
    assert "equalization_formula" in grant_types_emitted
    assert "equalization_performance" in grant_types_emitted

    # Total columns must be excluded — row count must be exactly 8 (one per grant type).
    assert len(rows) == len(EXPECTED_GRANT_TYPES), (
        f"expected {len(EXPECTED_GRANT_TYPES)} rows (8 grant types), got {len(rows)}; "
        f"grant_types={grant_types_emitted!r}"
    )

    # Spot-check equalization values.
    eq_min = next((r for r in rows if r["grant_type"] == "equalization_minimum"), None)
    assert eq_min is not None
    assert eq_min["amount_npr"] == pytest.approx(111000.0)

    eq_formula = next((r for r in rows if r["grant_type"] == "equalization_formula"), None)
    assert eq_formula is not None
    assert eq_formula["amount_npr"] == pytest.approx(222000.0)

    eq_perf = next((r for r in rows if r["grant_type"] == "equalization_performance"), None)
    assert eq_perf is not None
    assert eq_perf["amount_npr"] == pytest.approx(333000.0)


def test_total_columns_excluded() -> None:
    """_match_grant_type must return None for any header containing 'total'."""
    from mof_fiscal_transfers.parser import _match_grant_type

    for header in (
        "Total Equalization",
        "Total Conditional",
        "Total Special",
        "Grand Total",
        "Total",
        "total equalization grant",
    ):
        assert _match_grant_type(header) is None, (
            f"_match_grant_type({header!r}) should return None (total column)"
        )

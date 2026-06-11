"""NRB Financial Corporations Survey (FCS) XLSX parser -- deterministic Python.

Source: NRB quarterly XLSX files published at
https://www.nrb.org.np/category/financial-corporations-survey/

Sheet layout (FCS sheet -- stable across releases):
    Row 2  -- table title
    Row 3  -- unit note (Rs. in million)
    Row 4  -- "Aggregates (Headings)" / "Mid-Month" label headers
    Row 5  -- date headers in "YYYY Month" format (e.g. "2025 July", "2025 Oct")
    Row 6  -- (empty separator)
    Rows 7-26 -- data; column B = indicator label, columns C+ = time-series values

Target rows:
    Row  7 -- Foreign Assets, Net        -> nrb-fcs-net-foreign-assets-annual
    Row 18 -- Credit to Private Sector   -> nrb-fcs-credit-private-annual
    Row 21 -- Liquid Liabilities         -> nrb-fcs-m2-annual

"Liquid Liabilities" (Row 21) is the M2 equivalent per IMF MFSM 2016.
Annual slugs (-annual) are named for their most meaningful period (mid-July =
end of Nepali fiscal year in Ashadh); all column observations are emitted
so the validation layer can filter by period type.

Versioning: bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import (
    BsMonth,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ParserError,
    ParserResult,
    ReportingPeriodType,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-financial-corporations-survey"

# Target rows: (1-based row number on FCS sheet, indicator slug, label fragment)
_TARGET_ROWS: Final[tuple[tuple[int, str, str], ...]] = (
    (7,  "nrb-fcs-net-foreign-assets-annual", "Foreign Assets, Net"),
    (18, "nrb-fcs-credit-private-annual",     "Credit to Private Sector"),
    (21, "nrb-fcs-m2-annual",                 "Liquid Liabilities"),
)

_UNIT: Final[str] = "npr_million"
_SHEET_NAME: Final[str] = "FCS"
_DATE_ROW: Final[int] = 5       # 1-based row number of date headers
_NRB_PUB_LAG_DAYS: Final[int] = 60  # approximate NRB publication lag

# AD month name (lower / abbreviated) -> BS month
_AD_MONTH_TO_BS: Final[dict[str, BsMonth]] = {
    "january": "Poush",    "jan": "Poush",
    "february": "Magh",    "feb": "Magh",
    "march": "Falgun",     "mar": "Falgun",
    "april": "Chait",      "apr": "Chait",
    "may": "Baisakh",
    "june": "Jestha",      "jun": "Jestha",
    "july": "Ashadh",      "jul": "Ashadh",
    "august": "Shrawan",   "aug": "Shrawan",
    "september": "Bhadra", "sep": "Bhadra",
    "october": "Ashwin",   "oct": "Ashwin",
    "november": "Kartik",  "nov": "Kartik",
    "december": "Mangsir", "dec": "Mangsir",
}

# BS months in fiscal-year order (month 1 = Shrawan, month 12 = Ashadh).
_FY_ORDER: Final[tuple[BsMonth, ...]] = (
    "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush",
    "Magh", "Falgun", "Chait", "Baisakh", "Jestha", "Ashadh",
)
_BS_MONTH_FY_POS: Final[dict[str, int]] = {
    m: i + 1 for i, m in enumerate(_FY_ORDER)
}

_DATE_RE: Final[re.Pattern[str]] = re.compile(r"(\d{4})\s+(\w+)", re.IGNORECASE)


def _parse_date_header(header: str) -> tuple[BsMonth, int] | None:
    """Parse "YYYY Month" -> (bs_month, bs_year). Returns None on failure.

    BS year derivation mirrors periods.mid_month_ad:
        ad_month >= 7  ->  bs_year = ad_year + 57
        ad_month <  7  ->  bs_year = ad_year + 56
    """
    from _common.periods import _BS_MONTH_TO_AD_MONTH  # noqa: PLC0415

    m = _DATE_RE.search(header.strip())
    if not m:
        return None
    ad_year = int(m.group(1))
    bs_month = _AD_MONTH_TO_BS.get(m.group(2).lower())
    if bs_month is None:
        return None
    ad_month = _BS_MONTH_TO_AD_MONTH[bs_month]
    bs_year = ad_year + 57 if ad_month >= 7 else ad_year + 56
    return bs_month, bs_year


def _bs_fy_start(bs_month: BsMonth, bs_year: int) -> int:
    """Return BS fiscal-year start year for a given BS month + year.

    Months 1-9 (Shrawan-Chait): fy_start = bs_year.
    Months 10-12 (Baisakh-Ashadh): fy_start = bs_year - 1.
    """
    pos = _BS_MONTH_FY_POS[bs_month]
    return bs_year if pos <= 9 else bs_year - 1


def _reporting_period_type(bs_month: BsMonth) -> ReportingPeriodType:
    """Map a BS month to the appropriate ReportingPeriodType."""
    if bs_month == "Ashadh":
        return "annual"
    if bs_month in ("Ashwin", "Poush", "Chait"):
        return "quarterly"
    return "monthly"


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse one NRB FCS quarterly XLSX; emit monetary system indicators.

    Arguments:
        source_document_path: filesystem path to the downloaded XLSX.
        source_document_id: opaque FK from source_documents; threaded through
            for orchestrator symmetry.

    Returns:
        ParserResult with status, staging_rows, errors.
    """
    _ = source_document_id

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source file not found: {path}",
            )],
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl failed to open file: {exc}",
            )],
        )

    if _SHEET_NAME not in wb.sheetnames:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="ColumnMissing",
                error_detail=(
                    f"sheet {_SHEET_NAME!r} not found; "
                    f"available: {wb.sheetnames}"
                ),
            )],
        )

    ws = wb[_SHEET_NAME]
    all_rows: list[list[object]] = [
        list(row) for row in ws.iter_rows(values_only=True)
    ]
    errors: list[ParserError] = []

    if len(all_rows) < _DATE_ROW:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="PageLayoutChanged",
                error_detail=f"sheet has fewer than {_DATE_ROW} rows",
            )],
        )

    # Parse date headers from row 5 (1-based -> index 4).
    date_row = all_rows[_DATE_ROW - 1]
    date_cols: list[tuple[int, BsMonth, int]] = []
    for col_idx, cell in enumerate(date_row):
        if cell is None:
            continue
        parsed = _parse_date_header(str(cell))
        if parsed is not None:
            bs_month, bs_year = parsed
            date_cols.append((col_idx, bs_month, bs_year))

    if not date_cols:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="PageLayoutChanged",
                error_detail="no parseable date headers found in row 5",
            )],
        )

    # Heuristic publication date: latest observation + lag.
    latest_bs_month, latest_bs_year = date_cols[-1][1], date_cols[-1][2]
    pub_ad = mid_month_ad(latest_bs_month, latest_bs_year) + timedelta(
        days=_NRB_PUB_LAG_DAYS
    )
    latest_fy = _bs_fy_start(latest_bs_month, latest_bs_year)
    pub_bs = f"~{fiscal_year_label(latest_fy)} (heuristic)"

    _dummy_dt = datetime(2000, 1, 1, tzinfo=UTC)
    base = StagingRowDraft(
        indicator_slug_raw="",
        value=0.0,
        unit=_UNIT,
        reporting_period_type="annual",
        reporting_period_bs="",
        reporting_period_ad_start=_dummy_dt,
        reporting_period_ad_end=_dummy_dt,
        publication_date_ad=pub_ad,
        publication_date_bs=pub_bs,
        fiscal_year_bs="",
        fiscal_year_ad_label="",
        confidence_grade_proposed="A",
        parser_notes=None,
    )

    staging_rows: list[StagingRowDraft] = []
    slugs_found: set[str] = set()

    for row_num_1based, slug, label_fragment in _TARGET_ROWS:
        row_idx = row_num_1based - 1
        if row_idx >= len(all_rows):
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail=(
                    f"indicator {slug!r}: expected row {row_num_1based} "
                    f"but sheet has only {len(all_rows)} rows"
                ),
            ))
            continue

        data_row = all_rows[row_idx]
        label_cell = data_row[1] if len(data_row) > 1 else None
        label_str = str(label_cell).strip() if label_cell is not None else ""

        if label_fragment.lower() not in label_str.lower():
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail=(
                    f"indicator {slug!r}: label mismatch at row {row_num_1based}; "
                    f"expected {label_fragment!r}, got {label_str!r}"
                ),
                source_excerpt=label_str,
            ))
            continue

        row_had_value = False
        for col_idx, bs_month, bs_year in date_cols:
            if col_idx >= len(data_row):
                continue
            cell_val = data_row[col_idx]
            if cell_val is None:
                continue
            try:
                value = float(cell_val)
            except (TypeError, ValueError):
                continue

            fy_start = _bs_fy_start(bs_month, bs_year)
            fy_bs = fiscal_year_label(fy_start)
            fy_ad = fiscal_year_ad_label(fy_start)
            obs_dt = mid_month_ad(bs_month, bs_year)
            period_type = _reporting_period_type(bs_month)
            period_bs = f"{fy_bs} {bs_month}"

            staging_rows.append(replace(
                base,
                indicator_slug_raw=slug,
                value=value,
                reporting_period_type=period_type,
                reporting_period_bs=period_bs,
                reporting_period_ad_start=obs_dt,
                reporting_period_ad_end=obs_dt,
                fiscal_year_bs=fy_bs,
                fiscal_year_ad_label=fy_ad,
            ))
            row_had_value = True

        if row_had_value:
            slugs_found.add(slug)
        else:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail=(
                    f"indicator {slug!r}: no numeric values in any date column"
                ),
            ))

    if not staging_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors or [ParserError(
                error_class="PageLayoutChanged",
                error_detail="no indicator rows extracted",
            )],
        )

    all_target_slugs = {s for _, s, _ in _TARGET_ROWS}
    status = "success" if slugs_found >= all_target_slugs else "partial"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


def _main() -> None:
    """CLI entrypoint used by the Node ingestion orchestrator.

    Argv: parser.py <source_document_path> <source_document_id>
    Writes JSON to stdout. Exit codes:
        0: parser ran (status may still be 'failure')
        2: usage error
        1: catastrophic crash
    """
    import json
    import sys
    from dataclasses import asdict

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    payload = asdict(result)
    for row in payload.get("staging_rows", []):
        for key in (
            "reporting_period_ad_start",
            "reporting_period_ad_end",
            "publication_date_ad",
        ):
            val = row.get(key)
            if isinstance(val, datetime):
                row[key] = val.isoformat()

    json.dump(payload, sys.stdout)


if __name__ == "__main__":
    _main()

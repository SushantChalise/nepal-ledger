"""NRB historical Workers' Remittances (long annual series) — deterministic parser.

Source: NRB "Trade and Balance of Payments" workbook, sheet ``BOP 2000-`` —
the long annual Balance-of-Payments summary (FY 2000/01 onward, "In Million
Rupees"). Row label **"Workers' remittances"** carries the annual remittance
inflow the NDRI Migration Atlas's Figure 13 trend is built from.

Why a dedicated parser (not ``scrapers/nrb_dne/parser.py``):
  - Different file (historical workbook) + a focused single-row extraction.
  - Different CONCEPT: this is the BPM5 "Workers' remittances" line, NOT the
    modern BPM6 "Personal transfers, Credit" that feeds ``dne-remittance-inflow``
    (parser nrb_dne v0.8.0). They overlap closely in recent years but are not the
    same measure, so per the Data Continuity Protocol we ship this as its OWN
    clearly-labelled series (slug ``dne-remittance-workers-historical``). The BPM5
    line ends at FY 2020/21 (961,054.6 npr_million); NRB switched to BPM6
    thereafter, so this series DOVETAILS with the modern BPM6 series — together
    they form the full long trend. We never splice across the BPM5→BPM6 break.

Determinism (ADR-0003): pure openpyxl table reading + ``_common.periods`` date
math. No LLM, no network. Same input ⇒ byte-identical ``ParserResult`` JSON.

Calendar (ADR-0013): the sheet's FY headers are AD fiscal years ("2000/01").
Each is converted to its BS fiscal year by +57 on the start year (FY 2000/01 AD
= BS 2057/58; FY 2020/21 AD = BS 2077/78), matching the nrb_dne AD→BS offset.

Output: a ``_common.types.ParserResult`` (the staging-pipeline contract) — one
annual ``StagingRowDraft`` per year column. Run ``python parser.py <xlsx>`` to
emit the JSON, or ``--verify`` for a human-readable reconciliation summary.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

from _common.periods import fiscal_year_ad_label, fiscal_year_label, mid_month_ad
from _common.types import ParserError, ParserResult, StagingRowDraft

PARSER_VERSION: Final[str] = "0.1.0"

SHEET: Final[str] = "BOP 2000-"
REMITTANCE_LABEL: Final[str] = "workersremittances"  # normalized match key
INDICATOR_SLUG: Final[str] = "dne-remittance-workers-historical"
UNIT: Final[str] = "npr_million"
AD_TO_BS_OFFSET: Final[int] = 57
_FY_RE: Final[re.Pattern[str]] = re.compile(r"^(19|20)(\d{2})\s*/\s*(\d{2})$")
_NOTE: Final[str] = (
    "BPM5 'Workers' remittances' from the historical Trade-and-Balance-of-Payments "
    "workbook (sheet 'BOP 2000-'); distinct concept from BPM6 dne-remittance-inflow; "
    "series ends FY2020/21 and dovetails with the BPM6 series."
)


def _norm(s: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower()) if s is not None else ""


def _leading_text(row: tuple[object, ...]) -> str:
    """The first non-blank text cell of a row (the row's label)."""
    return next((c for c in row if isinstance(c, str) and c.strip()), "")


def _ad_start_year(token: object) -> int | None:
    """'2000/01' → 2000 (AD fiscal-year start year), else None."""
    m = _FY_RE.match(str(token).strip()) if token is not None else None
    return int(m.group(1) + m.group(2)) if m else None


def parse(xlsx_path: Path) -> ParserResult:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[
                ParserError(
                    error_class="SheetMissing",
                    error_detail=f"sheet {SHEET!r} not found in {xlsx_path.name}",
                    source_excerpt=None,
                )
            ],
        )
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()

    # Header row = the row with the most AD-FY tokens → {col: ad_start_year}.
    fy_cols: dict[int, int] = {}
    for row in rows:
        cols = {j: y for j, c in enumerate(row) if (y := _ad_start_year(c)) is not None}
        if len(cols) > len(fy_cols):
            fy_cols = cols

    # Remittance row = first row whose leading text cell normalizes to the label.
    remit_row = next(
        (row for row in rows if _norm(_leading_text(row)) == REMITTANCE_LABEL),
        None,
    )

    errors: list[ParserError] = []
    if not fy_cols:
        errors.append(ParserError("PeriodUnparseable", "no fiscal-year header row found", None))
    if remit_row is None:
        errors.append(ParserError("RowMissing", "'Workers' remittances' row not found", None))
    if not fy_cols or remit_row is None:
        return ParserResult(status="failure", parser_version=PARSER_VERSION, errors=errors)

    staging_rows: list[StagingRowDraft] = []
    for col, ad_start in sorted(fy_cols.items()):
        cell = remit_row[col] if col < len(remit_row) else None
        if cell is None or not isinstance(cell, (int, float)):
            continue  # projected / blank year — never fabricate
        bs_start = ad_start + AD_TO_BS_OFFSET
        staging_rows.append(
            StagingRowDraft(
                indicator_slug_raw=INDICATOR_SLUG,
                value=round(float(cell), 4),
                unit=UNIT,
                reporting_period_type="annual",
                reporting_period_bs=fiscal_year_label(bs_start),
                reporting_period_ad_start=mid_month_ad("Shrawan", bs_start),
                reporting_period_ad_end=mid_month_ad("Shrawan", bs_start + 1),
                publication_date_ad=datetime(1970, 1, 1),  # set below to latest FY end
                publication_date_bs="",
                fiscal_year_bs=fiscal_year_label(bs_start),
                fiscal_year_ad_label=fiscal_year_ad_label(bs_start),
                confidence_grade_proposed="B",
                parser_notes=_NOTE,
            )
        )

    if not staging_rows:
        errors.append(ParserError("NoDataExtracted", "no numeric remittance values found", None))
        return ParserResult(status="failure", parser_version=PARSER_VERSION, errors=errors)

    # One publication date for the whole historical compilation: the end of the
    # latest fiscal year present (the workbook is published after its last point).
    pub = max(r.reporting_period_ad_end for r in staging_rows)
    pub_bs = max(staging_rows, key=lambda r: r.reporting_period_ad_end).reporting_period_bs
    staging_rows = [
        StagingRowDraft(**{**vars(r), "publication_date_ad": pub, "publication_date_bs": pub_bs})
        for r in staging_rows
    ]

    status = "partial" if errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


def _verify(result: ParserResult) -> None:
    rows = result.staging_rows
    sys.stderr.write(
        f"[remittance-history] status={result.status} v{result.parser_version} "
        f"rows={len(rows)} errors={len(result.errors)}\n"
    )
    if rows:
        f, l = rows[0], rows[-1]
        sys.stderr.write(
            f"  first: FY{f.fiscal_year_ad_label} (BS {f.fiscal_year_bs}) = {f.value:,} {f.unit}\n"
            f"  last:  FY{l.fiscal_year_ad_label} (BS {l.fiscal_year_bs}) = {l.value:,} {l.unit}\n"
        )
        # Reconcile the BPM5 endpoint (FY2020/21 ≈ NPR 961 bn) — confirms the
        # series + the BPM5→BPM6 dovetail point.
        anchor = next((r.value for r in rows if r.fiscal_year_ad_label == "2020/21"), None)
        if anchor is not None:
            ok = 950_000 <= float(anchor) <= 970_000
            sys.stderr.write(
                f"  reconcile FY2020/21 (BPM5 endpoint) = {anchor:,} npr_million "
                f"(≈ NPR {float(anchor) / 1000:,.0f} bn) — {'OK' if ok else 'OUT OF RANGE'}\n"
            )


def _main() -> None:
    args = sys.argv[1:]
    if not args:
        sys.stderr.write("usage: parser.py <xlsx_path> [--verify]\n")
        sys.exit(2)
    path = Path(args[0])
    if not path.exists():
        sys.stderr.write(f"file not found: {path}\n")
        sys.exit(2)
    result = parse(path)
    if "--verify" in args:
        _verify(result)
    else:
        json.dump(result.to_json_dict(), sys.stdout, ensure_ascii=True)
    sys.exit(0 if result.status in ("success", "partial") else 1)


if __name__ == "__main__":
    _main()

"""NRB Database on Nepalese Economy — Financial Sector XLSX parser.

Source: NRB Financial Sector datasets (https://www.nrb.org.np/database-on-nepalese-economy/financial-sector/)

Strategy:
    Four XLSX files cover all 8 target slugs. Each file has a different
    date-header layout; the parser handles each independently using openpyxl
    direct cell access. All date parsing is deterministic (no LLM) and maps
    AD calendar months to BS months via the canonical periods module.

    Files and indicators:
        1. Loans-of-the-BFIs-Sectorwise.xlsx (sheet: Sect_creditTotal)
           - nrb-finsec-loans-agriculture-monthly → '1 Agriculture' row
           - nrb-finsec-loans-manufacturing-monthly → '3 Productions' row
           - nrb-finsec-loans-realestate-monthly → '9.11 Real Estates' row
           Year in row 4, month in row 5 (AD month names, one per column).

        2. Monetary-survey.xlsx (sheet: MA)
           - nrb-finsec-m1-level-monthly → column C (M1, Rs. in million)
           - nrb-finsec-m2-level-monthly → column E (M2, Rs. in million)
           Year in column A (appears only at Mid-Jan each year), month in col B
           (e.g. 'Mid-Dec', 'Mid-Jan'). Data starts row 4.

        3. Structure-of-interest-rates-1.xlsx (sheet: Historical Int. Rate Final)
           - nrb-finsec-deposit-rate-monthly → row 34 (G. Weighted Average Deposit Rate)
           - nrb-finsec-lending-rate-monthly → row 35 (H. Weighted Average Lending Rate)
           Year in row 2, month in row 3 (AD month names as column headers from col 2).

        4. NEPSE.xlsx (sheet: NEPSE)
           - nrb-finsec-nepse-index-monthly → NEPSE Index (Month End)
           FY label 'YYYY/YY' in col A (rows 3+), month columns Aug–Jul (cols 2–13).
           Note: 'August' = end-of-Shrawan per NRB documentation.

Date handling:
    AD calendar month names map to BS months (mid-month approximation):
        Jul→Shrawan, Aug→Bhadra, Sep→Ashwin, Oct→Kartik, Nov→Mangsir, Dec→Poush,
        Jan→Magh, Feb→Falgun, Mar→Chait, Apr→Baisakh, May→Jestha, Jun→Ashadh

    For NEPSE (fiscal-year layout), August column = end of Shrawan of that FY,
    January = end of Poush, etc. (confirmed by NRB's own note in the file).

    BS year derivation:
        AD months Jul–Dec: BS year = AD year + 57
        AD months Jan–Jun: BS year = AD year + 56

    Fiscal year:
        BS months Shrawan–Poush (AD Jul–Dec): FY start = BS year
        BS months Magh–Ashadh (AD Jan–Jun): FY start = BS year - 1

Versioning:
    Bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import (
    BsMonth,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ParserError,
    ParserResult,
    ParserStatus,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "nrb-db-financial-sector"

# ── AD month name → BS month mapping ─────────────────────────────────────────

# Canonical mapping: AD month name (various spellings) → (BS month, is_first_half)
# is_first_half=True → AD months Jul–Dec where BS year = AD year + 57
# is_first_half=False → AD months Jan–Jun where BS year = AD year + 56
_AD_MONTH_TO_BS: Final[dict[str, BsMonth]] = {
    "jan": "Magh",
    "january": "Magh",
    "feb": "Falgun",
    "february": "Falgun",
    "mar": "Chait",
    "march": "Chait",
    "apr": "Baisakh",
    "april": "Baisakh",
    "may": "Jestha",
    "jun": "Ashadh",
    "june": "Ashadh",
    "jul": "Shrawan",
    "july": "Shrawan",
    "aug": "Bhadra",
    "august": "Bhadra",
    "sep": "Ashwin",
    "sept": "Ashwin",
    "september": "Ashwin",
    "oct": "Kartik",
    "october": "Kartik",
    "nov": "Mangsir",
    "november": "Mangsir",
    "dec": "Poush",
    "december": "Poush",
}

# AD months that fall in the second half of the AD year (Jul–Dec).
# For these months: BS year = AD year + 57.
# For Jan–Jun: BS year = AD year + 56.
_SECOND_HALF_AD_MONTHS: Final[frozenset[str]] = frozenset({
    "jul", "july", "aug", "august", "sep", "sept", "september",
    "oct", "october", "nov", "november", "dec", "december",
})

# FY start (BS) month position in fiscal year:
# Shrawan=1..Poush=6 (AD Jul-Dec): FY start = BS year
# Magh=7..Ashadh=12 (AD Jan-Jun): FY start = BS year - 1
_FIRST_HALF_FY_BS_MONTHS: Final[frozenset[BsMonth]] = frozenset({
    "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush",
})

# NEPSE: month-column label → (BS month name, is_second_half_ad)
# Aug col = end of Shrawan (AD Jul/Aug boundary), confirmed by NRB note.
# The month column represents "end of that AD month = mid BS month".
# NRB: "August represents Mid August that represents end of month Shrawan"
# so Aug col → Shrawan, Sep → Bhadra, etc.
_NEPSE_COL_TO_BS: Final[dict[str, BsMonth]] = {
    "aug": "Shrawan",
    "sep": "Bhadra",
    "oct": "Ashwin",
    "nov": "Kartik",
    "dec": "Mangsir",
    "jan": "Poush",
    "feb": "Magh",
    "mar": "Falgun",
    "apr": "Chait",
    "may": "Baisakh",
    "jun": "Jestha",
    "jul": "Ashadh",
}

# NEPSE column position (0-indexed from data start, col B=1) → BS month label
# cols: Aug=1, Sep=2, Oct=3, Nov=4, Dec=5, Jan=6, Feb=7, Mar=8, Apr=9, May=10, Jun=11, Jul=12
_NEPSE_COL_LABELS: Final[tuple[str, ...]] = (
    "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar", "apr", "may", "jun", "jul",
)


def _bs_year_from_ad(ad_year: int, ad_month_key: str) -> int:
    """Derive BS year from AD year and month key (lowercase month name)."""
    if ad_month_key in _SECOND_HALF_AD_MONTHS:
        return ad_year + 57
    return ad_year + 56


def _fy_bs_start(bs_month: BsMonth, bs_year: int) -> int:
    """Return fiscal-year start BS year for a given BS month + BS year."""
    if bs_month in _FIRST_HALF_FY_BS_MONTHS:
        return bs_year
    return bs_year - 1


def _normalize_month_key(raw: object) -> str | None:
    """Normalize a cell value to a lowercase month key. Returns None if not a month."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    # Strip 'mid-' prefix used by Monetary survey (e.g. 'Mid-Dec')
    s = re.sub(r"^mid[-\s]+", "", s)
    return s if s in _AD_MONTH_TO_BS else None


def _make_row(
    *,
    slug: str,
    value: float,
    unit: str,
    bs_month: BsMonth,
    bs_year: int,
    pub_ad: datetime,
) -> StagingRowDraft:
    """Construct a StagingRowDraft for a monthly indicator point."""
    mid = mid_month_ad(bs_month, bs_year)
    fy_start = _fy_bs_start(bs_month, bs_year)
    return StagingRowDraft(
        indicator_slug_raw=slug,
        value=value,
        unit=unit,
        reporting_period_type="monthly",
        reporting_period_bs=f"{fiscal_year_label(fy_start)} {bs_month}",
        reporting_period_ad_start=mid,
        reporting_period_ad_end=mid,
        publication_date_ad=pub_ad,
        publication_date_bs=f"~{fiscal_year_label(fy_start)} (heuristic)",
        fiscal_year_bs=fiscal_year_label(fy_start),
        fiscal_year_ad_label=fiscal_year_ad_label(fy_start),
        confidence_grade_proposed="A",
    )


# ── File parsers ──────────────────────────────────────────────────────────────


def _parse_loans(
    path: Path,
    pub_ad: datetime,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse Loans-of-the-BFIs-Sectorwise.xlsx, sheet Sect_creditTotal.

    Row 4 (index 3) = AD year integers (integer, repeated for each month).
    Row 5 (index 4) = AD month names (e.g. 'Jul', 'Aug', ..., 'December').
    Row 6+ = sector rows with ' 1 Agriculture' etc. as row label in col A.
    """
    SHEET = "Sect_creditTotal"
    TARGETS: dict[str, str] = {
        "1 agriculture": "nrb-finsec-loans-agriculture-monthly",
        "3 productions": "nrb-finsec-loans-manufacturing-monthly",
        "9.11 real estates": "nrb-finsec-loans-realestate-monthly",
    }
    UNIT = "npr_million"

    rows: list[StagingRowDraft] = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"Loans XLSX open failed: {exc}",
        ))
        return rows

    try:
        if SHEET not in wb.sheetnames:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail=f"Loans XLSX: sheet '{SHEET}' not found; found: {wb.sheetnames}",
            ))
            return rows

        ws = wb[SHEET]
        all_rows = list(ws.iter_rows(values_only=True))

        year_row = all_rows[3]   # Row 4 (0-indexed: 3)
        month_row = all_rows[4]  # Row 5 (0-indexed: 4)

        # Build col→(bs_month, bs_year) index for all valid data columns.
        col_date: dict[int, tuple[BsMonth, int]] = {}
        current_year: int | None = None
        for c, (yr_cell, mo_cell) in enumerate(zip(year_row, month_row)):
            if yr_cell is not None and isinstance(yr_cell, (int, float)):
                current_year = int(yr_cell)
            if current_year is None:
                continue
            mo_key = _normalize_month_key(mo_cell)
            if mo_key is None:
                continue
            bs_month = _AD_MONTH_TO_BS.get(mo_key)
            if bs_month is None:
                continue
            bs_year = _bs_year_from_ad(current_year, mo_key)
            col_date[c] = (bs_month, bs_year)

        if not col_date:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail="Loans XLSX: no valid date columns found in rows 4-5",
            ))
            return rows

        # Find target sector rows.
        found_slugs: set[str] = set()
        for row in all_rows[5:]:  # data starts row 6
            label = row[0]
            if label is None:
                continue
            normalized_label = str(label).strip().lower()
            # Strip leading digits and spaces to match
            matched_slug: str | None = None
            for target_key, slug in TARGETS.items():
                if normalized_label == target_key or normalized_label.lstrip("0123456789. ") == target_key.lstrip("0123456789. "):
                    matched_slug = slug
                    break
                # Also try: strip spaces from label
                stripped = re.sub(r"\s+", " ", normalized_label).strip()
                if stripped == target_key or stripped.endswith(target_key):
                    matched_slug = slug
                    break

            if matched_slug is None:
                continue
            found_slugs.add(matched_slug)

            for c, (bs_month, bs_year) in col_date.items():
                cell_val = row[c]
                if cell_val is None:
                    continue
                try:
                    value = float(cell_val)
                except (ValueError, TypeError):
                    continue
                rows.append(_make_row(
                    slug=matched_slug,
                    value=value,
                    unit=UNIT,
                    bs_month=bs_month,
                    bs_year=bs_year,
                    pub_ad=pub_ad,
                ))

        for target_key, slug in TARGETS.items():
            if slug not in found_slugs:
                errors.append(ParserError(
                    error_class="ColumnMissing",
                    error_detail=f"Loans XLSX: row '{target_key}' not found in sheet '{SHEET}'",
                ))
    finally:
        wb.close()

    return rows


def _parse_monetary(
    path: Path,
    pub_ad: datetime,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse Monetary-survey.xlsx, sheet MA.

    Row 1: title
    Row 2: column headers (Year, Month, M1, None, M2, ...)
    Row 3: unit row
    Row 4+: data rows
    Col A (index 0): year integer, appears only at 'Mid-Jan' rows
    Col B (index 1): month string e.g. 'Mid-Aug', 'Mid-Jan'
    Col C (index 2): M1 value
    Col E (index 4): M2 value
    """
    SHEET = "MA"
    UNIT = "npr_million"
    M1_SLUG = "nrb-finsec-m1-level-monthly"
    M2_SLUG = "nrb-finsec-m2-level-monthly"

    rows: list[StagingRowDraft] = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"Monetary XLSX open failed: {exc}",
        ))
        return rows

    try:
        if SHEET not in wb.sheetnames:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail=f"Monetary XLSX: sheet '{SHEET}' not found; found: {wb.sheetnames}",
            ))
            return rows

        ws = wb[SHEET]
        all_rows = list(ws.iter_rows(values_only=True))

        # Data starts at row 4 (index 3). Year is in col A, propagated forward
        # (it only appears explicitly at Mid-Jan but the year applies to the
        # preceding months: e.g. rows before Jan 1998 are FY1997).
        # Better: carry year forward from last explicit year cell.
        current_year: int | None = None
        m1_found = m2_found = False

        for row in all_rows[3:]:  # skip 3 header rows
            yr_cell = row[0]
            mo_cell = row[1]
            m1_cell = row[2]
            m2_cell = row[4] if len(row) > 4 else None

            if yr_cell is not None and isinstance(yr_cell, (int, float)):
                current_year = int(yr_cell)

            if current_year is None:
                continue

            mo_key = _normalize_month_key(mo_cell)
            if mo_key is None:
                continue
            bs_month = _AD_MONTH_TO_BS.get(mo_key)
            if bs_month is None:
                continue

            bs_year = _bs_year_from_ad(current_year, mo_key)

            if m1_cell is not None:
                try:
                    m1_val = float(m1_cell)
                    rows.append(_make_row(
                        slug=M1_SLUG,
                        value=m1_val,
                        unit=UNIT,
                        bs_month=bs_month,
                        bs_year=bs_year,
                        pub_ad=pub_ad,
                    ))
                    m1_found = True
                except (ValueError, TypeError):
                    pass

            if m2_cell is not None:
                try:
                    m2_val = float(m2_cell)
                    rows.append(_make_row(
                        slug=M2_SLUG,
                        value=m2_val,
                        unit=UNIT,
                        bs_month=bs_month,
                        bs_year=bs_year,
                        pub_ad=pub_ad,
                    ))
                    m2_found = True
                except (ValueError, TypeError):
                    pass

        if not m1_found:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail="Monetary XLSX: M1 column (col C) had no parseable values",
            ))
        if not m2_found:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail="Monetary XLSX: M2 column (col E) had no parseable values",
            ))
    finally:
        wb.close()

    return rows


def _parse_interest_rates(
    path: Path,
    pub_ad: datetime,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse Structure-of-interest-rates-1.xlsx, sheet 'Historical Int. Rate Final'.

    Row 1: title
    Row 2 (index 1): year integers as column headers (from col B onwards)
    Row 3 (index 2): month names as column headers
    Row 34 (index 33): G. Weighted Average Deposit Rate
    Row 35 (index 34): H. Weighted Average Lending Rate
    """
    SHEET = "Historical Int. Rate Final"
    DEPOSIT_SLUG = "nrb-finsec-deposit-rate-monthly"
    LENDING_SLUG = "nrb-finsec-lending-rate-monthly"
    UNIT = "percent_per_annum"

    rows: list[StagingRowDraft] = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"Interest Rate XLSX open failed: {exc}",
        ))
        return rows

    try:
        if SHEET not in wb.sheetnames:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail=f"Interest Rate XLSX: sheet '{SHEET}' not found; found: {wb.sheetnames}",
            ))
            return rows

        ws = wb[SHEET]
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 35:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail=f"Interest Rate XLSX: expected ≥35 rows, got {len(all_rows)}",
            ))
            return rows

        year_row = all_rows[1]   # Row 2 (0-indexed: 1)
        month_row = all_rows[2]  # Row 3 (0-indexed: 2)
        deposit_row = all_rows[33]  # Row 34
        lending_row = all_rows[34]  # Row 35

        # Build col→(bs_month, bs_year) for all valid data columns.
        col_date: dict[int, tuple[BsMonth, int]] = {}
        current_year: int | None = None
        for c in range(1, len(year_row)):  # skip col A (row labels)
            yr_cell = year_row[c]
            if yr_cell is not None and isinstance(yr_cell, (int, float)):
                current_year = int(yr_cell)
            if current_year is None:
                continue
            mo_key = _normalize_month_key(month_row[c] if c < len(month_row) else None)
            if mo_key is None:
                continue
            bs_month = _AD_MONTH_TO_BS.get(mo_key)
            if bs_month is None:
                continue
            bs_year = _bs_year_from_ad(current_year, mo_key)
            col_date[c] = (bs_month, bs_year)

        if not col_date:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail="Interest Rate XLSX: no valid date columns found in rows 2-3",
            ))
            return rows

        dep_found = lend_found = False
        for c, (bs_month, bs_year) in col_date.items():
            dep_cell = deposit_row[c] if c < len(deposit_row) else None
            lend_cell = lending_row[c] if c < len(lending_row) else None

            if dep_cell is not None:
                try:
                    dep_val = float(dep_cell)
                    rows.append(_make_row(
                        slug=DEPOSIT_SLUG,
                        value=dep_val,
                        unit=UNIT,
                        bs_month=bs_month,
                        bs_year=bs_year,
                        pub_ad=pub_ad,
                    ))
                    dep_found = True
                except (ValueError, TypeError):
                    pass

            if lend_cell is not None:
                try:
                    lend_val = float(lend_cell)
                    rows.append(_make_row(
                        slug=LENDING_SLUG,
                        value=lend_val,
                        unit=UNIT,
                        bs_month=bs_month,
                        bs_year=bs_year,
                        pub_ad=pub_ad,
                    ))
                    lend_found = True
                except (ValueError, TypeError):
                    pass

        if not dep_found:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail="Interest Rate XLSX: deposit rate row (row 34) had no parseable values",
            ))
        if not lend_found:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail="Interest Rate XLSX: lending rate row (row 35) had no parseable values",
            ))
    finally:
        wb.close()

    return rows


def _parse_nepse(
    path: Path,
    pub_ad: datetime,
    errors: list[ParserError],
) -> list[StagingRowDraft]:
    """Parse NEPSE.xlsx, sheet NEPSE.

    Row 1: title
    Row 2: column headers — 'Fiscal Year', 'Aug', 'Sep', ..., 'Jul'
    Row 3+: data rows — col A = 'YYYY/YY' FY label, cols B-M = monthly index values.

    Per NRB note: 'August represents Mid August that represents end of month Shrawan.'
    Therefore col 'Aug' = Shrawan, 'Sep' = Bhadra, ..., 'Jul' = Ashadh.
    BS year for each cell:
        Aug–Dec (Shrawan–Mangsir): BS FY start year (e.g. '2025/26' → BS 2082)
        Jan–Jul (Poush–Ashadh): BS FY start year + 1
    """
    SHEET = "NEPSE"
    SLUG = "nrb-finsec-nepse-index-monthly"
    UNIT = "index_points"

    rows: list[StagingRowDraft] = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        errors.append(ParserError(
            error_class="EncodingError",
            error_detail=f"NEPSE XLSX open failed: {exc}",
        ))
        return rows

    try:
        if SHEET not in wb.sheetnames:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail=f"NEPSE XLSX: sheet '{SHEET}' not found; found: {wb.sheetnames}",
            ))
            return rows

        ws = wb[SHEET]
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 3:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail="NEPSE XLSX: fewer than 3 rows",
            ))
            return rows

        # Row 2 (index 1) has column headers: Fiscal Year, Aug, Sep, ... Jul
        header_row = all_rows[1]

        # Build col→(bs_month, bs_position) where bs_position 1-6=first half, 7-12=second half
        # First half of FY: Shrawan(Aug), Bhadra(Sep), Ashwin(Oct), Kartik(Nov), Mangsir(Dec), Poush(Jan)
        # Second half of FY: Magh(Feb), Falgun(Mar), Chait(Apr), Baisakh(May), Jestha(Jun), Ashadh(Jul)
        _FIRST_HALF_FY: frozenset[str] = frozenset({"aug", "sep", "oct", "nov", "dec", "jan"})

        col_month: dict[int, tuple[BsMonth, bool]] = {}  # col → (bs_month, is_first_half_fy)
        for c, cell in enumerate(header_row):
            if cell is None or c == 0:
                continue
            mo_key = str(cell).strip().lower()
            bs_month = _NEPSE_COL_TO_BS.get(mo_key)
            if bs_month is None:
                continue
            is_first_half = mo_key in _FIRST_HALF_FY
            col_month[c] = (bs_month, is_first_half)

        if not col_month:
            errors.append(ParserError(
                error_class="PageLayoutChanged",
                error_detail="NEPSE XLSX: no valid month columns found in header row",
            ))
            return rows

        found_any = False
        _FY_LABEL_RE = re.compile(r"^(\d{4})/(\d{2})$")

        for row in all_rows[2:]:  # data rows start at row 3 (index 2)
            fy_label = row[0]
            if fy_label is None:
                continue
            fy_str = str(fy_label).strip()
            m = _FY_LABEL_RE.match(fy_str)
            if not m:
                continue  # source/notes rows at the bottom

            # FY label 'YYYY/YY' e.g. '2025/26' → AD FY start = 2025 → BS FY start = 2025+57 = 2082
            ad_fy_start = int(m.group(1))
            bs_fy_start = ad_fy_start + 57

            for c, (bs_month, is_first_half_fy) in col_month.items():
                cell_val = row[c] if c < len(row) else None
                if cell_val is None:
                    continue
                try:
                    value = float(cell_val)
                except (ValueError, TypeError):
                    continue

                # BS year: first half of FY (Aug-Jan) → bs_fy_start; second half → bs_fy_start+1
                bs_year = bs_fy_start if is_first_half_fy else bs_fy_start + 1

                rows.append(_make_row(
                    slug=SLUG,
                    value=value,
                    unit=UNIT,
                    bs_month=bs_month,
                    bs_year=bs_year,
                    pub_ad=pub_ad,
                ))
                found_any = True

        if not found_any:
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail="NEPSE XLSX: no parseable data rows found",
            ))
    finally:
        wb.close()

    return rows


# ── Top-level parse() ─────────────────────────────────────────────────────────

_TARGET_SLUGS: Final[frozenset[str]] = frozenset({
    "nrb-finsec-loans-realestate-monthly",
    "nrb-finsec-loans-agriculture-monthly",
    "nrb-finsec-loans-manufacturing-monthly",
    "nrb-finsec-nepse-index-monthly",
    "nrb-finsec-m2-level-monthly",
    "nrb-finsec-m1-level-monthly",
    "nrb-finsec-lending-rate-monthly",
    "nrb-finsec-deposit-rate-monthly",
})

# File name patterns to locate each file when given a directory.
_FILE_PATTERN_LOANS = re.compile(r"loans.*sectorwise.*\.xlsx$", re.IGNORECASE)
_FILE_PATTERN_MONETARY = re.compile(r"monetary.?survey.*\.xlsx$", re.IGNORECASE)
_FILE_PATTERN_INTEREST = re.compile(r"structure.?of.?interest.?rates.*\.xlsx$", re.IGNORECASE)
_FILE_PATTERN_NEPSE = re.compile(r"nepse.*\.xlsx$", re.IGNORECASE)


def _find_file(directory: Path, pattern: re.Pattern[str]) -> Path | None:
    """Find a file in directory matching pattern. Returns None if not found."""
    for f in directory.iterdir():
        if f.is_file() and pattern.match(f.name):
            return f
    return None


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse NRB Financial Sector XLSX files from a directory.

    Arguments:
        source_document_path: filesystem path to the DIRECTORY containing the
            four XLSX files, or path to a single XLSX file (in which case only
            that file's indicators are extracted).
        source_document_id: opaque FK from source_documents; threaded through
            for orchestrator symmetry.

    Returns:
        ParserResult with status, staging_rows, errors.
    """
    _ = source_document_id

    input_path = Path(source_document_path)

    if not input_path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source path not found: {input_path}",
            )],
        )

    # Determine if we were given a directory or a single file.
    if input_path.is_dir():
        directory = input_path
        loans_path = _find_file(directory, _FILE_PATTERN_LOANS)
        monetary_path = _find_file(directory, _FILE_PATTERN_MONETARY)
        interest_path = _find_file(directory, _FILE_PATTERN_INTEREST)
        nepse_path = _find_file(directory, _FILE_PATTERN_NEPSE)
    else:
        # Single file: infer which one it is.
        directory = input_path.parent
        name = input_path.name
        loans_path = input_path if _FILE_PATTERN_LOANS.match(name) else None
        monetary_path = input_path if _FILE_PATTERN_MONETARY.match(name) else None
        interest_path = input_path if _FILE_PATTERN_INTEREST.match(name) else None
        nepse_path = input_path if _FILE_PATTERN_NEPSE.match(name) else None

    # Publication date: use fixture mtime as heuristic.
    try:
        mtime = input_path.stat().st_mtime
        pub_ad = datetime.fromtimestamp(mtime, tz=UTC)
    except OSError:
        pub_ad = datetime(2026, 1, 31, tzinfo=UTC)  # fallback

    errors: list[ParserError] = []
    staging_rows: list[StagingRowDraft] = []

    if loans_path is not None:
        staging_rows.extend(_parse_loans(loans_path, pub_ad, errors))
    else:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="Loans-of-the-BFIs-Sectorwise.xlsx not found in source directory",
        ))

    if monetary_path is not None:
        staging_rows.extend(_parse_monetary(monetary_path, pub_ad, errors))
    else:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="Monetary-survey.xlsx not found in source directory",
        ))

    if interest_path is not None:
        staging_rows.extend(_parse_interest_rates(interest_path, pub_ad, errors))
    else:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="Structure-of-interest-rates-1.xlsx not found in source directory",
        ))

    if nepse_path is not None:
        staging_rows.extend(_parse_nepse(nepse_path, pub_ad, errors))
    else:
        errors.append(ParserError(
            error_class="ColumnMissing",
            error_detail="NEPSE.xlsx not found in source directory",
        ))

    if not staging_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors or [ParserError(
                error_class="PageLayoutChanged",
                error_detail="no indicators extracted from any XLSX file",
            )],
        )

    found_slugs = {r.indicator_slug_raw for r in staging_rows}
    missing_slugs = _TARGET_SLUGS - found_slugs
    if missing_slugs:
        for slug in sorted(missing_slugs):
            errors.append(ParserError(
                error_class="ColumnMissing",
                error_detail=f"target slug '{slug}' produced no rows",
            ))

    status: ParserStatus = "success" if not errors else "partial"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


def _main() -> None:
    """CLI entrypoint used by the Node ingestion orchestrator.

    Argv: ``parser.py <source_document_path> <source_document_id>``.
    Writes JSON to stdout.
    Exit codes:
      0: parser ran (status may still be 'failure'; consumer reads stdout)
      2: usage error
      1: catastrophic crash (let Python propagate)
    """
    import json
    import sys

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    json.dump(result.to_json_dict(), sys.stdout)


if __name__ == "__main__":
    _main()

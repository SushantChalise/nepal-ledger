"""Pytest fixtures for the NRB DNE parser.

Builds synthetic XLSX fixtures via openpyxl — regenerated each run, not
committed as binaries. Fixtures mimic the External Sector "Foreign Exchange
Reserves" DNE file layout (wide format: indicators as rows, FY periods as
columns).

Fixture variants:
    happy_path_xlsx      — valid 3-indicator annual + 2 monthly columns
    empty_workbook_xlsx  — workbook with a blank sheet only
    ambiguous_unit_xlsx  — no unit in preamble; ambiguous row
    bad_period_xlsx      — one unparseable period header column
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"


def _ensure_fixture_dir() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------


def _build_happy_path(path: Path) -> None:
    """External-sector Foreign Exchange Reserves style XLSX.

    Layout (1-indexed openpyxl rows):
      Row 1: title "Foreign Exchange Reserves (in million US$)"
      Row 2: header periods — "2080/81", "2081/82", "2082/83", "Shrawan 2082", "Bhadra 2082"
      Row 3: "Total Foreign Exchange Reserves"   1500.0  2100.0  2300.0  2250.0  2280.0
      Row 4: "Gold Reserves"                      100.0   120.0   130.0   125.0   127.0
      Row 5: "Foreign Currency Assets"            1400.0  1980.0  2170.0  2125.0  2153.0
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "External Sector"

    ws.cell(row=1, column=1, value="Foreign Exchange Reserves (in million US$)")
    # Header row.
    ws.cell(row=2, column=1, value="Indicators")
    ws.cell(row=2, column=2, value="2080/81")
    ws.cell(row=2, column=3, value="2081/82")
    ws.cell(row=2, column=4, value="2082/83")
    ws.cell(row=2, column=5, value="Shrawan 2082")
    ws.cell(row=2, column=6, value="Bhadra 2082")

    # Data rows.
    indicators = [
        ("Total Foreign Exchange Reserves", [1500.0, 2100.0, 2300.0, 2250.0, 2280.0]),
        ("Gold Reserves", [100.0, 120.0, 130.0, 125.0, 127.0]),
        ("Foreign Currency Assets", [1400.0, 1980.0, 2170.0, 2125.0, 2153.0]),
    ]
    for row_offset, (label, values) in enumerate(indicators):
        r = 3 + row_offset
        ws.cell(row=r, column=1, value=label)
        for col_offset, val in enumerate(values):
            ws.cell(row=r, column=2 + col_offset, value=val)

    wb.save(str(path))


def _build_empty_workbook(path: Path) -> None:
    """Workbook with a single blank sheet — no parseable data."""
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Empty"
    # Write only a title row — no period header.
    ws.cell(row=1, column=1, value="This sheet has no data")
    wb.save(str(path))


def _build_ambiguous_unit(path: Path) -> None:
    """XLSX with no unit annotation in preamble and no unit-bearing sheet name.

    The parser should emit UnitAmbiguous and still produce rows.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # No unit hint row — just jump to the header.
    ws.cell(row=1, column=1, value="Indicator")
    ws.cell(row=1, column=2, value="2081/82")
    ws.cell(row=1, column=3, value="2082/83")

    ws.cell(row=2, column=1, value="Remittance Inflows")
    ws.cell(row=2, column=2, value=950.0)
    ws.cell(row=2, column=3, value=1050.0)

    wb.save(str(path))


def _build_bad_period(path: Path) -> None:
    """XLSX with one valid period column and one malformed period column.

    Malformed: "2082/invalid" — looks like a period (contains "2082") but
    won't parse; should emit PeriodUnparseable.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fiscal Sector"

    ws.cell(row=1, column=1, value="Government Revenue (Rs. in million)")
    ws.cell(row=2, column=1, value="Indicator")
    ws.cell(row=2, column=2, value="2081/82")       # valid
    ws.cell(row=2, column=3, value="2082/invalid")  # malformed

    ws.cell(row=3, column=1, value="Tax Revenue")
    ws.cell(row=3, column=2, value=500000.0)
    ws.cell(row=3, column=3, value=550000.0)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Pytest fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def happy_path_xlsx() -> Path:
    p = FIXTURE_DIR / "happy_path.xlsx"
    if not p.exists():
        _build_happy_path(p)
    return p


@pytest.fixture(scope="session")
def empty_workbook_xlsx() -> Path:
    p = FIXTURE_DIR / "empty_workbook.xlsx"
    if not p.exists():
        _build_empty_workbook(p)
    return p


@pytest.fixture(scope="session")
def ambiguous_unit_xlsx() -> Path:
    p = FIXTURE_DIR / "ambiguous_unit.xlsx"
    if not p.exists():
        _build_ambiguous_unit(p)
    return p


@pytest.fixture(scope="session")
def bad_period_xlsx() -> Path:
    p = FIXTURE_DIR / "bad_period.xlsx"
    if not p.exists():
        _build_bad_period(p)
    return p


def _build_bs_fy_suffix(path: Path) -> None:
    """Fixture: BS FY column headers with NRB revision/provisional suffixes.

    Mirrors the real BoP BPM6 pattern but using BS-era years (2079/80R,
    2080/81P, 2081/82) so the parser can detect them.

    Layout:
      Row 1: title "External Debt (Rs. in million)"
      Row 2: "Indicator", "2079/80R", "2080/81P", "2081/82"
      Row 3: "Total External Debt",  5000.0,  5500.0,  6000.0
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "External Debt"
    ws.cell(row=1, column=1, value="External Debt (Rs. in million)")
    ws.cell(row=2, column=1, value="Indicator")
    ws.cell(row=2, column=2, value="2079/80R")   # revised suffix
    ws.cell(row=2, column=3, value="2080/81P")   # provisional suffix
    ws.cell(row=2, column=4, value="2081/82")    # plain
    ws.cell(row=3, column=1, value="Total External Debt")
    ws.cell(row=3, column=2, value=5000.0)
    ws.cell(row=3, column=3, value=5500.0)
    ws.cell(row=3, column=4, value=6000.0)
    wb.save(str(path))


def _build_ad_year_sheet(path: Path) -> None:
    """Fixture: AD-calendar-year FY column headers (2021/22, 2022/23).

    Mirrors the Migrant Workers / Foreign Trade real-file pattern where NRB
    uses AD fiscal years instead of BS. The parser should emit PeriodUnparseable
    with an explicit AD-year diagnostic rather than a bare NoDataExtracted.

    Layout:
      Row 1: "Migrant Workers by Country"
      Row 2: blank
      Row 3: "Country", "2021/22", "2022/23"
      Row 4: "Qatar", 45000, 50000
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Country"
    ws.cell(row=1, column=1, value="Migrant Workers by Country")
    ws.cell(row=3, column=1, value="Country")
    ws.cell(row=3, column=2, value="2021/22")
    ws.cell(row=3, column=3, value="2022/23")
    ws.cell(row=4, column=1, value="Qatar")
    ws.cell(row=4, column=2, value=45000)
    ws.cell(row=4, column=3, value=50000)
    wb.save(str(path))


@pytest.fixture(scope="session")
def bs_fy_suffix_xlsx() -> Path:
    p = FIXTURE_DIR / "bs_fy_suffix.xlsx"
    if not p.exists():
        _build_bs_fy_suffix(p)
    return p


@pytest.fixture(scope="session")
def ad_year_sheet_xlsx() -> Path:
    p = FIXTURE_DIR / "ad_year_sheet.xlsx"
    if not p.exists():
        _build_ad_year_sheet(p)
    return p


# ---------------------------------------------------------------------------
# v0.4.0 fixtures — non-standard AD layouts (ADR-0013 follow-up)
# ---------------------------------------------------------------------------


def _build_year_month_header(path: Path) -> None:
    """Fixture: the Foreign-exchange-reserves two-row monthly header.

    Layout (1-indexed openpyxl rows; data starts at col 3 = openpyxl col index 3):
      Row 1: title "Gross Foreign Assets of the Banking Sector"
      Row 2: unit "(Rs in Million)"
      Row 3: cols 3..: integer AD YEARS — 2001 spans Aug-Dec, 2002 spans Jan-Mar.
             The year cell appears only on the first month of each year-block
             (sparse) to exercise the forward-fill; remaining cells blank.
      Row 4: cols 3..: AD MONTH names (mixed abbreviated/full) — Aug, Sep, Oct,
             Nov, Dec, Jan, Feb, March. (8 months ≥ _MIN_MONTH_HEADER_CELLS=6.)
      Rows 5-6: indicator label in col 1 (a sub-item label in col 2 is joined),
             values across the month columns.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "FX Reserves"
    ws.cell(row=1, column=1, value="Gross Foreign Assets of the Banking Sector")
    ws.cell(row=2, column=1, value="(Rs in Million)")
    # Year row (row 3): partially populated — enough year cells to pass detection
    # (≥ _MIN_YEAR_HEADER_CELLS), with gaps to exercise the forward-fill.
    # Columns 3..10 carry 8 monthly periods: Aug2001..Mar2002.
    ws.cell(row=3, column=3, value=2001)  # over "Aug"
    ws.cell(row=3, column=4, value=2001)  # over "Sep"
    # cols 5-7 (Oct, Nov, Dec) blank → forward-filled to 2001.
    ws.cell(row=3, column=8, value=2002)  # over "Jan"
    ws.cell(row=3, column=9, value=2002)  # over "Feb"
    # col 10 (March) blank → forward-filled to 2002.
    # Month row (row 4): mixed abbreviated/full names.
    months = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "March"]
    for i, m in enumerate(months):
        ws.cell(row=4, column=3 + i, value=m)
    # Data rows.
    rows_data = [
        ("A. Nepal Rastra Bank", None, [100.0, 110.0, 120.0, 130.0, 140.0, 150.0, 160.0, 170.0]),
        ("Gold Reserves", None, [10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0]),
    ]
    for r_off, (label, sub, vals) in enumerate(rows_data):
        r = 5 + r_off
        ws.cell(row=r, column=1, value=label)
        if sub is not None:
            ws.cell(row=r, column=2, value=sub)
        for i, v in enumerate(vals):
            ws.cell(row=r, column=3 + i, value=v)
    wb.save(str(path))


def _build_year_month_dup(path: Path) -> None:
    """Fixture: two-row monthly header with a REPEATED (year, month) column.

    Mirrors the real FX-reserves quirk of two adjacent "Oct 2025" columns holding
    different values. The parser must emit both (no data loss), flag the duplicate
    in parser_notes, and emit a single PeriodAmbiguous error.

    Months row: Aug Sep Oct Nov Dec Jan Oct  (the trailing "Oct" repeats Oct 2025).
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "FX Reserves"
    ws.cell(row=1, column=1, value="Gross Foreign Assets (Rs in Million)")
    # Year row: 2025 on the first three month cells (≥ _MIN_YEAR_HEADER_CELLS),
    # remaining cells blank → forward-filled to 2025 across the whole row.
    ws.cell(row=3, column=3, value=2025)
    ws.cell(row=3, column=4, value=2025)
    ws.cell(row=3, column=5, value=2025)
    months = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Oct"]
    for i, m in enumerate(months):
        ws.cell(row=4, column=3 + i, value=m)
    ws.cell(row=5, column=1, value="Total Reserves")
    # The two "Oct" columns (index 2 and 6) carry DIFFERENT values.
    vals = [100.0, 200.0, 300.0, 400.0, 500.0, 600.0, 999.0]
    for i, v in enumerate(vals):
        ws.cell(row=5, column=3 + i, value=v)
    wb.save(str(path))


def _build_long_panel(path: Path) -> None:
    """Fixture: the Exchange-rate long panel.

    Layout:
      Row 1: title
      Row 2-3: sub-header rows naming the value columns
      Row 4: "Fiscal Year", "Month", "Buying", "Selling", "Middle Rate"
      Row 5..: FY in col 1 (sparse — only first month of each FY), AD month in
               col 2, three numeric value columns. Includes one aggregate
               "Annual Average" row that must be SKIPPED (not a calendar month).
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Time series"
    ws.cell(row=1, column=3, value="Nepalese Exchange Rate (Rupee per USD)")
    ws.cell(row=2, column=3, value="Month End")
    ws.cell(row=4, column=1, value="Fiscal Year")
    ws.cell(row=4, column=2, value="Month")
    ws.cell(row=4, column=3, value="Buying")
    ws.cell(row=4, column=4, value="Selling")
    ws.cell(row=4, column=5, value="Middle Rate")
    # FY 2022/23: July (AD 2022) .. August (AD 2022); FY label only on first month.
    panel = [
        ("2022/23", "July", 130.0, 130.6, 130.3),
        (None, "August", 131.0, 131.6, 131.3),
        (None, "Annual Average", 130.5, 131.1, 130.8),  # aggregate → skipped
        ("2023/24", "January", 132.0, 132.6, 132.3),  # AD 2024 (Jan of FY 2023/24)
    ]
    for r_off, (fy, month, buy, sell, mid) in enumerate(panel):
        r = 5 + r_off
        if fy is not None:
            ws.cell(row=r, column=1, value=fy)
        ws.cell(row=r, column=2, value=month)
        ws.cell(row=r, column=3, value=buy)
        ws.cell(row=r, column=4, value=sell)
        ws.cell(row=r, column=5, value=mid)
    wb.save(str(path))


def _build_transposed(path: Path) -> None:
    """Fixture: the Tourist-arrivals transposed layout.

    Layout:
      Row 1: title "Tourist Arrivals"
      Row 2: "Year", "Jan", "Feb", "Mar", ... "Dec", "Total"
      Row 3..: integer AD year in col 1, 12 monthly values + an annual Total
               column (which must be ignored — it is not a month).
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Tourist Arrival"
    ws.cell(row=1, column=1, value="Tourist Arrivals (Number)")
    header = ["Year", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"]
    for i, h in enumerate(header):
        ws.cell(row=2, column=1 + i, value=h)
    years = [
        (1992, [100, 110, 120, 130, 140, 150, 160, 170, 180, 190, 200, 210], 1860),
        (1993, [101, 111, 121, 131, 141, 151, 161, 171, 181, 191, 201, 211], 1872),
    ]
    for r_off, (yr, vals, total) in enumerate(years):
        r = 3 + r_off
        ws.cell(row=r, column=1, value=yr)
        for i, v in enumerate(vals):
            ws.cell(row=r, column=2 + i, value=v)
        ws.cell(row=r, column=14, value=total)  # Total column — must be ignored
    wb.save(str(path))


def _build_foreign_trade_commodities(path: Path) -> None:
    """Fixture: the Foreign-Trade "Export Import Major Commodities" matrix.

    Two sections (one Export, one Import), each a wide MONTHLY panel: a section
    title in col 0, an "S.No. | Major Commodities | <FY>" header row whose FY
    label is sparse (only the first column of each 12-month block), a repeating
    AD month-name row (Aug → next Jul), then commodity rows (S.No. col 0, label
    col 1, monthly values from col 2).

    Two fiscal-year blocks (2012/2013, 2013/2014) × 12 months = 24 value columns,
    so columns 2..25. To exercise the structural FY-advance fix, the SECOND block's
    FY label cell is left BLANK (a merged-cell artifact in the real file) — the
    parser must still place it in 2013/14, not collapse it onto 2012/13.

    A trailing "TOTAL" row per section must be skipped. Commodity labels include
    "G.I. pipe" (whose "G.I." must NOT be stripped as an enumerator — it
    distinguishes it from "M.S. Pipe") to lock in the dimension-slug behaviour.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Export Import Major Commodities"
    months = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan",
              "Feb", "Mar", "Apr", "May", "Jun", "Jul"]

    def _write_section(start_row: int, title: str, commodities: list[tuple[str, float]]) -> int:
        r = start_row
        ws.cell(row=r, column=1, value=title)
        r += 1
        # Header row: S.No. | Major Commodities | FY (sparse: block 1 only).
        ws.cell(row=r, column=1, value="S.No.")
        ws.cell(row=r, column=2, value="Major Commodities")
        ws.cell(row=r, column=3, value="2012/2013")     # block 1 FY (col 3)
        # block 2 FY cell (col 15) deliberately BLANK → structural advance.
        header_row = r
        r += 1
        # Month row: 24 columns (2 blocks × 12 months), cols 3..26.
        for i in range(24):
            ws.cell(row=r, column=3 + i, value=months[i % 12])
        r += 1
        # Commodity rows: same value in every month column (value = base + month).
        for sno, (label, base_val) in enumerate(commodities, start=1):
            ws.cell(row=r, column=1, value=sno)
            ws.cell(row=r, column=2, value=label)
            for i in range(24):
                ws.cell(row=r, column=3 + i, value=base_val + i)
            r += 1
        # Trailing TOTAL row (must be skipped).
        ws.cell(row=r, column=2, value="TOTAL")
        for i in range(24):
            ws.cell(row=r, column=3 + i, value=9999.0)
        r += 2  # blank spacer row after the section
        _ = header_row
        return r

    next_row = _write_section(
        1,
        "Export of Major Commodities to India",
        [("Cardamom", 100.0), ("G.I. pipe", 10.0), ("M.S. Pipe", 20.0)],
    )
    _write_section(
        next_row,
        "Import of Major Commodities from China",
        [("Crude Soyabean Oil", 500.0)],
    )
    wb.save(str(path))


def _build_fx_reserve_slug(path: Path) -> None:
    """Fixture: an FX-reserves-shaped sheet that exercises slug cleanup (v0.5.0).

    Two-row monthly header (integer AD years over AD month names), with row labels
    that carry outline enumerators and a sub-label that repeats across sections —
    the exact shapes that previously produced artifact slugs:
      "A. Nepal Rastra Bank (1+2)"  → dne-nepal-rastra-bank        (enum + (1+2) stripped)
      "C. Gross Foreign Exchange…"  → dne-gross-foreign-exchange-reserve
      "Convertible" (under A.)      → dne-convertible
      "Convertible" (under C.)      → dne-convertible-gross-foreign-exchange-reserve
                                       (collision qualified by section parent, NOT -rNN)
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "FX Reserves"
    ws.cell(row=1, column=1, value="Gross Foreign Assets of the Banking Sector")
    ws.cell(row=2, column=1, value="(Rs in Million)")
    # Year row (row 3) + month row (row 4): 6 months Aug..Jan 2025, cols 3..8.
    for c in (3, 4, 5):
        ws.cell(row=3, column=c, value=2025)
    months = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
    for i, m in enumerate(months):
        ws.cell(row=4, column=3 + i, value=m)
    # Data rows: enumerated parents in col 0, repeating sub-labels in col 1.
    body = [
        ("A. Nepal Rastra Bank (1+2)", None, 100.0),
        (None, "Convertible", 60.0),                       # under A.
        ("C. Gross Foreign Exchange Reserve", None, 150.0),
        (None, "Convertible", 90.0),                       # under C. → collision
    ]
    for r_off, (col0, col1, val) in enumerate(body):
        r = 5 + r_off
        if col0 is not None:
            ws.cell(row=r, column=1, value=col0)
        if col1 is not None:
            ws.cell(row=r, column=2, value=col1)
        for i in range(6):
            ws.cell(row=r, column=3 + i, value=val + i)
    wb.save(str(path))


@pytest.fixture(scope="session")
def foreign_trade_commodities_xlsx() -> Path:
    # Named "Foreign-Trade.xlsx" so parse_dne's filename-stem dispatch
    # (_DIMENSIONAL_FILE_STEMS) routes it to the dimensional path end-to-end.
    p = FIXTURE_DIR / "Foreign-Trade.xlsx"
    if not p.exists():
        _build_foreign_trade_commodities(p)
    return p


@pytest.fixture(scope="session")
def fx_reserve_slug_xlsx() -> Path:
    p = FIXTURE_DIR / "fx_reserve_slug.xlsx"
    if not p.exists():
        _build_fx_reserve_slug(p)
    return p


@pytest.fixture(scope="session")
def year_month_header_xlsx() -> Path:
    p = FIXTURE_DIR / "year_month_header.xlsx"
    if not p.exists():
        _build_year_month_header(p)
    return p


@pytest.fixture(scope="session")
def year_month_dup_xlsx() -> Path:
    p = FIXTURE_DIR / "year_month_dup.xlsx"
    if not p.exists():
        _build_year_month_dup(p)
    return p


@pytest.fixture(scope="session")
def long_panel_xlsx() -> Path:
    p = FIXTURE_DIR / "long_panel.xlsx"
    if not p.exists():
        _build_long_panel(p)
    return p


@pytest.fixture(scope="session")
def transposed_xlsx() -> Path:
    p = FIXTURE_DIR / "transposed.xlsx"
    if not p.exists():
        _build_transposed(p)
    return p


# ---------------------------------------------------------------------------
# v0.6.0 fixtures — real-sector annual column-series + Provincial-GDP dimensional
# ---------------------------------------------------------------------------


def _build_national_accounts(path: Path) -> None:
    """Fixture: the National-Accounts annual column-series layout (GDP headline).

    Mirrors the real file's "GDP Series_Nominal" and "GDP Series_Real" sheets:
    annual FY labels stacked DOWN col 0 (a "Year" column) with named-indicator value
    columns to the right, and a 2-3 row header above the first FY row. Includes:
      - "GDP Series_Nominal": col1 "Nominal GDP (Rs. in billion)" (allowlisted) +
        an "As Percent of GDP" sub-column (col2, NOT allowlisted → must be ignored).
      - "GDP Series_Real": col1 real-GDP growth %, col6 real GDP (Rs. in billion),
        col9 per-capita GDP (USD), col12 GDP deflator — all allowlisted.
      - A trailing "Source:" footer row inside the data block (must be skipped, not
        error). One FY label carries a "R"/"P" revision suffix to exercise stripping.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    nom = wb.active
    assert nom is not None
    nom.title = "GDP Series_Nominal"
    nom.cell(row=1, column=1, value="Real Sector Indicators")
    nom.cell(row=3, column=1, value="Year")  # openpyxl row 3 == rows index 2
    nom.cell(row=3, column=2, value="Nominal GDP (Rs. in billion)")
    nom.cell(row=3, column=3, value="As Percent of GDP*")  # NOT allowlisted
    # Data rows: FY label col1, nominal GDP col2, a percent col3 (ignored).
    nom_data = [
        ("2079/80", 5366.99, 24.0),
        ("2080/81R", 5709.09, 24.7),  # revision suffix → stripped to 2080/81
        ("2081/82P", 6107.22, 25.1),
    ]
    for r_off, (fy, gdp, pct) in enumerate(nom_data):
        r = 5 + r_off
        nom.cell(row=r, column=1, value=fy)
        nom.cell(row=r, column=2, value=gdp)
        nom.cell(row=r, column=3, value=pct)
    nom.cell(row=5 + len(nom_data), column=1, value="Source: Various issues")  # footer

    real = wb.create_sheet("GDP Series_Real")
    real.cell(row=1, column=1, value="Real Sector Indicators")
    real.cell(row=3, column=1, value="Year")
    real.cell(row=3, column=2, value="Real GDP Growth Rate \n(at purchasers'  price)")
    real.cell(row=3, column=7, value="Real GDP \n(at purchasers' price) \n(Rs. in billion)")
    real.cell(row=3, column=10, value="Per Capita GDP \n(in USD)")
    real.cell(row=3, column=13, value="GDP Deflator2")
    # Columns: 1=FY, 2=growth%, 7=real GDP bn, 10=per-capita USD, 13=deflator.
    # ≥3 FY rows so the annual column-series anchor (≥3 consecutive FY in col 0) fires.
    real_data = [
        ("2078/79", 1.983, 2530.0, 1390.0, 110.0),
        ("2079/80", 3.665, 2580.0, 1443.46, 115.37),
        ("2080/81R", 4.606, 2674.39, 1496.21, 124.64),
    ]
    for r_off, (fy, growth, rgdp, pc, defl) in enumerate(real_data):
        r = 5 + r_off
        real.cell(row=r, column=1, value=fy)
        real.cell(row=r, column=2, value=growth)
        real.cell(row=r, column=7, value=rgdp)
        real.cell(row=r, column=10, value=pc)
        real.cell(row=r, column=13, value=defl)
    wb.save(str(path))


def _build_cpi(path: Path) -> None:
    """Fixture: the Consumer-Price-Index annual column-series layout.

    Mirrors "CPI_National": FY labels down col 0, with a 2-row header — row "Index"
    over "Overall" (col1, allowlisted → dne-cpi) and "Percentage Change" over
    "Overall" (col4, allowlisted → dne-inflation-rate). Sub-group columns are not
    present here (they would not be allowlisted). One blank trailing row.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "CPI_National"
    ws.cell(row=1, column=1, value="National Consumer Price Index")
    ws.cell(row=2, column=1, value="Base Year : 2014/15 = 100")
    ws.cell(row=5, column=1, value="Fiscal Year")  # openpyxl row 5 == index 4
    ws.cell(row=5, column=2, value="Index")
    ws.cell(row=5, column=5, value="Percentage Change")
    ws.cell(row=6, column=2, value="Overall")
    ws.cell(row=6, column=5, value="Overall")
    cpi_data = [
        ("2079/80", 157.64, 7.74),
        ("2080/81", 166.22, 5.44),
        ("2081/82", 175.0, 5.28),
    ]
    for r_off, (fy, idx, chg) in enumerate(cpi_data):
        r = 7 + r_off
        ws.cell(row=r, column=1, value=fy)
        ws.cell(row=r, column=2, value=idx)
        ws.cell(row=r, column=5, value=chg)
    wb.save(str(path))


def _build_provincial_gdp(path: Path) -> None:
    """Fixture: the Provincial-GDP banner layout (GDP by province, ADR-0015).

    Mirrors the real "Tables" sheet Table 1 (nominal, current prices) EXACTLY in
    stride: each province block spans 7 consecutive FY columns under a province-name
    banner, with a BS-FY row then an AD-FY row beneath, industry rows, and a
    "Gross Domestic Product (GDP)" total row per province. Two provinces + a trailing
    "Total GVA" block that MUST be excluded (not a province):
      cols  3-9  : Koshi      (7 FY cols)
      cols 10-16 : Bagamati   (7 FY cols)
      cols 17-23 : Total GVA  (excluded)
    The "(GDP)" total row is preceded by a "...at basic prices" and a "Taxes..." row
    to confirm the substring match picks the right (headline) row. To keep magnitude
    assertions tied to specific FYs, only the first two FY columns of each block carry
    distinct asserted values; the rest are filled with plausible (non-colliding) data.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Tables"
    ws.cell(row=1, column=1, value="Table 1: Provincial Annual GDP")
    ws.cell(row=2, column=1, value="(at current prices, in million)")
    span = 7  # FY columns per province block — matches the real file
    blocks = [(3, "Koshi"), (3 + span, "Bagamati"), (3 + 2 * span, "Total GVA")]
    # Banner row (openpyxl row 4 == index 3).
    ws.cell(row=4, column=2, value="Industrial Classification")  # ignored banner
    for start, name in blocks:
        ws.cell(row=4, column=start, value=name)
    # BS FY row (row 5) + AD FY row (row 6): 7 distinct FY per block.
    bs = ["2075/76", "2076/77", "2077/78", "2078/79", "2079/80", "2080/81R", "2081/82P"]
    ad = ["2018/19", "2019/20", "2020/21", "2021/22", "2022/23", "2023/24", "2024/25"]
    for start, _name in blocks:
        for i in range(span):
            ws.cell(row=5, column=start + i, value=bs[i])
            ws.cell(row=6, column=start + i, value=ad[i])

    # The headline GDP-total row's asserted values for FY index 5 (2080/81R) and 6
    # (2081/82P), per block: Koshi, Bagamati, Total GVA.
    headline_5 = {"Koshi": 908830.0, "Bagamati": 1964517.0, "Total GVA": 2873000.0}
    headline_6 = {"Koshi": 970743.0, "Bagamati": 2100000.0, "Total GVA": 3070000.0}
    basic_5 = {"Koshi": 804000.0, "Bagamati": 1700000.0, "Total GVA": 2504000.0}

    def _block_vals(start_val: float) -> list[float]:
        # 7 ascending, distinct values for the 5 filler FY columns + 2 asserted.
        return [start_val + i * 1000.0 for i in range(span)]

    # Build body rows. For the headline + basic rows we override FY cols 5 & 6.
    def _row_for(base: float, override5: dict[str, float] | None,
                 override6: dict[str, float] | None) -> None:
        for start, name in blocks:
            vals = _block_vals(base)
            if override5 is not None:
                vals[5] = override5[name]
            if override6 is not None:
                vals[6] = override6[name]
            for i, v in enumerate(vals):
                ws.cell(row=_row_for.r, column=start + i, value=v)  # type: ignore[attr-defined]

    _row_for.r = 7  # type: ignore[attr-defined]
    ws.cell(row=7, column=2, value="Agriculture, forestry and fishing")
    _row_for(150000.0, None, None)
    _row_for.r = 8  # type: ignore[attr-defined]
    ws.cell(row=8, column=2, value="Gross Domestic Product  (GDP) at basic prices")
    _row_for(700000.0, basic_5, None)
    _row_for.r = 9  # type: ignore[attr-defined]
    ws.cell(row=9, column=2, value="Taxes less subsidies on products")
    _row_for(100000.0, None, None)
    _row_for.r = 10  # type: ignore[attr-defined]
    ws.cell(row=10, column=2, value="Gross Domestic Product (GDP)")
    _row_for(800000.0, headline_5, headline_6)
    wb.save(str(path))


@pytest.fixture(scope="session")
def national_accounts_xlsx() -> Path:
    # Named "National-Accounts.xlsx" so parse()'s _REAL_SECTOR_FILE_STEMS dispatch
    # routes it to the real-sector single-series path end-to-end.
    p = FIXTURE_DIR / "National-Accounts.xlsx"
    if not p.exists():
        _build_national_accounts(p)
    return p


@pytest.fixture(scope="session")
def cpi_xlsx() -> Path:
    p = FIXTURE_DIR / "Consumer-Price-Index.xlsx"
    if not p.exists():
        _build_cpi(p)
    return p


@pytest.fixture(scope="session")
def provincial_gdp_xlsx() -> Path:
    # Named "Provincial-GDP-2024-25.xlsx" so parse_dne's _DIMENSIONAL_FILE_STEMS
    # dispatch routes it to the province dimensional path end-to-end.
    p = FIXTURE_DIR / "Provincial-GDP-2024-25.xlsx"
    if not p.exists():
        _build_provincial_gdp(p)
    return p


# ---------------------------------------------------------------------------
# v0.7.0 fixture — Migrant-Workers-Remittance "Country" sheet (HEADCOUNT, ADR-0015)
# ---------------------------------------------------------------------------


def _build_migrant_workers(path: Path) -> None:
    """Fixture: the Migrant-Workers-Remittance "Country" sheet (HEADCOUNT matrix).

    Mirrors the real 3-row header EXACTLY in stride: a sparse fiscal-year banner
    (row 3 == index 2) at the head of each 12-month×3-subcol block, an AD month label
    ("Mid-Aug" … "Mid-Jul", at each group's first/Male column) on row 4 (== index 3),
    and a "Male"/"Female"/"Total" sub-header on row 5 (== index 4). Country rows start
    on row 6 (== index 5); the value read for a fact is the group's "Total" column.

    To keep the suite fast, this fixture uses TWO month groups per FY (Mid-Aug,
    Mid-Sep) across TWO fiscal years (2021/22, 2022/23), plus — for FY2022/23 — a
    THIRD group whose label is a DUPLICATE "Mid-Aug" (the real file's stray-month
    mislabel), to exercise the PeriodAmbiguous path without dropping data.

    Layout (openpyxl 1-indexed cols; each group = [Male, Female, Total]):
      FY 2021/22 : cols  2-4  (Mid-Aug),  5-7  (Mid-Sep)
      FY 2022/23 : cols  8-10 (Mid-Aug), 11-13 (Mid-Sep), 14-16 (Mid-Aug DUP)
    Country rows: Qatar, Malaysia, and a "Total" aggregate row + a "Nepal" all-zero
    placeholder row — both MUST be excluded as dimensions.
    """
    _ensure_fixture_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Country"
    ws.cell(row=1, column=1, value="Migrant workers by Country")
    # Row 3 (index 2): FY banner — sparse, at the first column of each FY block.
    ws.cell(row=3, column=1, value="Country")
    ws.cell(row=3, column=2, value="2021/22")
    ws.cell(row=3, column=8, value="2022/23")
    # Row 4 (index 3): AD month label at each group's Male (first) column.
    #   FY2021/22: Mid-Aug @2, Mid-Sep @5;  FY2022/23: Mid-Aug @8, Mid-Sep @11,
    #   and a DUPLICATE Mid-Aug @14 (the source mislabel quirk).
    for col, label in ((2, "Mid-Aug"), (5, "Mid-Sep"), (8, "Mid-Aug"),
                       (11, "Mid-Sep"), (14, "Mid-Aug")):
        ws.cell(row=4, column=col, value=label)
    # Row 5 (index 4): Male/Female/Total repeating across all 5 groups (cols 2..16).
    for g_start in (2, 5, 8, 11, 14):
        ws.cell(row=5, column=g_start, value="Male")
        ws.cell(row=5, column=g_start + 1, value="Female")
        ws.cell(row=5, column=g_start + 2, value="Total")

    # Data rows (row 6+ == index 5+). For each group we set Male, Female and an
    # explicit Total (Male+Female) so a fact equals the Total column, not a sum.
    # Group order across cols: (2021/22 Aug),(2021/22 Sep),(2022/23 Aug),
    #                          (2022/23 Sep),(2022/23 Aug DUP).
    # Each tuple below is (Male, Female, Total) per group, left→right.
    bodies: list[tuple[str, list[tuple[int, int, int]]]] = [
        ("Qatar", [(100, 5, 105), (110, 6, 116), (120, 7, 127), (130, 8, 138), (0, 0, 0)]),
        ("Malaysia", [(40, 1, 41), (45, 2, 47), (50, 3, 53), (55, 4, 59), (0, 0, 0)]),
        # "Nepal" is an all-zero placeholder in the real file — must be EXCLUDED.
        ("Nepal", [(0, 0, 0), (0, 0, 0), (0, 0, 0), (0, 0, 0), (0, 0, 0)]),
        # "Total" aggregate row — must be EXCLUDED as a country dimension.
        ("Total", [(140, 6, 146), (155, 8, 163), (170, 10, 180), (185, 12, 197), (0, 0, 0)]),
    ]
    for r_off, (country, groups) in enumerate(bodies):
        r = 6 + r_off
        ws.cell(row=r, column=1, value=country)
        for g_idx, (male, female, total) in enumerate(groups):
            g_start = 2 + g_idx * 3
            ws.cell(row=r, column=g_start, value=male)
            ws.cell(row=r, column=g_start + 1, value=female)
            ws.cell(row=r, column=g_start + 2, value=total)
    wb.save(str(path))


@pytest.fixture(scope="session")
def migrant_workers_xlsx() -> Path:
    # Named "Migrant-Workers-Remittance.xlsx" so parse_dne's _DIMENSIONAL_FILE_STEMS
    # dispatch routes it to the country HEADCOUNT dimensional path end-to-end.
    p = FIXTURE_DIR / "Migrant-Workers-Remittance.xlsx"
    if not p.exists():
        _build_migrant_workers(p)
    return p

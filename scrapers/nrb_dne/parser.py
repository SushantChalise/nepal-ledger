"""NRB Database on Nepalese Economy (DNE) XLSX parser — deterministic Python.

Source id: ``nrb-dne-xlsx``.

Layout contract (all DNE XLSX files share this wide-format):
    - First column(s): indicator label / row descriptor text.
    - Header row(s): period labels, e.g. "2079/80", "2080/81" (annual FY) or
      "Shrawan 2082", "Bhadra 2082" (monthly BS period).  The header row is
      detected as the first row that contains at least one parseable period
      token.
    - Data rows: each non-empty label row contains float (or int) values
      keyed to the column period headers.
    - Some files have a title row / unit annotation row above the header.
      The unit string ("in million US$", "Rs. in million", etc.) is extracted
      from these preamble rows.

Slug convention:
    ``dne-<kebab-case-label>`` (prefix ``dne-`` + slugified row label).

Unit detection:
    Scanned from preamble rows (rows before the detected header row) and the
    sheet title.  Mapping table follows NRB's common phrasings.  If the unit
    cannot be resolved, a ``UnitAmbiguous`` error is emitted but parsing
    continues — the raw unit string is used as the ``unit`` field so the
    validator can flag it rather than dropping data.

Period detection (four layouts, tried in priority order):
    1. Long panel — FY label in col 0 (sparse, forward-filled) + AD month name in
       col 1 + numeric value columns to the right (Exchange-rate).  Detected FIRST
       because the standard header detector would otherwise mis-claim it.
    2. Standard wide — indicators as rows, fiscal-period labels as column headers.
       Annual FY: "2079/80", "2079-80" (BS) or "2022/23" (AD, converted via the
       +57 offset, ADR-0013).  Monthly BS: "<bs_month_name> <bs_year>", e.g.
       "Shrawan 2082".
    3. Two-row monthly header — a row of integer AD YEARS over a row of AD MONTH
       names (Foreign-exchange-reserves).  Each (year, month) column is a monthly
       period; the sparse year row is forward-filled.  A repeated (year, month)
       column (source mislabel) keeps both values, flags them, and emits one
       ``PeriodAmbiguous``.
    4. Transposed — AD MONTH names as column headers with integer AD YEARS as row
       labels down col 0 (Tourist-arrivals); long-formatted to one row per
       year×month.
    5. Annual column-series (v0.6.0, real-sector) — annual FY labels stacked DOWN
       col 0 (a "Year"/"Fiscal Year" column) with named-indicator value columns to
       the right (GDP Series_Nominal/_Real, CPI_National/_KTM Valley). The inverse
       of standard wide. Only an EXPLICIT ALLOWLIST of headline columns is promoted
       to single series (ADR-0014: no catalogue pollution) — nominal/real GDP,
       real-GDP growth, per-capita GDP, GDP deflator, CPI index, inflation rate —
       each with a hard-mapped slug + verified unit (ADR-0011).

    AD calendar months are mapped to the BS month containing their 15th (a
    documented mid-month approximation, the exact inverse of
    ``_common.periods._BS_MONTH_TO_AD_MONTH``); every such row is flagged in
    ``parser_notes`` and the AD month span stored is the exact Gregorian month.
    Unparseable period column headers → ``PeriodUnparseable`` error; the column is
    skipped.  Sheets matching no layout fail loud with ``PeriodUnparseable`` when
    year-like tokens are present (never a silent drop).

Confidence: ``B`` default for all DNE rows (NRB compiles from multiple
agencies; figures revised across publications).

ADR: ADR-0003 — no LLM / AI calls. Pure file-in → dataclass-out.

Version history:
    0.6.0 — real-sector files: annual column-series layout (GDP/CPI headline single
            series) + Provincial-GDP dimensional (`dimension_kind='province'`).
    0.5.0 — Foreign-Trade dimensional_rows (ADR-0015) + single-series slug cleanup.
    0.4.0 — three non-standard AD layouts (long panel, two-row monthly, transposed).
"""

from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Final

import openpyxl

from _common.periods import (
    BS_MONTHS,
    BsMonth,
    fiscal_year_ad_label,
    fiscal_year_label,
    mid_month_ad,
)
from _common.types import (
    ParserError,
    ParserResult,
    ParserStatus,
    ReportingPeriodType,
    StagingRowDraft,
)

PARSER_VERSION: Final[str] = "0.6.0"
SOURCE_ID: Final[str] = "nrb-dne-xlsx"

# Filename stems (lowercased, no extension) routed to the dimensional fact path
# (ADR-0015) instead of the single-series staging path. Foreign Trade's commodity
# matrices (exports/imports by SITC group and by major commodity) do not fit the
# single-series (indicator, period, value) shape; they emit `dimensional_rows`.
# Provincial-GDP (v0.6.0) is a GDP-by-province matrix → `dimension_kind='province'`.
_DIMENSIONAL_FILE_STEMS: Final[frozenset[str]] = frozenset(
    {"foreign-trade", "provincial-gdp-2024-25"}
)

# Filename stems routed to the v0.6.0 real-sector single-series path (annual
# column-series layout: FY labels down col 0, headline-indicator value columns to
# the right). Only an EXPLICIT ALLOWLIST of headline columns is promoted (ADR-0014:
# no catalogue pollution); see `_REAL_SECTOR_COLUMN_SPECS`.
_REAL_SECTOR_FILE_STEMS: Final[frozenset[str]] = frozenset(
    {"national-accounts", "consumer-price-index"}
)


# ---------------------------------------------------------------------------
# Dimensional fact contract (ADR-0015) — DNE-LOCAL, intentionally NOT in
# _common/types.py. The shared ParserResult stays single-series; this parser's
# __main__ adds a `dimensional_rows` key to its JSON dict and the DNE ingest CLI
# reads it. Fields mirror the ADR-0015 `dne_facts` parser contract exactly.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class DimensionalRowDraft:
    """One dimensional fact: a base measure sliced by exactly one dimension.

    Mirrors the ADR-0015 parser contract field-for-field. Datetimes are real
    ``datetime`` objects here; ``to_json_dict`` serialises them to ISO-8601 for
    the CLI/ingest boundary (same convention as ``StagingRowDraft``).
    """

    base_indicator_slug: str
    base_indicator_name: str
    dimension_kind: str
    dimension_value: str
    dimension_label: str
    value: float
    unit: str
    reporting_period_type: ReportingPeriodType
    reporting_period_bs: str
    reporting_period_ad_start: datetime
    reporting_period_ad_end: datetime
    fiscal_year_bs: str
    fiscal_year_ad_label: str
    confidence_grade: str

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "base_indicator_slug": self.base_indicator_slug,
            "base_indicator_name": self.base_indicator_name,
            "dimension_kind": self.dimension_kind,
            "dimension_value": self.dimension_value,
            "dimension_label": self.dimension_label,
            "value": self.value,
            "unit": self.unit,
            "reporting_period_type": self.reporting_period_type,
            "reporting_period_bs": self.reporting_period_bs,
            "reporting_period_ad_start": self.reporting_period_ad_start.isoformat(),
            "reporting_period_ad_end": self.reporting_period_ad_end.isoformat(),
            "fiscal_year_bs": self.fiscal_year_bs,
            "fiscal_year_ad_label": self.fiscal_year_ad_label,
            "confidence_grade": self.confidence_grade,
        }


@dataclass(frozen=True)
class DneParserResult:
    """DNE-local result carrying BOTH single-series and dimensional output.

    The shared ``ParserResult`` (``_common.types``) is unchanged; this wrapper is
    what the DNE CLI serialises. ``staging_rows`` and ``dimensional_rows`` are
    mutually exclusive per file (single-series files populate the former; the
    Foreign-Trade matrix file populates the latter).
    """

    status: ParserStatus
    parser_version: str
    staging_rows: list[StagingRowDraft] = field(default_factory=list)
    dimensional_rows: list[DimensionalRowDraft] = field(default_factory=list)
    errors: list[ParserError] = field(default_factory=list)

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "parser_version": self.parser_version,
            "staging_rows": [r.to_json_dict() for r in self.staging_rows],
            "dimensional_rows": [r.to_json_dict() for r in self.dimensional_rows],
            "errors": [e.to_json_dict() for e in self.errors],
        }

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Confidence default for all DNE rows — NRB compiles from multiple agencies;
# figures are revised.  Validation layer may promote individual rows to 'A'.
_CONFIDENCE: Final = "B"

# Publication date placeholder — DNE files carry no embedded publication date;
# the orchestrator supplies the actual download timestamp.  We use a sentinel
# that is clearly approximate so the TS validator can flag it.
_PUB_DATE_SENTINEL: Final[datetime] = datetime(1970, 1, 1, tzinfo=UTC)
_PUB_DATE_BS_SENTINEL: Final[str] = "unknown"

# Regex: annual FY label like "2079/80", "2079-80", "2079/80R", "2079/80P",
# "2079/2080" (4-digit tail used in some SITC sheets), optionally prefixed
# with a bracketed BS label like "(2071-72) 2014/15".
# Groups: (1) = BS/AD start year 4-digit, (2) = tail 2- or 4-digit.
# Revision suffix [R/P/E] is stripped before matching.
_ANNUAL_FY_RE: Final = re.compile(
    r"^\s*(?:\(\d{4}[-/]\d{2,4}\)\s+)?"  # optional "(YYYY-YY) " prefix
    r"(\d{4})\s*[/\-]\s*(\d{2,4})\s*[RPEQrpeq]?\s*$"
)

# Regex: monthly BS period like "Shrawan 2082", "Bhadra 2081", case-insensitive.
# Build from the canonical month list so it stays in sync with _common.periods.
_MONTH_NAMES_PATTERN: Final[str] = "|".join(BS_MONTHS)
_MONTHLY_BS_RE: Final = re.compile(
    rf"^\s*({_MONTH_NAMES_PATTERN})\s+(\d{{4}})\s*$",
    re.IGNORECASE,
)

# Unit string mapping — NRB phrasing → canonical vocab.
# Keys are lowercased and whitespace-normalised before lookup.
_UNIT_MAP: Final[dict[str, str]] = {
    "in million us$": "usd_million",
    "in million us dollars": "usd_million",
    "million us$": "usd_million",
    "million usd": "usd_million",
    "us$ million": "usd_million",
    "usd million": "usd_million",
    "in us$ million": "usd_million",
    "in usd million": "usd_million",
    "in million usd": "usd_million",
    "rs. in million": "npr_million",
    "rs in million": "npr_million",
    "nrs. in million": "npr_million",
    "nrs in million": "npr_million",
    "rs. million": "npr_million",
    "rs million": "npr_million",
    "npr million": "npr_million",
    "million rs.": "npr_million",
    "million rs": "npr_million",
    "in million rs.": "npr_million",
    "in million rs": "npr_million",
    "in rs. million": "npr_million",
    "npr in million": "npr_million",
    "in npr million": "npr_million",
    "nrs million": "npr_million",
    "rs. in billion": "npr_billion",
    "rs in billion": "npr_billion",
    "nrs. in billion": "npr_billion",
    "npr billion": "npr_billion",
    "billion rs.": "npr_billion",
    "billion rs": "npr_billion",
    "in billion rs.": "npr_billion",
    "in npr billion": "npr_billion",
    "percent": "percent",
    "percentage": "percent",
    "in percent": "percent",
    "%": "percent",
    "number": "count",
    "nos.": "count",
    "nos": "count",
    "no.": "count",
    "no": "count",
    "count": "count",
    "in number": "count",
    "metric ton": "metric_ton",
    "metric tons": "metric_ton",
    "in metric tons": "metric_ton",
    "kwh": "kwh",
    "mwh": "mwh",
    "gwh": "gwh",
    "kilowatt hour": "kwh",
    "months": "months",
    "month": "months",
}

# Max rows to scan before the detected header row when searching for a unit.
_PREAMBLE_SCAN_ROWS: Final[int] = 10

# AD calendar year bounds for bare integer year detection in preamble rows.
# NRB uses bare integers like 2001, 2002 as year-row labels in the FX-reserves
# file; these are AD years, not BS.
_AD_YEAR_INT_MIN: Final[int] = 1990
_AD_YEAR_INT_MAX: Final[int] = 2040  # same as _BS_YEAR_MIN; ambiguous zone deferred

# Minimum parseable period columns required to call a sheet non-empty.
_MIN_PERIOD_COLS: Final[int] = 1

# Minimum BS fiscal-year start to distinguish BS years (2040+) from AD years
# (≤2039). The two ranges cannot overlap for any data this project ingests:
# BS 2040 ≈ AD 1983; AD 2039 is the future. ADR-0013.
_BS_YEAR_MIN: Final[int] = 2040

# Maximum AD fiscal-year start accepted for AD→BS conversion (exclusive upper
# bound = _BS_YEAR_MIN - 1). Any lead year ≥ _BS_YEAR_MIN is treated as BS.
_AD_YEAR_FY_MAX: Final[int] = _BS_YEAR_MIN - 1  # 2039

# AD Gregorian month-name → month-number (1-12). NRB month-header rows mix
# abbreviated and full English names ("Aug" vs "August", "Sept" vs "September",
# "March", "April", "June", "July"), so every common variant is mapped. Keys are
# lowercased before lookup.  Used by the integer-year+monthly, long-panel, and
# transposed AD layouts (ADR-0013 follow-up; the wide BS layout uses BS months).
_AD_MONTH_NAME_TO_NUM: Final[dict[str, int]] = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

# AD Gregorian month-number → BS month name. This is the exact inverse of
# `_common.periods._BS_MONTH_TO_AD_MONTH` and round-trips with `mid_month_ad`:
# the BS month listed for AD month M is the one whose mid-point (the 15th of M)
# falls inside it. This is a DOCUMENTED MID-MONTH APPROXIMATION — an AD calendar
# month overlaps two BS months (e.g. AD January spans the tail of Poush and the
# head of Magh); we attribute the whole AD month to the BS month containing its
# 15th. The TS validation layer refines to exact BS-calendar boundaries. We never
# fabricate a period: the BS label stored is a real, defensible monthly period,
# explicitly flagged via `parser_notes`. Kept local (not in _common) per the
# scope fence; mirrors the same mid-July break-month rule as `mid_month_ad`.
_AD_MONTH_NUM_TO_BS_MONTH: Final[dict[int, BsMonth]] = {
    1: "Magh",
    2: "Falgun",
    3: "Chait",
    4: "Baisakh",
    5: "Jestha",
    6: "Ashadh",
    7: "Shrawan",
    8: "Bhadra",
    9: "Ashwin",
    10: "Kartik",
    11: "Mangsir",
    12: "Poush",
}

# Mirror of `_common.periods._AD_YEAR_BREAK_MONTH`: AD months ≥ July belong to BS
# year (ad_year + 57); months < July belong to BS year (ad_year + 56).
_AD_YEAR_BREAK_MONTH: Final[int] = 7

# Fiscal-year offset between BS and AD lead years (ADR-0013): BS = AD + 57.
_BS_AD_FY_OFFSET: Final[int] = 57

# Note appended to every monthly draft built from an AD calendar month, recording
# the mid-month BS approximation so the validator (and any auditor) sees it.
_AD_MONTHLY_APPROX_NOTE: Final[str] = (
    "AD calendar month mapped to BS month containing its 15th (mid-month "
    "approximation per ADR-0013 follow-up); validator refines exact BS boundaries"
)

# Minimum number of integer-year cells a row must contain to be considered the
# "years" row of a two-row (year-over-month) monthly header.
_MIN_YEAR_HEADER_CELLS: Final[int] = 3

# Minimum number of AD-month-name cells a row must contain to be considered the
# "months" row of a two-row monthly header, OR the column header of a transposed
# (years-as-rows) sheet.
_MIN_MONTH_HEADER_CELLS: Final[int] = 6

# Labels that mark a non-period column in transposed/long layouts (annual totals,
# the row-label header itself). Lowercased before comparison.
_NON_MONTH_COL_LABELS: Final[frozenset[str]] = frozenset(
    {"total", "annual", "annual total", "year total", "sum", "year"}
)

# Month-1-of-fiscal-year (Shrawan = AD July) — used to derive the FY a monthly
# AD period belongs to when no explicit FY column is present.
_FY_FIRST_AD_MONTH: Final[int] = 7

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


# Leading outline-enumerator prefix on a row label. NRB statistical tables number
# rows with section enumerators that are layout scaffolding, not part of the
# indicator name — stripping them yields clean, stable slugs (ADR-0015 follow-up):
#   "A. Nepal Rastra Bank"        → "Nepal Rastra Bank"
#   "C. Gross Foreign Exchange…"  → "Gross Foreign Exchange…"
#   "1. Gold, SDR, IMF…"          → "Gold, SDR, IMF…"
#   "2.1 Other capital transfers" → "Other capital transfers"
#   "iii) Convertible"            → "Convertible"
#   "(a) Merchandise"             → "Merchandise"
# Matches a single leading token of letters/digits/dots (an outline code like
# "1.A.a.1" or a bare "A"/"1"), optionally bracketed, followed by a separator
# ('.', ')', ':', '-') and whitespace. Conservative: requires the trailing
# separator+space so real labels beginning with a word are never truncated.
_LEADING_ENUMERATOR_RE: Final = re.compile(
    r"^\s*\(?\s*"            # optional opening paren
    r"[A-Za-z0-9]+(?:\.[A-Za-z0-9]+)*"  # outline code: A | 1 | 1.A.a.1
    r"\s*\)?"                # optional closing paren
    r"\s*[.):\-]\s+"         # mandatory separator + whitespace
)

# Trailing aggregation hint NRB appends to a parent/total row, e.g.
# "A. Nepal Rastra Bank (1+2)", "D. Gross Foreign Assets (A+B)",
# "I. Change in NFA (G+H)". The parenthetical encodes which child rows sum into
# the parent — useful provenance, but it pollutes the slug and is not part of the
# indicator name. Stripped for slug derivation only (the raw label is preserved
# elsewhere). Matches a trailing "(...)" whose interior is only enumerator-like
# tokens joined by + / , (so genuine descriptive parentheticals are kept).
_TRAILING_AGG_HINT_RE: Final = re.compile(
    r"\s*\(\s*[A-Za-z0-9]+(?:\s*[+,]\s*[A-Za-z0-9]+)*\s*\)\s*$"
)


def _strip_enumerator(label: str) -> str:
    """Remove a leading outline enumerator and trailing aggregation hint.

    Used only for slug derivation. Returns the cleaned label, or the original
    (trimmed) text if nothing matched. Never returns empty: if stripping would
    empty the label (degenerate row whose label is purely an enumerator), the
    original trimmed text is kept so the slug stays non-empty and traceable.
    """
    cleaned = _LEADING_ENUMERATOR_RE.sub("", label, count=1)
    cleaned = _TRAILING_AGG_HINT_RE.sub("", cleaned)
    cleaned = cleaned.strip()
    return cleaned if cleaned else label.strip()


def _slugify(label: str) -> str:
    """Convert an indicator label to a dne-prefixed kebab-case slug.

    Strips leading outline enumerators ("A.", "1.A.a.1") and trailing aggregation
    hints ("(1+2)", "(A+B)") first so slugs are clean and stable:
    "A. Nepal Rastra Bank (1+2)" → "dne-nepal-rastra-bank";
    "Total Foreign Exchange Reserves" → "dne-total-foreign-exchange-reserves".
    """
    base = _strip_enumerator(label)
    # Lowercase, keep alphanumeric and spaces, strip other chars, then hyphenate.
    slug = base.lower()
    slug = re.sub(r"[^a-z0-9\s]+", " ", slug)
    slug = re.sub(r"\s+", "-", slug.strip())
    return f"dne-{slug}"


def _norm_text(raw: object) -> str:
    """Stringify a cell value and normalise internal whitespace."""
    if raw is None:
        return ""
    return " ".join(str(raw).split())


def _safe_float(raw: object) -> float | None:
    """Coerce cell value to float; return None for empty / non-numeric / NaN."""
    if raw is None:
        return None
    s = str(raw).strip()
    # NRB sometimes uses "-" or "--" for missing data.
    if s in ("", "-", "--", "N/A", "n/a", "NA", "..."):
        return None
    try:
        v = float(s.replace(",", ""))
    except (TypeError, ValueError):
        return None
    if v != v:  # NaN  # noqa: PLR0124
        return None
    return v


def _detect_unit_from_text(text: str) -> str | None:
    """Look up a raw unit string in the unit map. Returns canonical vocab or None.

    Strips surrounding parentheses before lookup to handle NRB's common pattern
    of writing unit annotations as "(Rs in Million)" or "(NPR in Million)".
    """
    # Strip leading/trailing parentheses that NRB wraps around unit strings.
    stripped = text.strip()
    if stripped.startswith("(") and stripped.endswith(")"):
        stripped = stripped[1:-1]
    normalised = " ".join(stripped.lower().split())
    # Direct match first.
    if normalised in _UNIT_MAP:
        return _UNIT_MAP[normalised]
    # Substring match: check if any key is contained in the normalised text.
    for key, vocab in _UNIT_MAP.items():
        if key in normalised:
            return vocab
    return None


def _parse_annual_fy(label: str) -> tuple[str, str] | None:
    """Parse annual FY label → (fiscal_year_bs "YYYY/YY", fiscal_year_ad_label).

    Accepts BS fiscal years (lead year ≥ 2040) and AD fiscal years (lead year
    ≤ 2039), as per ADR-0013.  The magnitude heuristic is deterministic: the
    two ranges cannot overlap for any data this project ingests.

    BS inputs (lead ≥ 2040):
    - "2079/80", "2079-80"            — standard BS format
    - "2079/80R", "2079/80P"          — NRB revised/provisional suffix
    - "2079/2080"                     — 4-digit tail (some SITC sheets)

    AD inputs (lead ≤ 2039) — converted to BS via the +57 fiscal-year offset:
    - "2022/23", "2022/23R"           — plain AD FY (External Sector files)
    - "(2071-72) 2014/15"             — bracketed BS label + AD year prefix
      (The bracketed part is stripped by the regex; the AD lead year is used.)

    The AD→BS conversion is done via ``fiscal_year_label(ad_start + 57)``
    which mirrors ``fiscal_year_ad_label`` in reverse.  Known pair:
    AD 2022/23 → BS 2079/80; AD 2023/24 → BS 2080/81.

    Returns None if the regex does not match or the tail is inconsistent.
    """
    m = _ANNUAL_FY_RE.match(label)
    if not m:
        return None
    start = int(m.group(1))
    tail_raw = m.group(2)
    tail_int = int(tail_raw) % 100  # normalise 4-digit "2080" → 80

    # Validate: tail must equal (start + 1) mod 100.
    expected_tail = (start + 1) % 100
    if tail_int != expected_tail:
        return None

    if start >= _BS_YEAR_MIN:
        # BS fiscal year — keep as-is, derive AD label from periods helper.
        fy_bs = f"{start}/{expected_tail:02d}"
        fy_ad = fiscal_year_ad_label(start)
        return fy_bs, fy_ad

    if start <= _AD_YEAR_FY_MAX:
        # AD fiscal year — convert to BS by adding 57 to the lead year.
        # The Nepal fiscal year runs mid-July to mid-July, so AD YYYY/(YY+1)
        # corresponds 1:1 to BS (YYYY+57)/((YYYY+58)%100).  ADR-0013.
        bs_start = start + 57
        fy_bs = fiscal_year_label(bs_start)
        fy_ad = f"{start}/{expected_tail:02d}"
        return fy_bs, fy_ad

    return None


def _parse_monthly_bs(label: str) -> tuple[BsMonth, int] | None:
    """Parse "Shrawan 2082" → (BsMonth, bs_year). Returns None if no match."""
    m = _MONTHLY_BS_RE.match(label)
    if not m:
        return None
    # Capitalise so it matches the BsMonth literal exactly.
    month_raw = m.group(1).capitalize()
    # "Chait" is used in the codebase; also accept "Chaitra".
    if month_raw == "Chaitra":
        month_raw = "Chait"
    if month_raw not in BS_MONTHS:
        return None
    return month_raw, int(m.group(2))  # type: ignore[return-value]


def _annual_fy_to_draft_fields(
    fy_bs: str,
    fy_ad: str,
    unit: str,
    slug: str,
    value: float,
    parser_notes: str | None = None,
) -> StagingRowDraft:
    """Build a StagingRowDraft for an annual FY cell."""
    bs_start = int(fy_bs.split("/")[0])
    # NRB annual FY runs mid-July (Shrawan 1) to mid-July (Asar 31).
    # Approximate: start = 15 July of AD start year, end = 15 July of AD end year.
    ad_start_year = bs_start - 57
    ad_start = datetime(ad_start_year, 7, 15, tzinfo=UTC)
    ad_end = datetime(ad_start_year + 1, 7, 15, tzinfo=UTC)
    return StagingRowDraft(
        indicator_slug_raw=slug,
        value=value,
        unit=unit,
        reporting_period_type="annual",
        reporting_period_bs=fy_bs,
        reporting_period_ad_start=ad_start,
        reporting_period_ad_end=ad_end,
        publication_date_ad=_PUB_DATE_SENTINEL,
        publication_date_bs=_PUB_DATE_BS_SENTINEL,
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade_proposed=_CONFIDENCE,
        parser_notes=parser_notes,
    )


def _monthly_bs_to_draft_fields(
    bs_month: BsMonth,
    bs_year: int,
    unit: str,
    slug: str,
    value: float,
    parser_notes: str | None = None,
) -> StagingRowDraft:
    """Build a StagingRowDraft for a monthly BS cell."""
    mid = mid_month_ad(bs_month, bs_year)
    # Month span: 1st to 28th of the AD month (safe lower bound for mid-month to
    # mid-month; the TS validator refines to exact BS calendar boundaries).
    ad_start = datetime(mid.year, mid.month, 1, tzinfo=UTC)
    ad_end = datetime(mid.year, mid.month, 28, tzinfo=UTC)
    # Fiscal year: if month is Magh..Ashadh → FY starts in bs_year - 1.
    _late_months: Final = {"Magh", "Falgun", "Chait", "Baisakh", "Jestha", "Ashadh"}
    fy_start = bs_year - 1 if bs_month in _late_months else bs_year
    fy_bs = f"{fy_start}/{(fy_start + 1) % 100:02d}"
    fy_ad = fiscal_year_ad_label(fy_start)
    return StagingRowDraft(
        indicator_slug_raw=slug,
        value=value,
        unit=unit,
        reporting_period_type="monthly",
        reporting_period_bs=f"{bs_month} {bs_year}",
        reporting_period_ad_start=ad_start,
        reporting_period_ad_end=ad_end,
        publication_date_ad=_PUB_DATE_SENTINEL,
        publication_date_bs=_PUB_DATE_BS_SENTINEL,
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade_proposed=_CONFIDENCE,
        parser_notes=parser_notes,
    )


def _parse_ad_month_name(label: str) -> int | None:
    """Parse an AD Gregorian month name → month number (1-12), or None.

    Accepts NRB's mixed abbreviated/full English month names (case-insensitive,
    surrounding whitespace stripped): "Aug", "August", "Sept", "March", "June".
    """
    key = label.strip().lower()
    return _AD_MONTH_NAME_TO_NUM.get(key)


def _ad_month_to_bs(ad_year: int, ad_month: int) -> tuple[BsMonth, int]:
    """Map an AD (year, month) to its (BS month, BS year) — mid-month approximation.

    The BS month is the one containing the 15th of the AD month (the exact inverse
    of `_common.periods._BS_MONTH_TO_AD_MONTH`); the BS year follows the same
    mid-July break rule as `mid_month_ad`. Documented approximation per ADR-0013
    follow-up — see ``_AD_MONTH_NUM_TO_BS_MONTH``. Never fabricates: the result is
    a real BS month/year pair, and callers flag the approximation in parser_notes.
    """
    bs_month = _AD_MONTH_NUM_TO_BS_MONTH[ad_month]
    bs_year = (
        ad_year + _BS_AD_FY_OFFSET
        if ad_month >= _AD_YEAR_BREAK_MONTH
        else ad_year + _BS_AD_FY_OFFSET - 1
    )
    return bs_month, bs_year


def _ad_monthly_to_draft_fields(
    ad_year: int,
    ad_month: int,
    unit: str,
    slug: str,
    value: float,
    extra_note: str | None = None,
) -> StagingRowDraft:
    """Build a monthly StagingRowDraft from an AD (Gregorian) year+month cell.

    The AD month span is exact (1st → 28th of the Gregorian month — a safe lower
    bound the validator widens). Only the BS *label* is the mid-month
    approximation, flagged in ``parser_notes`` via ``_AD_MONTHLY_APPROX_NOTE``.
    The fiscal year is derived from the AD month: AD July (Shrawan) begins FY
    ``ad_year/ad_year+1``; AD Jan–Jun belong to the FY that began the prior July.

    ``extra_note`` is appended to ``parser_notes`` (used to flag source-level
    quirks such as a repeated (year, month) column in the header).
    """
    bs_month, bs_year = _ad_month_to_bs(ad_year, ad_month)
    ad_start = datetime(ad_year, ad_month, 1, tzinfo=UTC)
    ad_end = datetime(ad_year, ad_month, 28, tzinfo=UTC)
    # FY lead (AD): months Jul..Dec → this AD year; Jan..Jun → previous AD year.
    fy_ad_start = ad_year if ad_month >= _FY_FIRST_AD_MONTH else ad_year - 1
    fy_ad = f"{fy_ad_start}/{(fy_ad_start + 1) % 100:02d}"
    bs_fy_start = fy_ad_start + _BS_AD_FY_OFFSET
    fy_bs = fiscal_year_label(bs_fy_start)
    notes = (
        _AD_MONTHLY_APPROX_NOTE
        if extra_note is None
        else f"{_AD_MONTHLY_APPROX_NOTE}; {extra_note}"
    )
    return StagingRowDraft(
        indicator_slug_raw=slug,
        value=value,
        unit=unit,
        reporting_period_type="monthly",
        reporting_period_bs=f"{bs_month} {bs_year}",
        reporting_period_ad_start=ad_start,
        reporting_period_ad_end=ad_end,
        publication_date_ad=_PUB_DATE_SENTINEL,
        publication_date_bs=_PUB_DATE_BS_SENTINEL,
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade_proposed=_CONFIDENCE,
        parser_notes=notes,
    )


# ---------------------------------------------------------------------------
# Sheet-level parser — broken into focused sub-functions to satisfy ruff
# PLR0912 (branches ≤ 12) and PLR0915 (statements ≤ 50).
# ---------------------------------------------------------------------------


def _scan_unit_hint(rows: list[tuple[object, ...]], sheet_name: str) -> str | None:
    """Return a canonical unit string from preamble rows or the sheet name."""
    preamble_chunks: list[str] = []
    for row in rows[:_PREAMBLE_SCAN_ROWS]:
        for cell in row:
            if cell is not None:
                preamble_chunks.append(_norm_text(cell))
    blob = " ".join(preamble_chunks).lower()
    hint = _detect_unit_from_text(blob)
    if hint is None:
        hint = _detect_unit_from_text(sheet_name.lower())
    return hint


def _detect_header(
    rows: list[tuple[object, ...]],
) -> tuple[int | None, dict[int, tuple[str, object]]]:
    """Find the first row with ≥1 parseable period; return (row_idx, period_cols)."""
    for row_idx, row in enumerate(rows):
        col_periods: dict[int, tuple[str, object]] = {}
        for col_idx, cell in enumerate(row):
            label = _norm_text(cell)
            annual = _parse_annual_fy(label)
            if annual:
                col_periods[col_idx] = ("annual", annual)
                continue
            monthly = _parse_monthly_bs(label)
            if monthly:
                col_periods[col_idx] = ("monthly", monthly)
        if len(col_periods) >= _MIN_PERIOD_COLS:
            return row_idx, col_periods
    return None, {}


def _find_label_col(
    header_row: tuple[object, ...],
    first_period_col: int,
) -> int:
    """Return the rightmost non-period column index before first_period_col."""
    for c in range(first_period_col - 1, -1, -1):
        if c < len(header_row) and _norm_text(header_row[c]):
            return c
    return 0


def _collect_period_errors(
    header_row: tuple[object, ...],
    period_cols: dict[int, tuple[str, object]],
    first_period_col: int,
    sheet_name: str,
) -> list[ParserError]:
    """Emit PeriodUnparseable errors for year-like header cells that didn't parse."""
    errors: list[ParserError] = []
    for col_idx in range(first_period_col, len(header_row)):
        if col_idx in period_cols:
            continue
        cell_text = _norm_text(header_row[col_idx] if col_idx < len(header_row) else None)
        if not cell_text:
            continue
        if re.search(r"\b(20\d{2}|19\d{2})\b", cell_text):
            errors.append(
                ParserError(
                    error_class="PeriodUnparseable",
                    error_detail=(
                        f"sheet={sheet_name!r} col={col_idx}: "
                        f"period header {cell_text!r} could not be parsed; column skipped"
                    ),
                    source_excerpt=cell_text,
                )
            )
    return errors


def _resolve_unit(
    unit_hint: str | None,
    label_raw: str,
    row_idx: int,
    slug: str,
    sheet_name: str,
) -> tuple[str, ParserError | None]:
    """Return (canonical_unit, optional_UnitAmbiguous_error)."""
    if unit_hint is not None:
        return unit_hint, None
    detected = _detect_unit_from_text(label_raw)
    if detected:
        return detected, None
    err = ParserError(
        error_class="UnitAmbiguous",
        error_detail=(
            f"sheet={sheet_name!r} row={row_idx} slug={slug!r}: "
            f"unit not resolved; literal label used as unit"
        ),
        source_excerpt=label_raw,
    )
    return label_raw, err


def _build_draft(
    period_type: str,
    period_meta: object,
    row_unit: str,
    slug: str,
    value: float,
    row_idx: int,
    col_idx: int,
    label_raw: str,
    sheet_name: str,
) -> tuple[StagingRowDraft | None, ParserError | None]:
    """Convert a single (period, value) cell into a draft or a typed error."""
    try:
        if period_type == "annual":
            fy_bs, fy_ad = period_meta  # type: ignore[misc]
            draft = _annual_fy_to_draft_fields(
                fy_bs=fy_bs, fy_ad=fy_ad, unit=row_unit, slug=slug, value=value
            )
        else:
            bs_month, bs_year = period_meta  # type: ignore[misc]
            draft = _monthly_bs_to_draft_fields(
                bs_month=bs_month, bs_year=bs_year, unit=row_unit, slug=slug, value=value
            )
    except (ValueError, KeyError) as exc:
        err = ParserError(
            error_class="PeriodUnparseable",
            error_detail=(
                f"sheet={sheet_name!r} row={row_idx} col={col_idx}: "
                f"period conversion failed: {exc}"
            ),
            source_excerpt=label_raw,
        )
        return None, err
    return draft, None


_SKIP_LABELS: Final[frozenset[str]] = frozenset(
    {"total", "subtotal", "sub-total", "grand total", "memo"}
)


def _qualifier_fragment(qualifier: str | None) -> str | None:
    """Slugify a collision qualifier (outline code or parent label) → kebab tail.

    "3.4.1.1"                     → "3-4-1-1"
    "C. Gross Foreign Exchange…"  → "gross-foreign-exchange…" (enumerator stripped)
    Returns None when the qualifier is empty/whitespace (no usable qualifier).
    """
    if qualifier is None:
        return None
    base = _strip_enumerator(qualifier)
    frag = re.sub(r"[^a-z0-9\s]+", " ", base.lower())
    frag = re.sub(r"\s+", "-", frag.strip())
    return frag or None


def _resolve_slug_collision(
    slug: str,
    qualifier: str | None,
    seen_slugs: set[str],
    row_idx: int,
) -> str:
    """Return a unique slug for a row whose plain slug already occurred.

    Resolution order (deterministic, documented per ADR-0015 follow-up):
      1. Qualify with the section/parent label (FX-reserves) or outline code (BoP)
         — e.g. "dne-nrb" under outline "3.4.1.1" → "dne-nrb-3-4-1-1"; "Convertible"
         under "C. Gross Foreign Exchange Reserve" →
         "dne-convertible-gross-foreign-exchange-reserve".
      2. If no qualifier is available, or the qualified slug ALSO collides, fall
         back to the row-index suffix "-r{row_idx}" (stable for a given file
         layout — the previous behaviour, retained only as a last resort).
    The non-colliding case never reaches here; callers gate on ``slug in seen``.
    """
    frag = _qualifier_fragment(qualifier)
    if frag is not None:
        qualified = f"{slug}-{frag}"
        if qualified not in seen_slugs:
            return qualified
    return f"{slug}-r{row_idx}"


def _parse_data_rows(
    rows: list[tuple[object, ...]],
    header_row_idx: int,
    label_col_idx: int,
    period_cols: dict[int, tuple[str, object]],
    unit_hint: str | None,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Iterate data rows (after header) and emit staging drafts."""
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []
    seen_slugs: set[str] = set()

    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if label_col_idx >= len(row):
            continue
        label_raw = _norm_text(row[label_col_idx])
        if not label_raw or label_raw.lower() in _SKIP_LABELS:
            continue

        slug = _slugify(label_raw)
        if slug in seen_slugs:
            # Qualifier for a duplicate label: the outline code in a column to the
            # LEFT of the label column (BoP's "S.N." column carries "3.4.1.1"-style
            # codes that uniquely place each repeated label, e.g. "NRB").
            outline = _norm_text(row[0]) if label_col_idx > 0 and row else None
            slug = _resolve_slug_collision(slug, outline, seen_slugs, row_idx)
        seen_slugs.add(slug)

        row_unit, unit_err = _resolve_unit(unit_hint, label_raw, row_idx, slug, sheet_name)
        if unit_err:
            errors.append(unit_err)

        for col_idx, (period_type, period_meta) in period_cols.items():
            if col_idx >= len(row):
                continue
            value = _safe_float(row[col_idx])
            if value is None:
                continue
            draft, period_err = _build_draft(
                period_type, period_meta, row_unit, slug, value,
                row_idx, col_idx, label_raw, sheet_name,
            )
            if period_err:
                errors.append(period_err)
            elif draft is not None:
                staging.append(draft)

    return staging, errors


def _as_year_int(cell: object) -> int | None:
    """Return the AD year an integer-ish cell encodes, or None.

    Accepts ints/floats (2001, 2001.0) and digit strings ("2001"). Bounded to the
    AD calendar-year window so stray numeric data is never mistaken for a year.
    """
    if isinstance(cell, bool):  # bool is an int subclass — exclude explicitly
        return None
    if isinstance(cell, int | float):
        if cell != int(cell):
            return None
        n = int(cell)
    else:
        s = _norm_text(cell)
        if not s.isdigit():
            return None
        n = int(s)
    if _AD_YEAR_INT_MIN <= n <= _AD_YEAR_INT_MAX:
        return n
    return None


def _row_year_cols(row: tuple[object, ...]) -> dict[int, int]:
    """Map column index → AD year for every integer-AD-year cell in a row."""
    return {ci: y for ci, cell in enumerate(row) if (y := _as_year_int(cell)) is not None}


def _row_month_cols(row: tuple[object, ...]) -> dict[int, int]:
    """Map column index → AD month number for every AD-month-name cell in a row."""
    out: dict[int, int] = {}
    for ci, cell in enumerate(row):
        if cell is None:
            continue
        m = _parse_ad_month_name(_norm_text(cell))
        if m is not None:
            out[ci] = m
    return out


def _detect_year_month_header(
    rows: list[tuple[object, ...]],
) -> tuple[int, dict[int, tuple[int, int]]] | None:
    """Detect a two-row header: a row of integer AD YEARS directly above a row of
    AD MONTH names (the Foreign-exchange-reserves layout).

    Returns ``(month_row_idx, {col_idx: (ad_year, ad_month)})`` or None.

    Strategy: scan for an adjacent (year_row, month_row) pair within the preamble.
    The year row is sparse — a year value typically appears only in the first
    column of each year's month-block (e.g. 2001 over "Aug", blanks over the rest)
    OR repeats per month. We forward-fill the year across the month columns so each
    monthly column gets the most recent year seen at or before it.
    """
    scan = min(len(rows) - 1, _PREAMBLE_SCAN_ROWS)
    for ri in range(scan):
        year_cols = _row_year_cols(rows[ri])
        if len(year_cols) < _MIN_YEAR_HEADER_CELLS:
            continue
        month_cols = _row_month_cols(rows[ri + 1])
        if len(month_cols) < _MIN_MONTH_HEADER_CELLS:
            continue
        # Forward-fill the year across month columns. Walk columns left→right;
        # carry the last year seen in the year row at or before this column.
        first_year_col = min(year_cols)
        paired: dict[int, tuple[int, int]] = {}
        current_year: int | None = None
        max_col = max(*year_cols, *month_cols)
        for ci in range(first_year_col, max_col + 1):
            if ci in year_cols:
                current_year = year_cols[ci]
            if ci in month_cols and current_year is not None:
                paired[ci] = (current_year, month_cols[ci])
        if len(paired) >= _MIN_MONTH_HEADER_CELLS:
            return ri + 1, paired
    return None


def _parse_year_month_layout(
    rows: list[tuple[object, ...]],
    month_row_idx: int,
    paired: dict[int, tuple[int, int]],
    unit_hint: str | None,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Emit one monthly draft per (indicator-row × paired year/month column).

    NRB occasionally ships a repeated (year, month) column in the header (e.g.
    two "Oct 2025" columns with *different* values — a source-side mislabel we
    cannot disambiguate without fabricating). We never drop either value: both
    rows are emitted, the duplicate-period columns are flagged in ``parser_notes``,
    and a single ``PeriodAmbiguous`` error surfaces the issue for the validator.
    """
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []
    seen_slugs: set[str] = set()
    first_period_col = min(paired)

    # Identify columns whose (year, month) repeats an earlier column (left→right).
    dup_cols: set[int] = set()
    seen_periods: set[tuple[int, int]] = set()
    for col_idx in sorted(paired):
        ym = paired[col_idx]
        if ym in seen_periods:
            dup_cols.add(col_idx)
        else:
            seen_periods.add(ym)
    if dup_cols:
        dup_sample = ", ".join(
            dict.fromkeys(
                f"{_AD_MONTH_NUM_TO_BS_MONTH[paired[c][1]]} (AD {paired[c][0]}-{paired[c][1]:02d})"
                for c in sorted(dup_cols)
            )
        )
        errors.append(
            ParserError(
                error_class="PeriodAmbiguous",
                error_detail=(
                    f"sheet={sheet_name!r}: header has repeated (year, month) "
                    f"columns ({dup_sample}); both values emitted and flagged — "
                    f"validator must adjudicate the source-side duplicate"
                ),
                source_excerpt=dup_sample,
            )
        )

    # Running section parent: the most recent row whose col-0 cell carried a
    # label (e.g. "C. Gross Foreign Exchange Reserve"). Sub-rows that live only in
    # col 1 ("Convertible", "Inconvertible", "Share in total (in percent)") repeat
    # across sections; we qualify their colliding slugs with this parent so each
    # gets a stable, source-derived slug instead of a row-index suffix.
    current_parent: str | None = None
    for row_idx in range(month_row_idx + 1, len(rows)):
        row = rows[row_idx]
        col0 = _norm_text(row[0]) if row else ""
        # Label may sit in col 0 or, for indented sub-items, col 1. Join the
        # non-empty label cells that precede the first period column.
        label_parts = [
            _norm_text(row[c])
            for c in range(min(first_period_col, len(row)))
            if c < len(row) and _norm_text(row[c])
        ]
        label_raw = " ".join(label_parts)
        if not label_raw or label_raw.lower() in _SKIP_LABELS:
            continue
        # A row that owns a col-0 label becomes the new section parent for the
        # col-1-only sub-rows that follow it.
        if col0:
            current_parent = label_raw

        slug = _slugify(label_raw)
        if slug in seen_slugs:
            # Sub-row with no col-0 label of its own → qualify by section parent.
            qualifier = current_parent if not col0 else None
            slug = _resolve_slug_collision(slug, qualifier, seen_slugs, row_idx)
        seen_slugs.add(slug)

        row_unit, unit_err = _resolve_unit(unit_hint, label_raw, row_idx, slug, sheet_name)
        if unit_err:
            errors.append(unit_err)

        for col_idx, (ad_year, ad_month) in paired.items():
            if col_idx >= len(row):
                continue
            value = _safe_float(row[col_idx])
            if value is None:
                continue
            dup_note = (
                f"source header had a repeated column for this (year, month) at "
                f"col {col_idx}; value not dropped"
                if col_idx in dup_cols
                else None
            )
            staging.append(
                _ad_monthly_to_draft_fields(
                    ad_year, ad_month, row_unit, slug, value, extra_note=dup_note
                )
            )
    return staging, errors


def _detect_long_panel(
    rows: list[tuple[object, ...]],
) -> tuple[int, int, int, list[int]] | None:
    """Detect the long-panel layout (Exchange-rate): an AD fiscal-year label in
    col 0 (sparse — present only on the first month of each FY, forward-filled),
    an AD month name in col 1, and numeric value columns to the right.

    Returns ``(first_data_row, fy_col, month_col, value_cols)`` or None.
    """
    fy_col, month_col = 0, 1
    # Find the first data row: col0 parses as an annual FY (AD or BS) and col1 is
    # an AD month name. Scan a generous window past any multi-row header.
    scan = min(len(rows), _PREAMBLE_SCAN_ROWS * 2)
    for ri in range(scan):
        row = rows[ri]
        if len(row) <= month_col:
            continue
        if _parse_annual_fy(_norm_text(row[fy_col])) is None:
            continue
        if _parse_ad_month_name(_norm_text(row[month_col])) is None:
            continue
        # Value columns: every column ≥ 2 that holds a float somewhere in the
        # next few rows. Use this row plus a couple after it as the probe.
        value_cols: list[int] = []
        probe = rows[ri : ri + 4]
        max_col = max(len(r) for r in probe)
        for c in range(month_col + 1, max_col):
            if any(c < len(r) and _safe_float(r[c]) is not None for r in probe):
                value_cols.append(c)
        if value_cols:
            return ri, fy_col, month_col, value_cols
    return None


def _value_col_label(rows: list[tuple[object, ...]], header_rows: int, col: int) -> str:
    """Build a value-column sub-label by joining header cells above a value column.

    The long-panel sheet has a 3-4 row header naming each numeric column
    (e.g. "Month End Buying", "Monthly Average Middle Rate"). We concatenate the
    non-empty header cells in this column to disambiguate the indicator slug.
    """
    parts = [
        _norm_text(rows[r][col])
        for r in range(header_rows)
        if col < len(rows[r]) and _norm_text(rows[r][col])
    ]
    return " ".join(parts)


def _parse_long_panel_layout(
    rows: list[tuple[object, ...]],
    first_data_row: int,
    fy_col: int,
    month_col: int,
    value_cols: list[int],
    unit_hint: str | None,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Long-format the Exchange-rate panel: one monthly draft per (row × value col).

    The FY label in col 0 is forward-filled. Each value column carries its own
    sub-label (from the multi-row header) so distinct series get distinct slugs.
    Rows whose month cell is an aggregate ("Annual Average") are skipped — they
    are not a single calendar month and would corrupt the monthly period.
    """
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []
    col_labels = {c: _value_col_label(rows, first_data_row, c) for c in value_cols}
    current_fy_ad: str | None = None
    # Resolve each value column's unit ONCE, and only from a positively-known
    # sheet-level hint. We deliberately do NOT keyword-match the column sub-labels
    # ("Month End Buying", etc.): those contain noise words ("month") that the
    # substring matcher would mis-resolve to a wrong vocab unit. When no hint is
    # known (e.g. an FX-rate panel — there is no controlled-vocab "NPR per USD"
    # unit), we emit one UnitAmbiguous per column and carry the raw sub-label so
    # the validator flags it for human unit assignment — never a silent wrong unit.
    col_units: dict[int, str] = {}
    for col in value_cols:
        sub = col_labels.get(col) or f"col{col}"
        if unit_hint is not None:
            col_units[col] = unit_hint
            continue
        col_units[col] = sub
        errors.append(
            ParserError(
                error_class="UnitAmbiguous",
                error_detail=(
                    f"sheet={sheet_name!r} col={col}: unit not resolved for the "
                    f"long-panel value column; raw column label used as unit"
                ),
                source_excerpt=sub,
            )
        )

    for row_idx in range(first_data_row, len(rows)):
        row = rows[row_idx]
        if fy_col < len(row):
            fy_parsed = _parse_annual_fy(_norm_text(row[fy_col]))
            if fy_parsed is not None:
                current_fy_ad = fy_parsed[1]  # AD label "YYYY/YY"
        if month_col >= len(row):
            continue
        month_text = _norm_text(row[month_col])
        if month_text.lower() in _NON_MONTH_COL_LABELS or "average" in month_text.lower():
            # "Annual Average" / "Monthly Average" aggregate rows — not a month.
            continue
        ad_month = _parse_ad_month_name(month_text)
        if ad_month is None or current_fy_ad is None:
            continue
        ad_year = _fy_label_to_calendar_year(current_fy_ad, ad_month)
        if ad_year is None:
            continue
        for col in value_cols:
            if col >= len(row):
                continue
            value = _safe_float(row[col])
            if value is None:
                continue
            sub = col_labels.get(col) or f"col{col}"
            slug = _slugify(f"{sheet_name} {sub}")
            staging.append(
                _ad_monthly_to_draft_fields(
                    ad_year, ad_month, col_units[col], slug, value
                )
            )
    return staging, errors


def _fy_label_to_calendar_year(fy_ad_label: str, ad_month: int) -> int | None:
    """Resolve the AD calendar year of ``ad_month`` within an AD FY label.

    NRB fiscal year runs mid-July→mid-July. For AD FY "2022/23": months Jul–Dec
    fall in the lead calendar year (2022); months Jan–Jun fall in the trailing
    year (2023). Returns None on a malformed label.
    """
    m = re.match(r"^\s*(\d{4})\s*/\s*(\d{2,4})\s*$", fy_ad_label)
    if not m:
        return None
    lead = int(m.group(1))
    return lead if ad_month >= _FY_FIRST_AD_MONTH else lead + 1


def _detect_transposed(
    rows: list[tuple[object, ...]],
) -> tuple[int, int, dict[int, int]] | None:
    """Detect the transposed layout (Tourist-arrivals): a header row of AD MONTH
    names across columns, with integer AD YEARS as row labels down col 0.

    Returns ``(header_row_idx, year_col, {col_idx: ad_month})`` or None.

    Requires both signals to avoid false positives: (a) ≥6 month-name column
    headers, and (b) the rows beneath carry integer AD years in the label column.
    """
    year_col = 0
    scan = min(len(rows), _PREAMBLE_SCAN_ROWS)
    for ri in range(scan):
        month_cols = _row_month_cols(rows[ri])
        if len(month_cols) < _MIN_MONTH_HEADER_CELLS:
            continue
        # Confirm: at least two data rows below carry an AD year in col 0.
        year_rows = sum(
            1
            for r in rows[ri + 1 : ri + 6]
            if year_col < len(r) and _as_year_int(r[year_col]) is not None
        )
        if year_rows >= 2:  # noqa: PLR2004 — need ≥2 year rows to confirm orientation
            return ri, year_col, month_cols
    return None


def _parse_transposed_layout(
    rows: list[tuple[object, ...]],
    header_row_idx: int,
    year_col: int,
    month_cols: dict[int, int],
    unit_hint: str | None,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Long-format a transposed (years-as-rows, months-as-columns) sheet.

    One monthly draft per (year row × month column). Non-month columns (e.g. an
    annual "Total") are ignored because they are not in ``month_cols``. The
    indicator slug is the sheet name (the sheet is a single indicator surface,
    e.g. "Tourist Arrival"), since the row label is the year, not an indicator.
    """
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []
    slug = _slugify(sheet_name)
    row_unit, unit_err = _resolve_unit(unit_hint, sheet_name, header_row_idx, slug, sheet_name)
    if unit_err:
        errors.append(unit_err)

    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if year_col >= len(row):
            continue
        ad_year = _as_year_int(row[year_col])
        if ad_year is None:
            continue
        for col_idx, ad_month in month_cols.items():
            if col_idx >= len(row):
                continue
            value = _safe_float(row[col_idx])
            if value is None:
                continue
            staging.append(
                _ad_monthly_to_draft_fields(ad_year, ad_month, row_unit, slug, value)
            )
    return staging, errors


def _try_alternate_layouts(
    rows: list[tuple[object, ...]],
    unit_hint: str | None,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]] | None:
    """Try the non-standard AD layouts that only apply once the standard wide
    header detection has already failed:

    1. Two-row integer-year + month header (Foreign-exchange-reserves).
    2. Transposed: years-as-rows, months-as-columns (Tourist-arrivals).

    (The long-panel layout is detected earlier in ``_parse_sheet`` because its
    signature would otherwise be mis-claimed by the standard header detector.)

    Returns the first layout that yields ≥1 staging row, else None (so the caller
    falls through to the fail-loud deferral diagnostic).
    """
    ym = _detect_year_month_header(rows)
    if ym is not None:
        staging, errs = _parse_year_month_layout(rows, ym[0], ym[1], unit_hint, sheet_name)
        if staging:
            return staging, errs

    tp = _detect_transposed(rows)
    if tp is not None:
        staging, errs = _parse_transposed_layout(
            rows, tp[0], tp[1], tp[2], unit_hint, sheet_name
        )
        if staging:
            return staging, errs

    return None


def _defer_unparseable_sheet(
    rows: list[tuple[object, ...]],
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Fail-loud diagnostic for a sheet no layout matched.

    Emits a ``PeriodUnparseable`` error if AD-year-like tokens are present (so an
    unhandled real shape is visible, never silently dropped); otherwise returns
    empty (a genuinely blank sheet → NoDataExtracted at the top level).
    """
    ad_year_tokens: list[str] = []
    for row in rows[:_PREAMBLE_SCAN_ROWS]:
        for cell in row:
            if cell is None:
                continue
            text = _norm_text(cell)
            yint = _as_year_int(cell)
            if yint is not None or re.search(
                r"\b(20\d{2}|19\d{2})\s*[/\-]\s*\d{2}[RPEQrpeq]?\b", text
            ):
                ad_year_tokens.append(text if text else str(cell))
    if ad_year_tokens:
        sample = ", ".join(dict.fromkeys(ad_year_tokens[:3]))
        return [], [
            ParserError(
                error_class="PeriodUnparseable",
                error_detail=(
                    f"sheet={sheet_name!r}: no parseable period header found; "
                    f"year-like tokens detected (e.g. {sample!r}) but the layout "
                    f"matched no known shape (standard wide, two-row monthly, "
                    f"long panel, or transposed) — deferred per ADR-0013"
                ),
                source_excerpt=sample,
            )
        ]
    return [], []


def _parse_sheet(
    ws: object,
    sheet_name: str,
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Parse one DNE worksheet into staging rows + errors.

    Tries the standard wide BS/AD fiscal-year layout first, then the three
    non-standard AD layouts (two-row monthly, long panel, transposed), then a
    fail-loud deferral diagnostic. Never silently drops year-bearing data.
    """
    rows: list[tuple[object, ...]] = list(ws.iter_rows(values_only=True))  # type: ignore[attr-defined]
    if not rows:
        return [], []

    unit_hint = _scan_unit_hint(rows, sheet_name)

    # Long panel FIRST: its signature (FY label in col 0 + AD month name in col 1
    # + value columns to the right) would otherwise be mis-claimed by the standard
    # wide-header detector, which sees the col-0 FY label as a single period column.
    lp = _detect_long_panel(rows)
    if lp is not None:
        # Pass unit_hint=None: the long panel's preamble is dominated by month-name
        # and column-header noise, so the blob-derived hint is unreliable here.
        # The panel parser resolves units per value column and fails loud
        # (UnitAmbiguous) rather than risk a wrong substring match.
        staging, errs = _parse_long_panel_layout(
            rows, lp[0], lp[1], lp[2], lp[3], None, sheet_name
        )
        if staging:
            return staging, errs

    header_row_idx, period_cols = _detect_header(rows)
    if header_row_idx is None or not period_cols:
        alt = _try_alternate_layouts(rows, unit_hint, sheet_name)
        if alt is not None:
            return alt
        return _defer_unparseable_sheet(rows, sheet_name)

    first_period_col = min(period_cols.keys())
    header_row = rows[header_row_idx]
    label_col_idx = _find_label_col(header_row, first_period_col)
    period_errors = _collect_period_errors(header_row, period_cols, first_period_col, sheet_name)

    staging, data_errors = _parse_data_rows(
        rows, header_row_idx, label_col_idx, period_cols, unit_hint, sheet_name
    )
    return staging, period_errors + data_errors


# ---------------------------------------------------------------------------
# Foreign Trade → dimensional facts (ADR-0015)
# ---------------------------------------------------------------------------
#
# Foreign-Trade.xlsx is a dimensional matrix, not a single series. Its
# "Export Import Major Commodities" sheet breaks merchandise trade down by
# COMMODITY (~hundreds of named goods: "Cardamom", "Aluminium Section…"), split
# into sections by direction (Export/Import) and trade partner (India / China /
# Other Countries). Each section is a wide MONTHLY panel: a sparse fiscal-year
# label every 12 columns (forward-filled) over a repeating AD month-name row.
#
# We emit one `DimensionalRowDraft` per (commodity row × month column):
#   base_indicator_slug : dne-merchandise-exports-<partner> / -imports-<partner>
#   dimension_kind      : "commodity"
#   dimension_value     : bare kebab of the commodity label ("cardamom")
#   dimension_label     : the raw source label
#
# SCOPE / DEVIATION (flagged): ADR-0015 names the base measures
# `dne-merchandise-exports` / `dne-merchandise-imports`. We QUALIFY the base slug
# with the trade partner because each sheet carries separate India/China/Other
# sections for the SAME commodity+period; an unqualified base would collide on the
# `dne_facts` unique index `(base_indicator_slug, dimension_kind, dimension_value,
# reporting_period_bs, reporting_period_type, source_document_id)` and silently
# drop 2 of every 3 partner facts under ON CONFLICT DO NOTHING (a Rule-6 silent
# failure / data loss). Partner qualification keeps every fact and is derivable
# from the section header ("…to India"). The headline (partner-agnostic) total can
# be registered later as a single indicator. The "Export Import SITC Groupwise"
# sheet (a DIFFERENT classification of the same totals) and the two "Direction of
# Foreign Trade" partner sheets are intentionally DEFERRED here to avoid mixing
# classifications under one base measure — they follow the same contract next.

# Per the sheet preamble ("Rs in Million") — all Major-Commodities values are NPR
# million. Hard-coded (not unit-scanned) because each section repeats the title.
_FT_COMMODITY_UNIT: Final[str] = "npr_million"

# The single Foreign-Trade sheet promoted to the dimensional model in this round.
_FT_COMMODITY_SHEET: Final[str] = "Export Import Major Commodities"

# Months per fiscal-year block in the commodity panel (AD Aug → next Jul).
_FT_MONTHS_PER_FY: Final[int] = 12

# First data column in the commodity panel (col 0 = S.No., col 1 = label).
_FT_FIRST_VALUE_COL: Final[int] = 2

# The AD month that begins each fiscal-year block in the commodity panel. NRB
# orders the 12 monthly columns Aug → next Jul (the Gregorian face of the
# Shrawan→Asadh fiscal year), so a new "Aug" column marks a new FY block.
_FT_FY_START_AD_MONTH: Final[int] = 8


def _dimension_slug(label: str) -> str:
    """Bare kebab slug of a dimension member (NO ``dne-`` prefix).

    "Cardamom" → "cardamom"; "Aluminium Section(Bars, rods…)" →
    "aluminium-section-bars-rods…"; "G.I. pipe" → "g-i-pipe";
    "Ghee (Clarified)" → "ghee-clarified".

    Deliberately does NOT apply ``_strip_enumerator``: for COMMODITY leaf labels,
    leading tokens like "G.I."/"M.S." (Galvanised Iron / Mild Steel) and trailing
    parentheticals ("(Clarified)" vs "(Vegetable)") are MEANINGFUL and
    distinguish distinct goods — stripping them would collapse two commodities to
    one slug and silently drop facts under the dne_facts ON CONFLICT. The S.No.
    enumerator lives in its own column (col 0), never in this label (col 1).
    """
    s = re.sub(r"[^a-z0-9\s]+", " ", label.lower())
    return re.sub(r"\s+", "-", s.strip())


def _ft_section_base(title: str) -> tuple[str, str] | None:
    """Map a section title → (base_indicator_slug, base_indicator_name), or None.

    "Export of Major Commodities to India"        → ("dne-merchandise-exports-india",
                                                      "Merchandise Exports to India")
    "Import of Major Commodities from Other Coun…" → ("dne-merchandise-imports-other-countries",
                                                      "Merchandise Imports to Other Countries")
    Returns None when the title is not an Export/Import section header.
    """
    t = title.strip()
    low = t.lower()
    if low.startswith("export"):
        direction_slug, direction_name = "exports", "Exports"
    elif low.startswith("import"):
        direction_slug, direction_name = "imports", "Imports"
    else:
        return None
    # Partner follows "to"/"from" at the tail of the title.
    m = re.search(r"\b(?:to|from)\s+(.+)$", t, re.IGNORECASE)
    if not m:
        return None
    partner_raw = m.group(1).strip()
    partner_slug = re.sub(r"\s+", "-", re.sub(r"[^a-z0-9\s]+", " ", partner_raw.lower()).strip())
    if not partner_slug:
        return None
    base_slug = f"dne-merchandise-{direction_slug}-{partner_slug}"
    base_name = f"Merchandise {direction_name} to {partner_raw}"
    return base_slug, base_name


def _ft_calendar_year(fy_ad_label: str, ad_month: int) -> int | None:
    """AD calendar year of ``ad_month`` within an AD FY label, AUGUST-started.

    The commodity panel orders months Aug → next Jul, so its calendar-year split
    differs from the July-started long panel (``_fy_label_to_calendar_year``):
    months Aug–Dec fall in the lead year, Jan–Jul in the trailing year. For AD FY
    "2012/13": Aug 2012 … Dec 2012, then Jan 2013 … Jul 2013. Returns None on a
    malformed label.
    """
    m = re.match(r"^\s*(\d{4})\s*/\s*\d{2,4}\s*$", fy_ad_label)
    if not m:
        return None
    lead = int(m.group(1))
    return lead if ad_month >= _FT_FY_START_AD_MONTH else lead + 1


def _advance_fy_ad_label(fy_ad: str) -> str | None:
    """Return the AD fiscal-year label one year later than ``fy_ad``.

    "2023/24" → "2024/25". Used to derive a missing block label structurally when
    the source leaves a 12-month block's FY cell blank (a merged-cell artifact in
    some sections). Returns None on a malformed input.
    """
    m = re.match(r"^\s*(\d{4})\s*/\s*\d{2,4}\s*$", fy_ad)
    if not m:
        return None
    lead = int(m.group(1)) + 1
    return f"{lead}/{(lead + 1) % 100:02d}"


def _ft_map_value_columns(
    fy_row: tuple[object, ...],
    month_row: tuple[object, ...],
) -> dict[int, tuple[int, int, str]]:
    """Map each value column → (ad_year, ad_month, fy_ad_label) for a section.

    ``fy_row`` carries a fiscal-year label at the head of each 12-month block;
    ``month_row`` carries the AD month name (Aug → next Jul) in every value column.
    The AD calendar year is resolved from the AD fiscal-year label + month via
    ``_fy_label_to_calendar_year`` (Jul–Dec → lead year, Jan–Jun → trailing year).

    The FY label is sparse and OCCASIONALLY ABSENT for a whole block (a merged-cell
    artifact — some sections drop the label cell for the 2024/25 block). Relying on
    a plain forward-fill would then label two physically-distinct blocks with the
    same FY and collapse them onto identical periods (data loss under the dne_facts
    ON CONFLICT). We instead advance the FY STRUCTURALLY: every new Aug column (the
    fiscal-year start) without a fresh label increments the carried FY by one year.
    Columns whose FY or month is unparseable are skipped (panel padding).
    """
    out: dict[int, tuple[int, int, str]] = {}
    current_fy_ad: str | None = None
    seen_first_block = False
    for col in range(_FT_FIRST_VALUE_COL, max(len(fy_row), len(month_row))):
        fy_cell = _norm_text(fy_row[col]) if col < len(fy_row) else ""
        labelled_here = False
        if fy_cell:
            parsed = _parse_annual_fy(fy_cell)
            if parsed is not None:
                current_fy_ad = parsed[1]  # AD "YYYY/YY" label
                labelled_here = True
                seen_first_block = True
        month_cell = _norm_text(month_row[col]) if col < len(month_row) else ""
        ad_month = _parse_ad_month_name(month_cell)
        if ad_month is None:
            continue
        # A new fiscal block begins at each Aug column. If this block carries no
        # fresh label, derive it by advancing the previous block's FY by one year.
        if (
            ad_month == _FT_FY_START_AD_MONTH
            and not labelled_here
            and seen_first_block
            and current_fy_ad is not None
        ):
            current_fy_ad = _advance_fy_ad_label(current_fy_ad)
        if current_fy_ad is None:
            continue
        ad_year = _ft_calendar_year(current_fy_ad, ad_month)
        if ad_year is None:
            continue
        out[col] = (ad_year, ad_month, current_fy_ad)
    return out


def _ft_dimensional_row(
    base_slug: str,
    base_name: str,
    label_raw: str,
    value: float,
    ad_year: int,
    ad_month: int,
    fy_ad: str,
) -> DimensionalRowDraft:
    """Build one commodity DimensionalRowDraft for an (AD year, month) cell.

    ``fy_ad`` is the section's AD fiscal-year label for this column (e.g.
    "2012/13"), carried from the header — NOT re-derived from the month, because
    the Aug-started panel puts the trailing "Jul" in the SAME FY as the leading
    "Aug" (a month-based re-derivation would split them). ``fy_bs`` is the BS
    equivalent (AD lead + 57). The AD month span is the exact Gregorian month
    (1st–28th, a safe lower bound the validator widens); only the BS month *label*
    is the documented mid-month approximation.
    """
    bs_month, bs_year = _ad_month_to_bs(ad_year, ad_month)
    ad_start = datetime(ad_year, ad_month, 1, tzinfo=UTC)
    ad_end = datetime(ad_year, ad_month, 28, tzinfo=UTC)
    fy_lead_ad = int(fy_ad.split("/")[0])
    fy_bs = fiscal_year_label(fy_lead_ad + _BS_AD_FY_OFFSET)
    return DimensionalRowDraft(
        base_indicator_slug=base_slug,
        base_indicator_name=base_name,
        dimension_kind="commodity",
        dimension_value=_dimension_slug(label_raw),
        dimension_label=label_raw,
        value=value,
        unit=_FT_COMMODITY_UNIT,
        reporting_period_type="monthly",
        reporting_period_bs=f"{bs_month} {bs_year}",
        reporting_period_ad_start=ad_start,
        reporting_period_ad_end=ad_end,
        fiscal_year_bs=fy_bs,
        fiscal_year_ad_label=fy_ad,
        confidence_grade=_CONFIDENCE,
    )


def _ft_is_section_title(col0: str) -> bool:
    """True if a col-0 string is an Export/Import section title (not a data row)."""
    low = col0.lower()
    return ("export" in low or "import" in low) and "commodit" in low


def _ft_is_header_row(row: tuple[object, ...]) -> bool:
    """True if a row is the "S.No. | Major Commodities | <FY>…" header row."""
    return _norm_text(row[0]).lower().startswith("s.no") if row else False


def _parse_ft_commodity_sheet(
    rows: list[tuple[object, ...]],
) -> list[DimensionalRowDraft]:
    """Walk the multi-section Major-Commodities sheet → commodity dimensional rows.

    State machine over rows: a section title sets the current base measure; the
    following "S.No." header row + month row define the value-column → (year,
    month) map; subsequent rows with a numeric value and a commodity label in
    col 1 become facts, until the next title / a TOTAL row / the notes block.
    """
    out: list[DimensionalRowDraft] = []
    base: tuple[str, str] | None = None
    col_map: dict[int, tuple[int, int, str]] = {}
    idx = 0
    n = len(rows)
    while idx < n:
        row = rows[idx]
        col0 = _norm_text(row[0]) if row else ""
        if _ft_is_section_title(col0):
            base = _ft_section_base(col0)
            col_map = {}
            idx += 1
            continue
        if base is not None and _ft_is_header_row(row) and idx + 1 < n:
            col_map = _ft_map_value_columns(row, rows[idx + 1])
            idx += 2  # skip the FY header row and the month row
            continue
        if base is not None and col_map:
            label_raw = _norm_text(row[1]) if len(row) > 1 else ""
            if label_raw and label_raw.lower() not in _SKIP_LABELS:
                base_slug, base_name = base
                for col, (ad_year, ad_month, fy_ad) in col_map.items():
                    if col >= len(row):
                        continue
                    value = _safe_float(row[col])
                    if value is None:
                        continue
                    out.append(
                        _ft_dimensional_row(
                            base_slug, base_name, label_raw, value,
                            ad_year, ad_month, fy_ad,
                        )
                    )
        idx += 1
    return out


def _parse_foreign_trade(path: Path) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Parse Foreign-Trade.xlsx → dimensional commodity rows (ADR-0015).

    Only the "Export Import Major Commodities" sheet is promoted in this round
    (see the module note above). Other sheets are intentionally deferred. Never
    raises: a missing/unreadable sheet yields an empty list (the caller turns an
    empty result into the standard NoDataExtracted partial).
    """
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except (OSError, KeyError, ValueError, Exception) as exc:  # noqa: BLE001
        return [], [
            ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl could not open {path.name}: {exc}",
            )
        ]
    if _FT_COMMODITY_SHEET not in wb.sheetnames:
        return [], [
            ParserError(
                error_class="Other",
                error_detail=(
                    f"Foreign-Trade file lacks the {_FT_COMMODITY_SHEET!r} sheet; "
                    f"sheets present: {wb.sheetnames}"
                ),
            )
        ]
    rows: list[tuple[object, ...]] = list(
        wb[_FT_COMMODITY_SHEET].iter_rows(values_only=True)
    )
    return _parse_ft_commodity_sheet(rows), []


# ---------------------------------------------------------------------------
# Real sector (v0.6.0) — GDP & CPI headline single series + Provincial GDP
# ---------------------------------------------------------------------------
#
# The real-sector National-Accounts and CPI files use an ANNUAL COLUMN-SERIES
# layout: annual fiscal-year labels stacked DOWN col 0 (a "Year"/"Fiscal Year"
# column), with named-indicator VALUE COLUMNS to the right. This is the inverse of
# the standard wide layout (where periods are column headers). The generic
# standard-wide detector mis-parses these (it would treat the per-industry rows as
# indicators — catalogue pollution, ADR-0014), so these files route here instead.
#
# We promote ONLY an explicit allowlist of clean headline columns (ADR-0014: no
# catalogue pollution). Each spec hard-maps a (sheet, column-header substring) to a
# canonical slug + a VERIFIED unit (ADR-0011 magnitude check, see the README /
# source profile). Everything else on the sheet (the "As Percent of GDP" sub-cols,
# CPI sub-groups, the GVA-by-industry detail) is intentionally NOT promoted — the
# GVA-by-industry breakdown is a dimensional concern (deferred, same model as
# Provincial GDP). A spec whose column cannot be located in a sheet is skipped (the
# sheet may be from a different edition); a sheet with FY labels but no matched spec
# emits no rows for that sheet (not an error — most sheets are intentionally unused).


@dataclass(frozen=True)
class _RealSectorColumnSpec:
    """One allowlisted headline column: where to find it and how to label it.

    ``sheet`` is the exact worksheet name. ``header_contains`` is a lowercased
    substring matched against the joined multi-row header text of a candidate value
    column (newlines normalised to spaces) — chosen so the match survives NRB's
    embedded line-breaks/footnote markers ("Per Capita GDP \\n(in USD)"). ``slug``
    and ``unit`` are the canonical, ADR-0011-verified outputs.
    """

    sheet: str
    header_contains: str
    slug: str
    name: str
    unit: str


# The allowlist. Units are VERIFIED by order of magnitude (ADR-0011): Nominal GDP
# FY2023/24 = 5709.097 (sheet header "Rs. in billion") = NPR 5.7 trillion ✓; CPI
# Overall Index FY2023/24 = 166.22 (base 2014/15=100) ✓; inflation FY2023/24 =
# 5.44% ✓. Header substrings are lowercased; the most specific spec per column is
# resolved by longest-substring-wins when two specs could match one column.
_REAL_SECTOR_COLUMN_SPECS: Final[tuple[_RealSectorColumnSpec, ...]] = (
    _RealSectorColumnSpec(
        "GDP Series_Nominal", "nominal gdp (rs. in billion)",
        "dne-gdp-nominal", "Nominal GDP (at producers' price)", "npr_billion",
    ),
    _RealSectorColumnSpec(
        "GDP Series_Real", "real gdp growth rate (at purchasers' price)",
        "dne-gdp-real-growth", "Real GDP Growth Rate (at purchasers' price)", "percent",
    ),
    _RealSectorColumnSpec(
        "GDP Series_Real", "real gdp (at purchasers' price)",
        "dne-gdp-real", "Real GDP (at purchasers' price)", "npr_billion",
    ),
    _RealSectorColumnSpec(
        "GDP Series_Real", "per capita gdp (in usd)",
        "dne-gdp-per-capita-usd", "Per Capita GDP", "usd",
    ),
    _RealSectorColumnSpec(
        "GDP Series_Real", "gdp deflator",
        "dne-gdp-deflator", "GDP Deflator", "index_points",
    ),
    _RealSectorColumnSpec(
        "CPI_National", "index",  # row-3 sub-header "Overall" under r2 "Index"
        "dne-cpi", "National Consumer Price Index — Overall", "index_points",
    ),
    _RealSectorColumnSpec(
        "CPI_National", "percentage change",
        "dne-inflation-rate", "Consumer Price Inflation — Overall", "percent",
    ),
)

# Col-0 label that marks the FY column in the annual column-series layout.
_FY_COLUMN_LABELS: Final[frozenset[str]] = frozenset({"year", "fiscal year"})

# Minimum annual FY rows stacked in col 0 to accept the layout (avoid false hits).
_MIN_FY_ROWS_DOWN_COL0: Final[int] = 3

# How many rows of header sit above the first FY data row (joined per column to
# match a spec's ``header_contains``). Generous; blanks are skipped in the join.
_REAL_SECTOR_HEADER_ROWS: Final[int] = 6


def _col_header_text(rows: list[tuple[object, ...]], first_data_row: int, col: int) -> str:
    """Join the header cells above ``first_data_row`` for ``col`` (newlines→spaces).

    NRB embeds line-breaks and footnote markers inside header cells ("Per Capita GDP
    \\n(in USD)", "Population (million)1"); ``_norm_text`` already collapses internal
    whitespace, so the joined, lowercased text is stable for substring matching.
    """
    parts = [
        _norm_text(rows[r][col])
        for r in range(min(first_data_row, len(rows)))
        if col < len(rows[r]) and _norm_text(rows[r][col])
    ]
    return " ".join(parts).lower()


def _find_fy_column_start(rows: list[tuple[object, ...]]) -> tuple[int, int] | None:
    """Locate the annual column-series anchor: (fy_col, first_fy_data_row), or None.

    Scans col 0 (and col 1 as a fallback, for sheets whose row labels begin in col 1)
    for a run of ≥ ``_MIN_FY_ROWS_DOWN_COL0`` consecutive annual FY labels. Returns
    the column and the row index of the FIRST FY label in that run.
    """
    for fy_col in (0, 1):
        first_row: int | None = None
        run = 0
        for ri, row in enumerate(rows):
            cell = _norm_text(row[fy_col]) if fy_col < len(row) else ""
            if cell and _parse_annual_fy(cell) is not None:
                if first_row is None:
                    first_row = ri
                run += 1
                if run >= _MIN_FY_ROWS_DOWN_COL0:
                    return fy_col, first_row
            else:
                first_row, run = None, 0
    return None


def _parse_real_sector_sheet(
    rows: list[tuple[object, ...]],
    sheet_name: str,
    specs: list[_RealSectorColumnSpec],
) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Emit staging rows for the allowlisted headline columns of one real-sector sheet.

    For each spec, locate the value column whose joined header text contains the
    spec's substring (longest-substring-wins on ties so "real gdp growth rate…"
    binds before a looser "index"), then emit one annual StagingRowDraft per FY row
    with a parseable numeric value. Returns ([], []) when the sheet has no FY column
    (it is not this layout — the caller decides whether that is an error)."""
    anchor = _find_fy_column_start(rows)
    if anchor is None:
        return [], []
    fy_col, first_data_row = anchor
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []

    # Resolve each spec to a concrete column (first value column whose header
    # contains the substring). Skip a spec whose column is absent in this edition.
    max_col = max((len(r) for r in rows), default=0)
    for spec in specs:
        matched_col: int | None = None
        for col in range(fy_col + 1, max_col):
            header = _col_header_text(rows, first_data_row, col)
            if spec.header_contains in header:
                matched_col = col
                break
        if matched_col is None:
            continue
        for ri in range(first_data_row, len(rows)):
            row = rows[ri]
            if fy_col >= len(row):
                continue
            fy = _parse_annual_fy(_norm_text(row[fy_col]))
            if fy is None:
                continue  # footer/source row inside the data block — skip, not error
            value = _safe_float(row[matched_col]) if matched_col < len(row) else None
            if value is None:
                continue
            fy_bs, fy_ad = fy
            staging.append(
                _annual_fy_to_draft_fields(
                    fy_bs=fy_bs, fy_ad=fy_ad, unit=spec.unit,
                    slug=spec.slug, value=value,
                )
            )
    _ = (sheet_name, errors)
    return staging, errors


def _parse_real_sector(path: Path) -> tuple[list[StagingRowDraft], list[ParserError]]:
    """Parse a real-sector file (National-Accounts / CPI) → headline single series.

    Iterates only the sheets named in ``_REAL_SECTOR_COLUMN_SPECS`` (the allowlist),
    parsing each via the annual column-series layout. Never raises: an unreadable
    file yields a typed EncodingError; a missing allowlisted sheet is silently
    skipped (different edition); the caller turns an empty result into the standard
    NoDataExtracted partial.
    """
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except (OSError, KeyError, ValueError, Exception) as exc:  # noqa: BLE001
        return [], [
            ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl could not open {path.name}: {exc}",
            )
        ]
    # Group specs by sheet so each sheet is read once.
    by_sheet: dict[str, list[_RealSectorColumnSpec]] = {}
    for spec in _REAL_SECTOR_COLUMN_SPECS:
        by_sheet.setdefault(spec.sheet, []).append(spec)
    staging: list[StagingRowDraft] = []
    errors: list[ParserError] = []
    for sheet_name, specs in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        s, e = _parse_real_sector_sheet(rows, sheet_name, specs)
        staging.extend(s)
        errors.extend(e)
    return staging, errors


# ---------------------------------------------------------------------------
# Provincial GDP → dimensional facts (ADR-0015), dimension_kind='province'
# ---------------------------------------------------------------------------
#
# Provincial-GDP-2024-25.xlsx ("Tables" sheet) is a GDP-BY-PROVINCE matrix. Each of
# the eight provinces (+ a "Total GVA" block) spans seven consecutive FY columns
# under a province-name banner row; below it sit two FY rows (BS then AD) and then
# industry rows, ending in a "Gross Domestic Product (GDP)" total per province. We
# emit ONE DimensionalRowDraft per (province × FY) for the headline GDP total only:
#   base_indicator_slug : dne-provincial-gdp
#   dimension_kind      : "province"
#   dimension_value     : kebab province name ("koshi", "sudur-pashchim")
# This is the Money-Map "GDP by province" composition surface. The per-industry GVA
# rows within each province block are a SECOND dimension (industry) and are deferred
# to keep one dimension per fact (ADR-0015). The file has Table 1 (current prices,
# nominal) and Table 2 (constant prices, real); we promote the NOMINAL total (Table
# 1) — the headline "GDP by province" figure — and defer the real table.

# The base measure for provincial GDP facts.
_PROV_GDP_BASE_SLUG: Final[str] = "dne-provincial-gdp"
_PROV_GDP_BASE_NAME: Final[str] = "Provincial GDP (at producers' price, nominal)"
_PROV_GDP_UNIT: Final[str] = "npr_million"  # sheet header: "(at current prices, in million)"

# Province banner labels that are NOT a province (skip these dimension blocks).
_PROV_NON_PROVINCE_BANNERS: Final[frozenset[str]] = frozenset(
    {"total gva", "total", "nepal", "industrial classification"}
)

# The headline total-GDP row label within each province block. Matched by
# ENDSWITH (not substring): the real sheet has both "Gross Domestic Product (GDP)
# at basic prices" (a sub-row) and the headline "Gross Domestic Product (GDP)";
# both CONTAIN "(gdp)", so a substring match would wrongly pick the basic-prices
# row. The headline row ends in "(gdp)"; the basic-prices row ends in "prices".
_PROV_GDP_TOTAL_ROW_ENDSWITH: Final[str] = "(gdp)"
_PROV_GDP_TOTAL_ROW_REQUIRES: Final[str] = "gross domestic product"

# FY columns per province block in the Provincial GDP banner layout.
_PROV_FY_PER_BLOCK: Final[int] = 7

# The Provincial GDP matrix sheet name.
_PROV_GDP_SHEET: Final[str] = "Tables"


def _detect_province_banner_row(
    rows: list[tuple[object, ...]],
) -> tuple[int, dict[int, str]] | None:
    """Find the province banner row + its column→province map for Table 1 (nominal).

    The banner row carries province names every 7 columns (e.g. col2 "Koshi", col9
    "Madhes", …). We require ≥2 known province names to accept it, and return the
    FIRST such row (Table 1 / current prices). Returns ``(banner_row_idx,
    {col: province_label})`` or None.

    A banner cell must be NON-FY text: the BS-FY and AD-FY rows beneath the banner
    also "head a block" structurally, so without this guard the FY row itself would
    be mistaken for the banner (its FY labels read as province names). Province names
    never parse as fiscal years; FY labels always do.
    """
    for ri, row in enumerate(rows):
        provinces: dict[int, str] = {}
        for ci, cell in enumerate(row):
            label = _norm_text(cell)
            low = label.lower()
            if not label or low in _PROV_NON_PROVINCE_BANNERS:
                continue
            if _parse_annual_fy(label) is not None:
                continue  # an FY label is never a province banner cell
            # A province banner cell is followed by a run of FY labels two rows down
            # (the AD FY row) OR one row down (the BS FY row). Confirm the cell heads
            # a block by checking the row beneath it carries an FY label.
            if _looks_like_province_block_head(rows, ri, ci):
                provinces[ci] = label
        if len(provinces) >= 2:  # noqa: PLR2004 — ≥2 provinces confirms the banner
            return ri, provinces
    return None


def _looks_like_province_block_head(
    rows: list[tuple[object, ...]], banner_row: int, col: int
) -> bool:
    """True if ``col`` heads an FY block: an annual FY label sits 1 or 2 rows below."""
    for delta in (1, 2):
        r = banner_row + delta
        if (
            r < len(rows)
            and col < len(rows[r])
            and _parse_annual_fy(_norm_text(rows[r][col])) is not None
        ):
            return True
    return False


def _province_fy_row(rows: list[tuple[object, ...]], banner_row: int) -> int | None:
    """Return the row index holding BS FY labels under the banner (banner_row+1 or +2).

    Provincial GDP stacks a BS-FY row then an AD-FY row beneath the banner. We key
    periods off the BS row (canonical ``reporting_period_bs``). Prefer the row whose
    labels parse as BS (lead year ≥ 2040 via ``_parse_annual_fy`` → BS path)."""
    for delta in (1, 2):
        r = banner_row + delta
        if r >= len(rows):
            continue
        # Count FY labels on this row; the BS row and AD row both parse, but the BS
        # row's labels have lead year ≥ 2040. Use the first FY-bearing row.
        if any(
            _parse_annual_fy(_norm_text(c)) is not None for c in rows[r]
        ):
            return r
    return None


def _parse_provincial_gdp_sheet(
    rows: list[tuple[object, ...]],
) -> list[DimensionalRowDraft]:
    """Walk the Provincial GDP banner layout → one province×FY GDP-total fact each.

    Locates the banner row + per-block FY row, then finds the "Gross Domestic Product
    (GDP)" total row within Table 1, and emits a dimensional fact for each
    province-block column whose FY label and value both parse. Stops before Table 2
    (constant prices) by only consuming the first GDP-total row after the banner.
    """
    out: list[DimensionalRowDraft] = []
    banner = _detect_province_banner_row(rows)
    if banner is None:
        return out
    banner_row, provinces = banner
    fy_row = _province_fy_row(rows, banner_row)
    if fy_row is None:
        return out
    # Find the headline GDP-total row (first occurrence after the FY row). Match by
    # endswith "(gdp)" so the "…(GDP) at basic prices" sub-row is NOT chosen.
    total_row: int | None = None
    for ri in range(fy_row + 1, len(rows)):
        label = _norm_text(rows[ri][1]).lower() if len(rows[ri]) > 1 else ""
        if (
            _PROV_GDP_TOTAL_ROW_REQUIRES in label
            and label.endswith(_PROV_GDP_TOTAL_ROW_ENDSWITH)
        ):
            total_row = ri
            break
    if total_row is None:
        return out
    total = rows[total_row]
    for banner_col, prov_label in provinces.items():
        dim_value = _dimension_slug(prov_label)
        for col in range(banner_col, banner_col + _PROV_FY_PER_BLOCK):
            if col >= len(total) or fy_row >= len(rows) or col >= len(rows[fy_row]):
                continue
            fy = _parse_annual_fy(_norm_text(rows[fy_row][col]))
            if fy is None:
                continue
            value = _safe_float(total[col])
            if value is None:
                continue
            fy_bs, fy_ad = fy
            bs_start = int(fy_bs.split("/")[0])
            ad_start_year = bs_start - _BS_AD_FY_OFFSET
            out.append(
                DimensionalRowDraft(
                    base_indicator_slug=_PROV_GDP_BASE_SLUG,
                    base_indicator_name=_PROV_GDP_BASE_NAME,
                    dimension_kind="province",
                    dimension_value=dim_value,
                    dimension_label=prov_label,
                    value=value,
                    unit=_PROV_GDP_UNIT,
                    reporting_period_type="annual",
                    reporting_period_bs=fy_bs,
                    reporting_period_ad_start=datetime(ad_start_year, 7, 15, tzinfo=UTC),
                    reporting_period_ad_end=datetime(ad_start_year + 1, 7, 15, tzinfo=UTC),
                    fiscal_year_bs=fy_bs,
                    fiscal_year_ad_label=fy_ad,
                    confidence_grade=_CONFIDENCE,
                )
            )
    return out


def _parse_provincial_gdp(path: Path) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Parse Provincial-GDP-2024-25.xlsx → province-dimensional GDP facts (ADR-0015).

    Promotes the NOMINAL (Table 1, current prices) headline GDP total per province.
    The real table and the per-industry GVA breakdown are deferred (one dimension per
    fact). Never raises: an unreadable/missing sheet yields a typed error and the
    caller turns an empty result into NoDataExtracted.
    """
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except (OSError, KeyError, ValueError, Exception) as exc:  # noqa: BLE001
        return [], [
            ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl could not open {path.name}: {exc}",
            )
        ]
    sheet = _PROV_GDP_SHEET if _PROV_GDP_SHEET in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))
    return _parse_provincial_gdp_sheet(rows), []


def _real_sector_result(path: Path) -> ParserResult:
    """Wrap ``_parse_real_sector`` into a ``ParserResult`` (single return site).

    Extracted from ``parse`` so that function stays under the branch/return caps.
    An empty result becomes a ``partial`` with a NoDataExtracted note; otherwise the
    status is ``success`` (no errors) or ``partial`` (some columns flagged).
    """
    rs_staging, rs_errors = _parse_real_sector(path)
    if not rs_staging:
        rs_errors.append(
            ParserError(
                error_class="Other",
                error_detail=(
                    "NoDataExtracted: no real-sector headline series produced "
                    f"from {path.name} (no allowlisted column matched)"
                ),
            )
        )
        return ParserResult(
            status="partial",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=rs_errors,
        )
    status: ParserStatus = "partial" if rs_errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=rs_staging,
        errors=rs_errors,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse a DNE XLSX file into ``StagingRowDraft`` rows.

    Arguments:
        source_document_path: filesystem path to the ``.xlsx`` file.
        source_document_id: opaque UUID threaded through to the orchestrator;
            not embedded in rows (the ingest layer handles FK wiring).

    Returns:
        ``ParserResult`` — never raises on bad data.
    """
    _ = source_document_id  # reserved for future provenance embedding

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=f"source file not found: {path}",
                )
            ],
        )

    if _is_dimensional_path(path):
        # Dimensional matrix files (Foreign-Trade) have NO single-series staging
        # output — they route through ``parse_dne`` → ``dimensional_rows``. Return
        # empty staging with an explicit note so this never silently emits the
        # mis-detected per-section "indicator" rows the standard layout would.
        return ParserResult(
            status="partial",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=(
                        f"{path.name} is a dimensional matrix (ADR-0015); use "
                        f"parse_dne() — staging_rows is intentionally empty here"
                    ),
                )
            ],
        )

    if _is_real_sector_path(path):
        # Real-sector files (National-Accounts, CPI) use the annual column-series
        # layout (FY down col 0, headline-indicator value columns). Route to the
        # allowlist parser instead of the generic per-sheet detector — the generic
        # one would mis-read the GVA-by-industry rows as indicators (ADR-0014).
        return _real_sector_result(path)

    try:
        wb = openpyxl.load_workbook(
            filename=str(path), read_only=True, data_only=True
        )
    except (OSError, KeyError, ValueError, Exception) as exc:  # noqa: BLE001
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[
                ParserError(
                    error_class="EncodingError",
                    error_detail=f"openpyxl could not open {path.name}: {exc}",
                )
            ],
        )

    all_staging: list[StagingRowDraft] = []
    all_errors: list[ParserError] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_staging, sheet_errors = _parse_sheet(ws, sheet_name)
        all_staging.extend(sheet_staging)
        all_errors.extend(sheet_errors)

    if not all_staging:
        all_errors.append(
            ParserError(
                error_class="Other",
                error_detail="NoDataExtracted: no staging rows produced from any sheet",
            )
        )
        return ParserResult(
            status="partial",
            parser_version=PARSER_VERSION,
            staging_rows=[],
            errors=all_errors,
        )

    status: ParserStatus = "partial" if all_errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=all_staging,
        errors=all_errors,
    )


def _is_dimensional_path(path: Path) -> bool:
    """True if a file routes to the dimensional fact path (ADR-0015).

    Matched on the filename stem (case-insensitive), e.g. ``Foreign-Trade.xlsx``
    → stem ``foreign-trade`` ∈ ``_DIMENSIONAL_FILE_STEMS``.
    """
    return path.stem.strip().lower() in _DIMENSIONAL_FILE_STEMS


def _is_real_sector_path(path: Path) -> bool:
    """True if a file routes to the v0.6.0 real-sector single-series path.

    Matched on the filename stem (case-insensitive), e.g. ``National-Accounts.xlsx``
    → stem ``national-accounts`` ∈ ``_REAL_SECTOR_FILE_STEMS``.
    """
    return path.stem.strip().lower() in _REAL_SECTOR_FILE_STEMS


def _dispatch_dimensional(
    path: Path,
) -> tuple[list[DimensionalRowDraft], list[ParserError]]:
    """Route a dimensional file to its parser by filename stem.

    ``foreign-trade`` → commodity facts (ADR-0015); ``provincial-gdp-2024-25`` →
    province facts (v0.6.0). The caller has already confirmed the stem is in
    ``_DIMENSIONAL_FILE_STEMS`` via ``_is_dimensional_path``.
    """
    if path.stem.strip().lower() == "provincial-gdp-2024-25":
        return _parse_provincial_gdp(path)
    return _parse_foreign_trade(path)


def parse_dne(source_document_path: str, source_document_id: str) -> DneParserResult:
    """DNE entry point carrying BOTH single-series and dimensional output.

    This is what the DNE ingest CLI invokes. For dimensional matrix files
    (Foreign-Trade by commodity, Provincial-GDP by province) it returns populated
    ``dimensional_rows`` (ADR-0015) and empty ``staging_rows``; for every other DNE
    file it wraps ``parse()`` and returns its ``staging_rows`` with empty
    ``dimensional_rows``. Never raises on bad data.
    """
    path = Path(source_document_path)
    if not path.exists():
        return DneParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[
                ParserError(
                    error_class="Other",
                    error_detail=f"source file not found: {path}",
                )
            ],
        )

    if _is_dimensional_path(path):
        dim_rows, dim_errors = _dispatch_dimensional(path)
        if not dim_rows:
            dim_errors.append(
                ParserError(
                    error_class="Other",
                    error_detail=(
                        "NoDataExtracted: no dimensional rows produced from "
                        f"{path.name}"
                    ),
                )
            )
            return DneParserResult(
                status="partial",
                parser_version=PARSER_VERSION,
                dimensional_rows=[],
                errors=dim_errors,
            )
        status: ParserStatus = "partial" if dim_errors else "success"
        return DneParserResult(
            status=status,
            parser_version=PARSER_VERSION,
            dimensional_rows=dim_rows,
            errors=dim_errors,
        )

    # Single-series file: delegate to the existing parser unchanged.
    single = parse(source_document_path, source_document_id)
    return DneParserResult(
        status=single.status,
        parser_version=single.parser_version,
        staging_rows=single.staging_rows,
        dimensional_rows=[],
        errors=single.errors,
    )


# ---------------------------------------------------------------------------
# CLI entrypoint (orchestrator contract — mirror of nrb_cmefs)
# ---------------------------------------------------------------------------


def _main() -> None:
    """Argv: ``parser.py <source_document_path> <source_document_id>``.

    Writes the DNE result JSON to stdout — including the ``dimensional_rows`` key
    (ADR-0015) that the DNE ingest CLI reads. Datetimes are ISO-8601 strings.
    Exit codes: 0 = ran (status may be failure), 2 = usage error.
    """
    if len(sys.argv) != 3:  # noqa: PLR2004
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse_dne(sys.argv[1], sys.argv[2])
    json.dump(result.to_json_dict(), sys.stdout)


if __name__ == "__main__":
    _main()

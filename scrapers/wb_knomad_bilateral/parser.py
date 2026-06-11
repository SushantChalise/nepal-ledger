"""World Bank / KNOMAD bilateral remittance matrix parser — deterministic Python.

Source: World Bank / KNOMAD Bilateral Remittance Estimates (annual Excel matrix).
Source ID: wb-knomad-bilateral-remittance

Background:
    NRB's Balance of Payments data (BPM6) gives Nepal's total remittance inflow
    but provides no corridor breakdown (India/Gulf/Malaysia). KNOMAD publishes
    an annual bilateral remittance matrix (~June each year) covering every
    country pair in USD millions, filling Gap 1 of the Fact Ledger.

Input format:
    An annual .xlsx file from the KNOMAD bilateral remittances page
    (https://www.knomad.org/data/remittances). Manual download required —
    the KNOMAD site requires authentication for file downloads.

    Sheet structure (may vary slightly by year):
      - A sheet whose name contains "bilateral" (case-insensitive), e.g.
        "Bilateral_Remittance" or "Bilateral_Remittance_Estimates"
      - Row 1 (index 0) = column headers = *receiving* country names.
        The Nepal column header may be "Nepal" or "Nepal, Fed. Dem. Rep."
      - Column 1 (index 0 per row) = *sending* country name
      - Values in USD millions (positive float or None)

    Semantics:
      - Row "India", Nepal column → money sent FROM India TO Nepal (inflow)
      - Row "Nepal", India column → money sent FROM Nepal TO India (outflow)

    We extract the Nepal *column* (inflows to Nepal) from key source countries.

AD → BS fiscal year mapping (approximation):
    KNOMAD publishes a matrix labelled with AD calendar year Y. This covers
    remittances received during AD year Y (calendar year Jan–Dec), which the
    World Bank assigns to Nepal's fiscal year starting mid-July of Y.
    Approximation: AD year Y → BS FY (Y + 57)/((Y + 58) % 100):
      - AD 2024 → BS FY 2081/82  (Jul 2024 – Jul 2025)
      - AD 2023 → BS FY 2080/81
      - AD 2022 → BS FY 2079/80

    The KNOMAD calendar-year figure is NOT the same as NRB's FY remittance
    total (which runs Shrawan–Ashadh). The parser notes this in parser_notes
    so the validation layer can apply appropriate tolerance.

Indicator slugs emitted:
    knomad-remittance-to-nepal-from-india-annual          (usd_million)
    knomad-remittance-to-nepal-from-qatar-annual          (usd_million)
    knomad-remittance-to-nepal-from-uae-annual            (usd_million)
    knomad-remittance-to-nepal-from-saudi-arabia-annual   (usd_million)
    knomad-remittance-to-nepal-from-kuwait-annual         (usd_million)
    knomad-remittance-to-nepal-from-bahrain-annual        (usd_million)
    knomad-remittance-to-nepal-from-oman-annual           (usd_million)
    knomad-remittance-to-nepal-from-malaysia-annual       (usd_million)
    knomad-remittance-to-nepal-from-usa-annual            (usd_million)
    knomad-remittance-to-nepal-from-australia-annual      (usd_million)
    knomad-remittance-to-nepal-from-japan-annual          (usd_million)
    knomad-remittance-to-nepal-from-korea-annual          (usd_million)
    knomad-remittance-to-nepal-total-annual               (usd_million)

Versioning:
    Bump PARSER_VERSION on any behaviour change.
"""

from __future__ import annotations

import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Final

import openpyxl

from _common.periods import fiscal_year_ad_label, fiscal_year_label
from _common.types import ParserError, ParserResult, StagingRowDraft

PARSER_VERSION: Final[str] = "0.1.0"
SOURCE_ID: Final[str] = "wb-knomad-bilateral-remittance"

# Source countries we track (substring match, case-insensitive).
# Value = slug suffix used in indicator_slug_raw.
# Order matters: more-specific matches must come before substrings they contain
# to avoid "korea" matching "north korea" as "korea". We match on
# normalised lower-case sending-country name.
_COUNTRY_SLUGS: Final[dict[str, str]] = {
    "india": "india",
    "qatar": "qatar",
    "united arab emirates": "uae",
    "saudi arabia": "saudi-arabia",
    "kuwait": "kuwait",
    "bahrain": "bahrain",
    "oman": "oman",
    "malaysia": "malaysia",
    "united states": "usa",
    "australia": "australia",
    "japan": "japan",
    # Korea (Republic) — match after "north korea" check is done via priority order
    "korea, rep": "korea",
    "korea rep": "korea",
    "south korea": "korea",
}

# Slug suffix for the "World" / grand-total row.
_TOTAL_SLUG_SUFFIX: Final[str] = "total"

# Publication approximation: KNOMAD releases the matrix in June of year Y+1
# (i.e. the 2024 matrix is released ~June 2025).
_PUB_MONTH: Final[int] = 6
_PUB_DAY: Final[int] = 15


def _find_bilateral_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet | None:
    """Return the first sheet whose name contains 'bilateral' (case-insensitive)."""
    for name in wb.sheetnames:
        if "bilateral" in name.lower():
            return wb[name]  # type: ignore[return-value]
    # Fallback: first sheet if only one exists
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]  # type: ignore[return-value]
    return None


def _find_nepal_col(header_row: list[object]) -> int | None:
    """Return the 0-based column index whose header contains 'Nepal'."""
    for i, cell in enumerate(header_row):
        if cell is not None and "nepal" in str(cell).lower():
            return i
    return None


def _infer_year_from_filename(path: Path) -> int | None:
    """Extract a 4-digit calendar year from the filename (e.g. '...2024.xlsx' → 2024)."""
    m = re.search(r"(20\d{2})", path.stem)
    if m:
        return int(m.group(1))
    return None


def _infer_year_from_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    """Scan the first row and column A for a 4-digit year cell."""
    for row in ws.iter_rows(max_row=3, max_col=10, values_only=True):
        for cell in row:
            if cell is not None:
                m = re.search(r"(20\d{2})", str(cell))
                if m:
                    return int(m.group(1))
    return None


def _match_source_country(name: str) -> str | None:
    """Return slug suffix if *name* matches a tracked source country, else None."""
    lower = name.lower().strip()
    # Total / World row
    if lower in ("world", "total", "grand total"):
        return _TOTAL_SLUG_SUFFIX
    # Exact-first pass: avoid substring false positives (north korea, etc.)
    for pattern, slug in _COUNTRY_SLUGS.items():
        if pattern in lower:
            # Extra guard: reject "north korea" matching "korea"
            if slug == "korea" and "north" in lower:
                continue
            return slug
    return None


def _bs_period_strings(ad_year: int) -> tuple[str, str, str, str]:
    """Return (reporting_period_bs, fiscal_year_bs, fiscal_year_ad_label, publication_date_bs)."""
    bs_fy_start = ad_year + 57
    fy_bs = fiscal_year_label(bs_fy_start)
    fy_ad = fiscal_year_ad_label(bs_fy_start)
    # Reporting period label: KNOMAD covers a calendar year, noted as such
    period_bs = f"CY{ad_year} (≈ FY {fy_bs})"
    # Publication date BS: approximate (KNOMAD publishes ~June of AD year+1)
    pub_bs = f"~{fiscal_year_label(bs_fy_start + 1)} Jestha (heuristic)"
    return period_bs, fy_bs, fy_ad, pub_bs


def parse(source_document_path: str, source_document_id: str) -> ParserResult:
    """Parse a KNOMAD bilateral remittance XLSX; emit corridor indicators for Nepal.

    Arguments:
        source_document_path: filesystem path to the downloaded XLSX.
        source_document_id: opaque FK from ``source_documents``; threaded
            through for orchestrator symmetry.

    Returns:
        ``ParserResult`` with ``status``, ``staging_rows``, ``errors``.
    """
    _ = source_document_id

    path = Path(source_document_path)
    if not path.exists():
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=f"source file not found: {path}",
            )],
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="EncodingError",
                error_detail=f"openpyxl failed to open workbook: {exc}",
            )],
        )

    ws = _find_bilateral_sheet(wb)
    if ws is None:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="ColumnMissing",
                error_detail=(
                    f"no sheet named 'bilateral*' found in workbook "
                    f"(sheets: {wb.sheetnames})"
                ),
            )],
        )

    # Read all rows (read_only mode: consume the iterator once)
    all_rows: list[list[object]] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    if not all_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail="bilateral sheet is empty",
            )],
        )

    header_row = all_rows[0]
    nepal_col = _find_nepal_col(header_row)
    if nepal_col is None:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="ColumnMissing",
                error_detail=(
                    "Nepal not found in header row — column header may have changed. "
                    f"Header (first 20 cols): {[str(v)[:20] if v else None for v in header_row[:20]]}"
                ),
            )],
        )

    # Infer calendar year from filename first, then sheet content
    ad_year = _infer_year_from_filename(path)
    if ad_year is None:
        ad_year = _infer_year_from_sheet(ws)  # type: ignore[arg-type]
    if ad_year is None:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="PeriodAmbiguous",
                error_detail=(
                    "cannot infer calendar year from filename or sheet content. "
                    "Filename must contain a 4-digit year (e.g. knomad-bilateral-2024.xlsx)."
                ),
            )],
        )

    period_bs, fy_bs, fy_ad, pub_bs = _bs_period_strings(ad_year)

    # AD date bounds: calendar year Jan 1 – Dec 31 (approximated to mid-month)
    reporting_ad_start = datetime(ad_year, 1, 15, tzinfo=UTC)
    reporting_ad_end = datetime(ad_year, 12, 15, tzinfo=UTC)
    publication_ad = datetime(ad_year + 1, _PUB_MONTH, _PUB_DAY, tzinfo=UTC)

    errors: list[ParserError] = []
    staging_rows: list[StagingRowDraft] = []

    for row in all_rows[1:]:  # skip header row
        if not row:
            continue
        sending_country_raw = row[0]
        if sending_country_raw is None:
            continue
        sending_country = str(sending_country_raw).strip()
        if not sending_country:
            continue

        slug_suffix = _match_source_country(sending_country)
        if slug_suffix is None:
            continue  # not a tracked country

        # Value in Nepal column
        raw_val = row[nepal_col] if nepal_col < len(row) else None
        if raw_val is None:
            # Missing value — skip silently (many corridor pairs are zero/blank)
            continue
        try:
            value = float(raw_val)
        except (ValueError, TypeError):
            errors.append(ParserError(
                error_class="ValueUnparseable",
                error_detail=(
                    f"sending country {sending_country!r}: "
                    f"cannot parse Nepal-column value {raw_val!r}"
                ),
                source_excerpt=f"row col0={sending_country!r} nepal_col={raw_val!r}",
            ))
            continue

        indicator_slug = f"knomad-remittance-to-nepal-from-{slug_suffix}-annual"
        if slug_suffix == _TOTAL_SLUG_SUFFIX:
            indicator_slug = "knomad-remittance-to-nepal-total-annual"

        staging_rows.append(StagingRowDraft(
            indicator_slug_raw=indicator_slug,
            value=value,
            unit="usd_million",
            reporting_period_type="annual",
            reporting_period_bs=period_bs,
            reporting_period_ad_start=reporting_ad_start,
            reporting_period_ad_end=reporting_ad_end,
            publication_date_ad=publication_ad,
            publication_date_bs=pub_bs,
            fiscal_year_bs=fy_bs,
            fiscal_year_ad_label=fy_ad,
            confidence_grade_proposed="A",
            parser_notes=(
                f"KNOMAD calendar-year {ad_year} matrix. "
                "Covers Jan–Dec (not Nepal FY Shrawan–Ashadh); "
                "approximate fiscal-year mapping applied. "
                "Sending country (row): "
                f"{sending_country!r}."
            ),
        ))

    if not staging_rows and not errors:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=[ParserError(
                error_class="Other",
                error_detail=(
                    "no tracked sending countries found in bilateral sheet. "
                    "Sheet may have changed structure or country names differ."
                ),
            )],
        )

    if not staging_rows:
        return ParserResult(
            status="failure",
            parser_version=PARSER_VERSION,
            errors=errors,
        )

    from _common.types import ParserStatus
    status: ParserStatus = "partial" if errors else "success"
    return ParserResult(
        status=status,
        parser_version=PARSER_VERSION,
        staging_rows=staging_rows,
        errors=errors,
    )


def _main() -> None:
    """CLI entrypoint used by the Node ingestion orchestrator.

    Argv: ``parser.py <source_document_path> <source_document_id>``.
    Writes JSON to stdout. Exit codes: 0 = ran (status may be 'failure'), 2 = usage error.
    """
    import json
    import sys
    from dataclasses import asdict

    if len(sys.argv) != 3:
        sys.stderr.write(
            "usage: parser.py <source_document_path> <source_document_id>\n"
        )
        sys.exit(2)

    result = parse(sys.argv[1], sys.argv[2])
    payload = asdict(result)
    for row in payload.get("staging_rows", []):
        for key in (
            "reporting_period_ad_start",
            "reporting_period_ad_end",
            "publication_date_ad",
        ):
            val = row.get(key)
            if isinstance(val, datetime):
                row[key] = val.isoformat()
    json.dump(payload, sys.stdout)


if __name__ == "__main__":
    _main()

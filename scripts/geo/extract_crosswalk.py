"""Emit the canonical 753 local-level crosswalk (federal_code ↔ name ↔ district).

This is the JOIN KEY table for the choropleth geometry pipeline (ADR-0025):
external boundary GeoJSON carries names + district but NOT the MoFAGA 8-digit
federal_code that `administrative_units.federal_code` / `entities.slug` use.
This script emits the authoritative (federal_code, name_en, name_ne,
district_en, type) rows so `build_admin_topojson` can resolve every geometry
feature to its federal_code by (district, name, type) match.

Source of truth: the MoF pre-cleaned workbook Sheet2 — the SAME table
`scripts/_seed-helpers/extract_local_levels.py` reads for the entities seed,
so the crosswalk is guaranteed consistent with the seeded entities. We read it
with openpyxl directly (no pandas dependency) so the pipeline runs in a bare
worktree venv.

Usage:
    python scripts/geo/extract_crosswalk.py <xlsx_path> [out_csv_path]

Writes UTF-8 CSV (default scripts/geo/crosswalk.csv) with header
    federal_code,name_en,name_ne,district_en,local_level_type
filtered to exactly the 753 real local levels (aggregator rows dropped).
We write the file directly (not stdout) because the Nepali (Devanagari)
names break Windows' cp1252 stdout redirection.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

CANONICAL_SHEET: Final[str] = "Sheet2"

_REAL_TYPES: Final[dict[str, str]] = {
    "Municipality": "municipality",
    "Rural Municipality": "rural_municipality",
    "Metropolitan City": "metropolitan_city",
    "Sub-Metropolitan City": "sub_metropolitan_city",
}

# Sheet2 column headers (must match the workbook; see extract_local_levels.py).
_COL_CODE: Final[str] = "Code"
_COL_DISTRICT: Final[str] = "District (English)"
_COL_NAME_NE: Final[str] = "Local Level Name (Nepali)"
_COL_NAME_EN: Final[str] = "Local Level Name (English)"
_COL_TYPE: Final[str] = "Local Level Type"


def _emit(path: Path) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[CANONICAL_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    idx = {name: header.index(name) for name in (
        _COL_CODE, _COL_DISTRICT, _COL_NAME_NE, _COL_NAME_EN, _COL_TYPE,
    )}
    out: list[list[str]] = []
    for raw in rows_iter:
        level_type = raw[idx[_COL_TYPE]]
        if not isinstance(level_type, str) or level_type not in _REAL_TYPES:
            continue
        code = raw[idx[_COL_CODE]]
        district = raw[idx[_COL_DISTRICT]]
        name_en = raw[idx[_COL_NAME_EN]]
        name_ne = raw[idx[_COL_NAME_NE]]
        if code is None or district is None or name_en is None or name_ne is None:
            continue
        if not isinstance(code, (int, float)):
            continue
        out.append([
            f"{int(code):08d}",
            str(name_en).strip(),
            str(name_ne).strip(),
            str(district).strip(),
            _REAL_TYPES[level_type],
        ])
    wb.close()
    return out


_DEFAULT_OUT: Final[str] = "scripts/geo/crosswalk.csv"


def _main() -> None:
    if len(sys.argv) not in (2, 3):
        sys.stderr.write("usage: extract_crosswalk.py <xlsx_path> [out_csv_path]\n")
        sys.exit(2)
    path = Path(sys.argv[1])
    if not path.exists():
        sys.stderr.write(f"file not found: {path}\n")
        sys.exit(2)
    out_path = Path(sys.argv[2]) if len(sys.argv) == 3 else Path(_DEFAULT_OUT)
    rows = _emit(path)
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        writer.writerow(["federal_code", "name_en", "name_ne", "district_en", "local_level_type"])
        writer.writerows(rows)
    sys.stderr.write(f"[extract_crosswalk] wrote {len(rows)} local levels -> {out_path}\n")


if __name__ == "__main__":
    _main()
